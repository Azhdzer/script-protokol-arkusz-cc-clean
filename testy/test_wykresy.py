# -*- coding: utf-8 -*-
"""
Testy zachowania wykresow w arkuszu obserwacji.

Znaleziony blad (istnial od poczatku, niezaleznie od panelu): openpyxl czyta
wykresy tylko czesciowo. Po cyklu wczytaj-zapisz z pliku znikaja WSZYSTKIE
serie danych i odwolania do kolumn — w arkuszu zostaja puste ramki z osiami.

Naprawa podmienia w gotowym .xlsx czesci 'xl/charts/*' na oryginalne
z szablonu. Te testy pilnuja, ze naprawa dziala i ze problem jest realny.
"""

import os
import re
import sys
import unittest
import zipfile

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl

import generuj_obserwacje as G
from wspolne import KORZEN, nowa_piaskownica

SZABLONY = ("szablon_LA_TH_2026 - obserwacje.xlsx",
            "xxx_LA_TH_2026 - obserwacje CC.xlsx")


def statystyki_wykresow(sciezka):
    """(liczba serii, liczba odwolan do komorek) we wszystkich wykresach pliku."""
    serie = odwolania = 0
    with zipfile.ZipFile(sciezka) as z:
        for nazwa in z.namelist():
            if nazwa.startswith("xl/charts/chart") and nazwa.endswith(".xml"):
                xml = z.read(nazwa).decode("utf-8", "replace")
                serie += xml.count("<c:ser>")
                odwolania += len(re.findall(r"<c:f>[^<]+</c:f>", xml))
    return serie, odwolania


class TestUtrataWykresowPrzezOpenpyxl(unittest.TestCase):
    """Najpierw dowodzimy, ze problem naprawde istnieje — inaczej test naprawy
    nic by nie znaczyl."""

    def test_openpyxl_gubi_serie_danych(self):
        for nazwa in SZABLONY:
            zrodlo = os.path.join(KORZEN, nazwa)
            if not os.path.exists(zrodlo):
                continue
            with self.subTest(szablon=nazwa):
                folder = nowa_piaskownica("wykresy_utrata")
                wynik = os.path.join(folder, "po_zapisie.xlsx")
                wb = openpyxl.load_workbook(zrodlo)
                wb.save(wynik)

                serie_przed, ref_przed = statystyki_wykresow(zrodlo)
                serie_po, ref_po = statystyki_wykresow(wynik)
                self.assertGreater(serie_przed, 0, "szablon bez wykresow — test bez sensu")
                self.assertEqual(serie_po, 0,
                                 "openpyxl przestal gubic serie — naprawa moze byc zbedna")
                self.assertEqual(ref_po, 0)


class TestPrzywracanieWykresow(unittest.TestCase):

    def przygotuj(self, nazwa):
        zrodlo = os.path.join(KORZEN, nazwa)
        if not os.path.exists(zrodlo):
            self.skipTest(f"brak szablonu {nazwa}")
        folder = nowa_piaskownica("wykresy_naprawa")
        wynik = os.path.join(folder, "obserwacja.xlsx")
        wb = openpyxl.load_workbook(zrodlo)
        wb.active.cell(row=2, column=1).value = "2026-08-13 14:00:00"
        wb.save(wynik)
        return zrodlo, wynik

    def test_serie_wracaja_do_pliku(self):
        for nazwa in SZABLONY:
            with self.subTest(szablon=nazwa):
                zrodlo, wynik = self.przygotuj(nazwa)
                oczekiwane = statystyki_wykresow(zrodlo)
                G._przywroc_wykresy_z_szablonu(zrodlo, wynik)
                self.assertEqual(statystyki_wykresow(wynik), oczekiwane)

    def test_plik_pozostaje_czytelny(self):
        zrodlo, wynik = self.przygotuj(SZABLONY[0])
        G._przywroc_wykresy_z_szablonu(zrodlo, wynik)
        wb = openpyxl.load_workbook(wynik)
        self.assertEqual(wb.active["A2"].value, "2026-08-13 14:00:00")

    def test_dane_arkusza_nie_sa_ruszane(self):
        """Podmieniamy wylacznie czesci wykresow — komorki zostaja nietkniete."""
        zrodlo, wynik = self.przygotuj(SZABLONY[0])
        wb = openpyxl.load_workbook(wynik)
        ws = wb.active
        for r in range(2, 12):
            ws.cell(row=r, column=2).value = r * 1.5
        wb.save(wynik)

        G._przywroc_wykresy_z_szablonu(zrodlo, wynik)
        wb2 = openpyxl.load_workbook(wynik)
        wartosci = [wb2.active.cell(row=r, column=2).value for r in range(2, 12)]
        self.assertEqual(wartosci, [r * 1.5 for r in range(2, 12)])

    def test_brak_wykresow_w_szablonie_nie_przeszkadza(self):
        folder = nowa_piaskownica("wykresy_brak")
        pusty = os.path.join(folder, "bez_wykresow.xlsx")
        wynik = os.path.join(folder, "wynik.xlsx")
        wb = openpyxl.Workbook()
        wb.save(pusty)
        wb.save(wynik)
        self.assertEqual(G._przywroc_wykresy_z_szablonu(pusty, wynik), 0)

    def test_bledna_sciezka_nie_wywala_generowania(self):
        """Wykresy to dodatek — ich brak nie moze przerwac tworzenia arkusza."""
        folder = nowa_piaskownica("wykresy_blad")
        wynik = os.path.join(folder, "wynik.xlsx")
        openpyxl.Workbook().save(wynik)
        self.assertEqual(
            G._przywroc_wykresy_z_szablonu(os.path.join(folder, "nie_ma.xlsx"), wynik), 0)
        self.assertTrue(os.path.exists(wynik))


if __name__ == "__main__":
    unittest.main(verbosity=2)
