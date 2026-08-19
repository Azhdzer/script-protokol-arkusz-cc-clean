# -*- coding: utf-8 -*-
"""
Testy pomijania pustych kolumn w arkuszu obserwacji.

Multimetr CC-04 zapisuje komplet kanalow Ch101..Ch108 razem z ich temperaturami
i rozrzutami. Gdy wzorcuje sie na 2 czujnikach zamiast 4, kolumny pozostalych
kanalow byly wpisywane do arkusza puste — z samym naglowkiem.

Zasada: sprawdzamy kazda kolumne od 2. wiersza w dol; jesli nie ma tam ani
jednej wartosci, kolumna w ogole nie trafia do arkusza. Sprawdzamy tylko obszar
objety naglowkami — dalej nie zagladamy.

Bezpieczenstwo: analiza i protokol pracuja na wierszach sparsowanych z pliku TXT
(`rows`), a NIE na kolumnach arkusza, wiec przesuniecie kolumn ich nie dotyczy.
Jedyne, co zalezy od pozycji, to odwolania wykresow — i te sa przeliczane.
"""

import os
import sys
import unittest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl.utils import get_column_letter

import generuj_obserwacje as G

# Uklad CC-04: 5 kolumn ramki, potem kanaly i ich pochodne (33 kolumny).
N_KOL_CC04 = len(G.CC04_KOLUMNY)


def wiersze(uzyte_indeksy, ile=20):
    """Wiersze, w ktorych dane sa tylko w ramce (0-4) i w podanych kolumnach."""
    dane = []
    for _ in range(ile):
        r = ["2026-08-13 14:00:00", "25.1", "30.5", "25.9", "48"] + [""] * (N_KOL_CC04 - 5)
        for i in uzyte_indeksy:
            r[i] = "109.68"
        dane.append(r)
    return dane


class TestWyborKolumn(unittest.TestCase):

    def litery(self, kolumny):
        return [get_column_letter(i + 1) for i in kolumny]

    def test_ramka_zostaje_zawsze(self):
        """Czas i nastawy zostaja, nawet gdy RHzadana jest pusta (pomiar temp-only)."""
        rows = [["2026-08-13 14:00:00", "25.1", "", "25.9", "", "109.7"]
                + [""] * (N_KOL_CC04 - 6)]
        kolumny = G._kolumny_z_danymi(rows, N_KOL_CC04)
        self.assertEqual(kolumny[:5], [0, 1, 2, 3, 4])

    def test_puste_kanaly_wypadaja(self):
        """Dane w Ch105 (J) i Ch107 (L) — reszta kanalow ma zniknac."""
        kolumny = G._kolumny_z_danymi(wiersze([9, 11]), N_KOL_CC04)
        self.assertEqual(self.litery(kolumny), ["A", "B", "C", "D", "E", "J", "L"])

    def test_kolumna_z_jedna_wartoscia_zostaje(self):
        """Wystarczy jeden odczyt w calej kolumnie, zeby ja zachowac."""
        rows = wiersze([9])
        rows[7][11] = "109.9"          # pojedyncza wartosc w Ch107
        kolumny = G._kolumny_z_danymi(rows, N_KOL_CC04)
        self.assertIn(11, kolumny)

    def test_komplet_kanalow_nic_nie_gubi(self):
        kolumny = G._kolumny_z_danymi(wiersze(range(5, N_KOL_CC04)), N_KOL_CC04)
        self.assertEqual(len(kolumny), N_KOL_CC04)

    def test_nie_zaglada_poza_naglowki(self):
        """Kolumny poza ukladem (dluzszy wiersz) nie sa brane pod uwage."""
        rows = wiersze([9])
        for r in rows:
            r.extend(["smiec", "smiec"])
        kolumny = G._kolumny_z_danymi(rows, N_KOL_CC04)
        self.assertTrue(all(i < N_KOL_CC04 for i in kolumny))

    def test_biale_znaki_to_nie_dane(self):
        rows = wiersze([9])
        for r in rows:
            r[10] = "   "
        self.assertNotIn(10, G._kolumny_z_danymi(rows, N_KOL_CC04))

    def test_brak_wierszy_zostawia_sama_ramke(self):
        self.assertEqual(G._kolumny_z_danymi([], N_KOL_CC04), [0, 1, 2, 3, 4])

    def test_kolejnosc_kolumn_zachowana(self):
        kolumny = G._kolumny_z_danymi(wiersze([20, 9, 11]), N_KOL_CC04)
        self.assertEqual(kolumny, sorted(kolumny))


class TestMapaKolumn(unittest.TestCase):
    """Mapa 'stary indeks -> nowy' jest podstawa przeliczenia wykresow."""

    def test_mapa_przesuwa_w_lewo(self):
        kolumny = G._kolumny_z_danymi(wiersze([9, 11]), N_KOL_CC04)
        mapa = G._raport_pominietych_kolumn(kolumny, G.CC04_KOLUMNY, N_KOL_CC04)
        self.assertEqual(mapa[9], 5)       # J -> F
        self.assertEqual(mapa[11], 6)      # L -> G

    def test_pominiete_nie_maja_wpisu(self):
        kolumny = G._kolumny_z_danymi(wiersze([9, 11]), N_KOL_CC04)
        mapa = G._raport_pominietych_kolumn(kolumny, G.CC04_KOLUMNY, N_KOL_CC04)
        self.assertNotIn(10, mapa)         # K = Ch106, nieuzyty

    def test_bez_pominiec_mapa_jest_tozsamoscia(self):
        kolumny = list(range(N_KOL_CC04))
        mapa = G._raport_pominietych_kolumn(kolumny, G.CC04_KOLUMNY, N_KOL_CC04)
        self.assertEqual(mapa, {i: i for i in range(N_KOL_CC04)})


class TestPrzeliczenieOdwolanWykresu(unittest.TestCase):
    """
    Wykresy szablonu celuja w konkretne litery ($J = Ch105). Po przesunieciu
    kolumn musza wskazywac nowe pozycje, inaczej pokazywalyby sasiednie dane.
    """

    def setUp(self):
        # Realny uklad: kanaly Ch105 (J) i Ch107 (L) plus kolumny wyliczane
        # tdp/temperatura/%RH (N, O, P), ktore sa zawsze wypelnione.
        kolumny = G._kolumny_z_danymi(wiersze([9, 11, 13, 14, 15]), N_KOL_CC04)
        self.mapa = G._raport_pominietych_kolumn(kolumny, G.CC04_KOLUMNY, N_KOL_CC04)

    def test_zakres_jest_przeliczony(self):
        nowy, _z = G._przelicz_odwolania_wykresu(
            "obserwacje!$J$2:$J$25066", self.mapa)
        self.assertEqual(nowy, "obserwacje!$F$2:$F$25066")

    def test_naglowek_serii_jest_przeliczony(self):
        nowy, _z = G._przelicz_odwolania_wykresu("obserwacje!$N$1", self.mapa)
        self.assertEqual(nowy, "obserwacje!$H$1")

    def test_kolumny_ramki_zostaja_na_miejscu(self):
        nowy, _z = G._przelicz_odwolania_wykresu(
            "obserwacje!$A$2:$A$99 obserwacje!$B$1", self.mapa)
        self.assertEqual(nowy, "obserwacje!$A$2:$A$99 obserwacje!$B$1")

    def test_zamiany_nie_nakladaja_sie(self):
        """J->F i N->H w jednym przebiegu; F nie moze zostac zamienione powtornie."""
        nowy, _z = G._przelicz_odwolania_wykresu(
            "$J$1 $N$1 $L$1", self.mapa)
        self.assertEqual(nowy, "$F$1 $H$1 $G$1")

    def test_usunieta_kolumna_jest_zgloszona(self):
        nowy, zgubione = G._przelicz_odwolania_wykresu("obserwacje!$K$1", self.mapa)
        self.assertEqual(zgubione, ["K"])
        self.assertEqual(nowy, "obserwacje!$K$1")   # zostawiamy bez zmian

    def test_pusta_mapa_nic_nie_zmienia(self):
        tresc = "obserwacje!$J$2:$J$99"
        nowy, zgubione = G._przelicz_odwolania_wykresu(tresc, {})
        self.assertEqual((nowy, zgubione), (tresc, []))



class TestSzerokoscDanych(unittest.TestCase):
    """
    Kolorowanie i znacznik numeru punktu musza trzymac sie faktycznej szerokosci
    arkusza. Zgloszenie: po pominieciu pustych kanalow cyfry do przeskakiwania
    miedzy punktami wyladowaly kilkanascie kolumn na prawo od danych.
    """

    def arkusz(self, ile_naglowkow, szerokosc_szablonu=33):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        for i in range(1, ile_naglowkow + 1):
            ws.cell(row=1, column=i).value = f"H{i}"
        # Komorki po prawej istnieja, ale sa puste — tak wyglada arkusz po
        # wyczyszczeniu pozostalosci szablonu.
        for i in range(ile_naglowkow + 1, szerokosc_szablonu + 1):
            ws.cell(row=1, column=i).value = None
        return ws

    def test_liczy_po_naglowkach_a_nie_po_max_column(self):
        ws = self.arkusz(15)
        self.assertEqual(ws.max_column, 33)          # szablon nadal „szeroki"
        self.assertEqual(G._szerokosc_danych(ws, 33), 15)

    def test_komplet_kolumn(self):
        ws = self.arkusz(33)
        self.assertEqual(G._szerokosc_danych(ws, 33), 33)

    def test_pusty_arkusz_wraca_do_domyslnej(self):
        ws = self.arkusz(0)
        self.assertEqual(G._szerokosc_danych(ws, 12), 12)

    def test_zatrzymuje_sie_na_pierwszej_dziurze(self):
        """Naglowki sa ciagle — pierwsza pusta komorka konczy obszar danych."""
        ws = self.arkusz(5)
        ws.cell(row=1, column=9).value = "osierocony"
        self.assertEqual(G._szerokosc_danych(ws, 33), 5)

class TestFormatyPoPrzesunieciu(unittest.TestCase):
    """
    Format liczbowy musi jechac RAZEM z danymi.

    Zgloszenie: po pominieciu pustych kanalow 'TPunktuRosy' 16,07 pokazywalo sie
    jako 16,0700. Kazda kolumna szablonu ma wlasny format ('Wskazania multimetru'
    -> '0.0000', reszta -> 'General'); dane wjechaly w kolumne o cudzym formacie
    i dostaly obce miejsca po przecinku.
    """

    def arkusz(self, formaty):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        for i, fmt in enumerate(formaty, start=1):
            ws.cell(row=1, column=i).value = f"H{i}"
            for w in range(2, 6):
                ws.cell(row=w, column=i).value = 1.0
                ws.cell(row=w, column=i).number_format = fmt
        return ws

    def test_format_jedzie_z_danymi(self):
        # kolumna 3 ('General') przesuwa sie na pozycje 1 ('0.0000')
        ws = self.arkusz(["0.0000", "0.0000", "General"])
        G._przenies_formaty_kolumn(ws, [2], ["0.0000", "0.0000", "General"], 4)
        self.assertEqual(ws.cell(row=2, column=1).number_format, "General")

    def test_wszystkie_wiersze_dostaja_format(self):
        ws = self.arkusz(["0.0000", "General"])
        G._przenies_formaty_kolumn(ws, [1], ["0.0000", "General"], 4)
        for w in range(2, 6):
            with self.subTest(wiersz=w):
                self.assertEqual(ws.cell(row=w, column=1).number_format, "General")

    def test_bez_przesuniecia_nic_sie_nie_zmienia(self):
        formaty = ["0.0000", "General", "0.000"]
        ws = self.arkusz(formaty)
        zmienione = G._przenies_formaty_kolumn(ws, [0, 1, 2], formaty, 4)
        self.assertEqual(zmienione, 0)
        self.assertEqual(ws.cell(row=2, column=1).number_format, "0.0000")

    def test_ten_sam_format_nie_wymaga_pracy(self):
        formaty = ["General", "General", "General"]
        ws = self.arkusz(formaty)
        self.assertEqual(G._przenies_formaty_kolumn(ws, [2], formaty, 4), 0)

    def test_brak_wierszy_nic_nie_robi(self):
        ws = self.arkusz(["0.0000", "General"])
        self.assertEqual(G._przenies_formaty_kolumn(ws, [1], ["0.0000", "General"], 0), 0)

    def test_odczyt_formatow_szablonu(self):
        ws = self.arkusz(["yyyy/mm/dd", "0.0000", "General"])
        self.assertEqual(G._formaty_kolumn_szablonu(ws, 3),
                         ["yyyy/mm/dd", "0.0000", "General"])


if __name__ == "__main__":
    unittest.main(verbosity=2)
