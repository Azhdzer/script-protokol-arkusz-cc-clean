# -*- coding: utf-8 -*-
"""
Test calego obiegu od poczatku do konca, na PRAWDZIWYCH danych projektu
(pomiar 188: dwa pliki TXT multimetru, PZ, Zestawienie, logi czterech
termohigrometrow Testo 174T).

Trzy kroki lecza po kolei w JEDNEJ piaskownicy — dokladnie tak, jak robi to
przycisk "Uruchom caly obieg":

    1. analizuj_excele    logi DUT            -> wyniki/<serial>_wynik.xlsx
    2. generuj_obserwacje TXT + PZ + wyniki   -> arkusz obserwacji + protokol
    3. generuj_arkusze    protokol            -> kopie Excel + swiadectwa Word

Krok 3 uruchamia Excel przez COM i trwa okolo dwoch minut. Aby go pominac:
    set CC_TESTY_SZYBKIE=1

Sprawdzamy nie tylko "czy sie nie wywalilo", ale czy USTAWIENIA Z PANELU widac
w gotowych dokumentach — podpis w protokole i numer swiadectwa w nazwie pliku Word.
"""

import glob
import os
import re
import sys
import unittest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl

from wspolne import (KORZEN, dostepne, nowa_piaskownica, uruchom_worker,
                     utworz_logi_probne, PLIKI_OBSERWACJA, FOLDERY_OBSERWACJA,
                     PLIKI_ARKUSZE, SZABLON_ARKUSZA, PROTOKOL_CC, OBSERWACJA_CC,
                     PROTOKOL_GOTOWY)

SZYBKIE = os.environ.get("CC_TESTY_SZYBKIE", "").strip() in ("1", "true", "tak")

# Ustawienia podstawiane z "panelu" — dobrane tak, by roznily sie od domyslnych,
# wiec ich obecnosc w wyniku dowodzi, ze panel naprawde steruje obiegiem.
PODPIS_TESTOWY = "Testowy Podpisujacy"
NR_SW_TESTOWY = "5000"

TXT_POMIARU = ["2026-08-06 12.10_188.txt", "2026-08-10 13.57_188.txt"]


class TestPelnyObieg(unittest.TestCase):
    """Jedna piaskownica, trzy kroki po kolei, potem sprawdzanie artefaktow."""

    folder = None
    wyniki = {}

    @classmethod
    def setUpClass(cls):
        brakujace = [p for p in (list(PLIKI_OBSERWACJA) + list(FOLDERY_OBSERWACJA))
                     if not os.path.exists(os.path.join(KORZEN, p))]
        if brakujace:
            raise unittest.SkipTest(f"brak danych wejsciowych: {brakujace}")

        pliki = dostepne(list(PLIKI_OBSERWACJA) + list(PLIKI_ARKUSZE))
        cls.folder = nowa_piaskownica("obieg", pliki, ["PZ", "excel_do_analizy"])

        # ── Krok 1 ────────────────────────────────────────────────────────
        cls.wyniki["analiza"] = uruchom_worker("analiza", cls.folder)

        # ── Krok 2 ────────────────────────────────────────────────────────
        cls.wyniki["obs"] = uruchom_worker("obserwacje", cls.folder, {
            "OBS_TXT_FILES": ";".join(TXT_POMIARU),
            "OBS_PODPIS": PODPIS_TESTOWY,
        })

        # ── Krok 3 ────────────────────────────────────────────────────────
        if SZYBKIE:
            cls.wyniki["ark"] = None
            return
        cls.wyniki["ark"] = uruchom_worker("arkusze", cls.folder, {
            "CC_PROTOKOL": PROTOKOL_CC,
            "CC_SZABLON": SZABLON_ARKUSZA,
            "GEN_NR_SW": NR_SW_TESTOWY,
        }, limit_s=1200)

    # ── pomocnicze ────────────────────────────────────────────────────────
    def sciezka(self, *czesci):
        return os.path.join(self.folder, *czesci)

    def kod_i_log(self, krok):
        kod, log = self.wyniki[krok]
        return kod, log

    # ══════════════════════════════════════════════════════════════════════
    # KROK 1 — analiza logow
    # ══════════════════════════════════════════════════════════════════════
    def test_1_analiza_konczy_sie_sukcesem(self):
        kod, log = self.kod_i_log("analiza")
        self.assertEqual(kod, 0, f"analizuj_excele zwrocil {kod}:\n{log[-3000:]}")

    def test_1_analiza_tworzy_plik_dla_kazdego_logu(self):
        zrodla = glob.glob(self.sciezka("excel_do_analizy", "*"))
        wyniki = glob.glob(self.sciezka("wyniki", "*_wynik.xlsx"))
        self.assertEqual(len(wyniki), len(zrodla))

    def test_1_wyniki_maja_znormalizowane_kolumny(self):
        for plik in glob.glob(self.sciezka("wyniki", "*_wynik.xlsx")):
            with self.subTest(plik=os.path.basename(plik)):
                wb = openpyxl.load_workbook(plik, read_only=True)
                ws = wb.active
                naglowki = [str(c.value or "") for c in next(ws.iter_rows(max_row=1))]
                wb.close()
                self.assertIn("Czas", naglowki[0])
                self.assertTrue(any("Temperatura" in n for n in naglowki))

    def test_1_wyniki_zawieraja_dane(self):
        for plik in glob.glob(self.sciezka("wyniki", "*_wynik.xlsx")):
            with self.subTest(plik=os.path.basename(plik)):
                wb = openpyxl.load_workbook(plik, read_only=True)
                liczba = wb.active.max_row
                wb.close()
                self.assertGreater(liczba, 100, "plik wynikow jest podejrzanie pusty")

    def test_1_powstaje_zestawienie_zbiorcze(self):
        self.assertTrue(os.path.exists(self.sciezka("wyniki", "zestawienie_pomiarow.xlsx")))

    # ══════════════════════════════════════════════════════════════════════
    # KROK 2 — obserwacja i protokol
    # ══════════════════════════════════════════════════════════════════════
    def test_2_obserwacja_konczy_sie_sukcesem(self):
        kod, log = self.kod_i_log("obs")
        self.assertEqual(kod, 0, f"generuj_obserwacje zwrocil {kod}:\n{log[-3000:]}")

    def test_2_powstaje_arkusz_obserwacji(self):
        self.assertTrue(os.path.exists(self.sciezka(OBSERWACJA_CC)),
                        f"brak {OBSERWACJA_CC}")

    def test_2_powstaje_protokol(self):
        self.assertTrue(os.path.exists(self.sciezka(PROTOKOL_CC)),
                        f"brak {PROTOKOL_CC}")

    def test_2_protokol_ma_wymagane_arkusze(self):
        wb = openpyxl.load_workbook(self.sciezka(PROTOKOL_CC), read_only=True)
        nazwy = wb.sheetnames
        wb.close()
        self.assertIn("Strona 2", nazwy)
        self.assertIn("Strona 3", nazwy)

    def test_2_strona2_wypelniona_przyrzadami_z_pz(self):
        """Tabela przyrzadow musi dostac cztery termohigrometry z PZ."""
        wb = openpyxl.load_workbook(self.sciezka(PROTOKOL_CC), data_only=True)
        ws = wb["Strona 2"]
        seriale = [ws.cell(w, 5).value for w in range(11, 15)]   # kolumna E
        wytworcy = [ws.cell(w, 2).value for w in range(11, 15)]  # kolumna B
        wb.close()
        self.assertEqual([str(s) for s in seriale],
                         ["37025098", "37025105", "37025156", "37025720"])
        self.assertTrue(all(w == "Testo" for w in wytworcy), wytworcy)

    def test_2_strona2_ma_rozdzielczosc_z_zestawienia(self):
        wb = openpyxl.load_workbook(self.sciezka(PROTOKOL_CC), data_only=True)
        ws = wb["Strona 2"]
        rozdzielczosci = [ws.cell(w, 11).value for w in range(11, 15)]  # kolumna K
        wb.close()
        self.assertTrue(all(r == 0.1 for r in rozdzielczosci), rozdzielczosci)

    def test_2_podpis_z_panelu_trafia_do_protokolu(self):
        """Dowod, ze ustawienie panelu widac w gotowym dokumencie."""
        wb = openpyxl.load_workbook(self.sciezka(PROTOKOL_CC), data_only=True)
        ws = wb["Strona 2"]
        znalezione = any(
            isinstance(k.value, str) and PODPIS_TESTOWY in k.value
            for wiersz in ws.iter_rows(min_row=1, max_row=40) for k in wiersz)
        wb.close()
        self.assertTrue(znalezione,
                        f"'{PODPIS_TESTOWY}' (OBS_PODPIS) nie trafil na Strone 2")

    def test_2_dane_srodowiskowe_z_kroku_1_sa_wpisane(self):
        """Strona 3 musi dostac odczyty z wyniki/ — to spina krok 1 z krokiem 2."""
        kod, log = self.kod_i_log("obs")
        self.assertIn("[WYNIKI]", log)
        self.assertRegex(log, r"Przyrzad 1 \(Q/R\) <- '3702\d+")

    def test_2_sklejenie_dwoch_plikow_txt(self):
        _kod, log = self.kod_i_log("obs")
        for nazwa in TXT_POMIARU:
            self.assertIn(nazwa, log)

    def test_2_punkty_spoza_zamowienia_sa_wyszarzane(self):
        """Wybor punktow wg PZ — jeden wsad komory obsluguje kilka zlecen."""
        _kod, log = self.kod_i_log("obs")
        self.assertIn("[PZ]", log)
        self.assertIn("spoza zamowienia", log)

    # ══════════════════════════════════════════════════════════════════════
    # KROK 3 — arkusze i swiadectwa
    # ══════════════════════════════════════════════════════════════════════
    def pomin_gdy_szybkie(self):
        if SZYBKIE:
            self.skipTest("CC_TESTY_SZYBKIE=1 — krok 3 (Excel COM) pominiety")

    def kopie_excel(self):
        return sorted(glob.glob(self.sciezka("188_LA_TH_2026 - ILAJ*.xlsx")))

    def swiadectwa(self):
        return sorted(glob.glob(self.sciezka("*_188_LA_TH_2026.docx")))

    def test_3_arkusze_koncza_sie_sukcesem(self):
        self.pomin_gdy_szybkie()
        kod, log = self.kod_i_log("ark")
        self.assertEqual(kod, 0, f"generuj_arkusze zwrocil {kod}:\n{log[-4000:]}")

    def test_3_powstaje_kopia_dla_kazdego_przyrzadu(self):
        self.pomin_gdy_szybkie()
        self.assertEqual(len(self.kopie_excel()), 4, self.kopie_excel())

    def test_3_nazwy_kopii_zawieraja_numery_fabryczne(self):
        self.pomin_gdy_szybkie()
        nazwy = " ".join(os.path.basename(p) for p in self.kopie_excel())
        for serial in ("37025098", "37025105", "37025156", "37025720"):
            with self.subTest(serial=serial):
                self.assertIn(serial, nazwy)

    def test_3_kopie_maja_arkusz_wyniki(self):
        self.pomin_gdy_szybkie()
        for plik in self.kopie_excel():
            with self.subTest(plik=os.path.basename(plik)):
                wb = openpyxl.load_workbook(plik, read_only=True)
                nazwy = wb.sheetnames
                wb.close()
                self.assertIn("Wyniki", nazwy)
                self.assertGreater(len(nazwy), 1, "kopia bez zakladek punktow")

    def test_3_powstaje_swiadectwo_dla_kazdej_kopii(self):
        self.pomin_gdy_szybkie()
        self.assertEqual(len(self.swiadectwa()), 4, self.swiadectwa())

    def test_3_numeracja_swiadectw_z_panelu(self):
        """GEN_NR_SW z panelu musi wyznaczyc numery kolejnych swiadectw."""
        self.pomin_gdy_szybkie()
        numery = sorted(
            int(re.match(r"(\d+)_", os.path.basename(p)).group(1))
            for p in self.swiadectwa())
        poczatek = int(NR_SW_TESTOWY)
        self.assertEqual(numery, [poczatek, poczatek + 1, poczatek + 2, poczatek + 3])

    def test_3_swiadectwa_nie_sa_puste(self):
        self.pomin_gdy_szybkie()
        for plik in self.swiadectwa():
            with self.subTest(plik=os.path.basename(plik)):
                self.assertGreater(os.path.getsize(plik), 50_000)

    def test_3_linki_zewnetrzne_przywrocone_na_serwer(self):
        """Kopia musi wyjsc z linkami na \\\\plum4, nie na piaskownice."""
        self.pomin_gdy_szybkie()
        _kod, log = self.kod_i_log("ark")
        self.assertIn(r"\\plum4\LabPomiarowe\Wzory.xls", log)
        self.assertIn("Przywrocono linki zewnetrzne", log)

    def test_3_nie_zglosil_bledow_krytycznych(self):
        self.pomin_gdy_szybkie()
        _kod, log = self.kod_i_log("ark")
        krytyczne = [w for w in log.splitlines() if "!!! BLAD" in w]
        self.assertFalse(krytyczne, "\n".join(krytyczne))


class TestWyborPlikowDoAnalizy(unittest.TestCase):
    """
    Zaznaczenie plikow w panelu musi REALNIE zawezic wsad kroku 1.

    Wczesniej lista w panelu miala pola wyboru, ale skrypt i tak przerabial
    caly folder — klikanie nic nie dawalo. To test, ktory tego pilnuje.
    """

    @classmethod
    def setUpClass(cls):
        # Logi generujemy sami — test ma dzialac takze wtedy, gdy folder
        # zlecenia jest juz posprzatany.
        cls.folder = nowa_piaskownica("wybor_analizy")
        cls.wszystkie = utworz_logi_probne(
            os.path.join(cls.folder, "excel_do_analizy"))
        cls.wybrany = cls.wszystkie[0]
        cls.kod, cls.log = uruchom_worker("analiza", cls.folder,
                                          {"ANL_PLIKI": cls.wybrany})

    def wyniki(self):
        return sorted(os.path.basename(p) for p in
                      glob.glob(os.path.join(self.folder, "wyniki", "*_wynik.xlsx")))

    def test_konczy_sie_sukcesem(self):
        self.assertEqual(self.kod, 0, self.log[-2000:])

    def test_powstaje_wynik_tylko_dla_zaznaczonego(self):
        self.assertEqual(len(self.wyniki()), 1, self.wyniki())

    def test_wynik_dotyczy_wlasciwego_pliku(self):
        rdzen = os.path.splitext(self.wybrany)[0]
        self.assertEqual(self.wyniki(), [f"{rdzen}_wynik.xlsx"])

    def test_log_wymienia_pominiete_pliki(self):
        self.assertIn("Pominięto", self.log)
        for nazwa in self.wszystkie[1:]:
            with self.subTest(plik=nazwa):
                self.assertIn(nazwa, self.log)

    def test_bez_zaznaczenia_przerabiane_sa_wszystkie(self):
        """Puste zaznaczenie = zachowanie sprzed panelu."""
        folder = nowa_piaskownica("wybor_analizy_puste")
        utworz_logi_probne(os.path.join(folder, "excel_do_analizy"))
        kod, log = uruchom_worker("analiza", folder, {"ANL_PLIKI": ""})
        self.assertEqual(kod, 0, log[-2000:])
        powstale = glob.glob(os.path.join(folder, "wyniki", "*_wynik.xlsx"))
        self.assertEqual(len(powstale), len(self.wszystkie))


class TestWyszarzonePrzyrzady(unittest.TestCase):
    """
    Wyszarzenie pomiarow przyrzadu na Stronie 3 ma go CALKOWICIE wykluczyc.

    Realne zgloszenie: uzytkownik wyszarzyl trzy przyrzady, zeby dostac arkusz
    i swiadectwo dla jednego. Dostal 4 kopie (trzy puste, po 12 kB) i 4
    swiadectwa. Lista kopii idzie ze Strony 2, wiec kolory na Stronie 3 same z
    siebie nie usuwaly przyrzadu z obiegu.
    """

    SZARY = (191, 191, 191)     # #BFBFBF — kolor "pomijane" z konfiguracji

    @classmethod
    def setUpClass(cls):
        if SZYBKIE:
            raise unittest.SkipTest("CC_TESTY_SZYBKIE=1 — test wymaga Excela")
        if not os.path.exists(os.path.join(KORZEN, PROTOKOL_GOTOWY)):
            raise unittest.SkipTest(f"brak gotowego protokolu: {PROTOKOL_GOTOWY}")

        # Krok 3 startuje z gotowego protokolu, wiec test nie zalezy od plikow
        # TXT poprzedniego zlecenia (te znikaja z folderu po zakonczeniu pracy).
        pliki = dostepne([PROTOKOL_GOTOWY] + list(PLIKI_ARKUSZE))
        cls.folder = nowa_piaskownica("wyszarzone", pliki, ["PZ"])

        cls._wyszarz_przyrzady(os.path.join(cls.folder, PROTOKOL_GOTOWY),
                               do_przyrzadu=3)

        cls.kod, cls.log = uruchom_worker("arkusze", cls.folder, {
            "CC_PROTOKOL": PROTOKOL_GOTOWY,
            "CC_SZABLON": SZABLON_ARKUSZA,
            "GEN_NR_SW": NR_SW_TESTOWY,
        }, limit_s=1200)

    @classmethod
    def _wyszarz_przyrzady(cls, sciezka, do_przyrzadu):
        """
        Maluje na szaro kolumny pomiarowe przyrzadow 1..do_przyrzadu.

        Uzywamy Excela (a nie openpyxl), bo zapis openpyxl potrafi zgubic czesc
        formatowania protokolu — test przestalby odzwierciedlac to, co robi
        uzytkownik recznie.
        """
        import xlwings as xw
        kol_od = 17                          # Q — pierwszy przyrzad
        kol_do = 17 + 2 * do_przyrzadu - 1   # ostatnia kolumna wyszarzanego
        app = xw.App(visible=False, add_book=False)
        try:
            app.api.DisplayAlerts = False
            wb = app.books.open(sciezka, update_links=False)
            ws = wb.sheets["Strona 3"]
            ws.range((20, kol_od), (54, kol_do)).color = cls.SZARY
            wb.save()
            wb.close()
        finally:
            try:
                app.quit()
            except Exception:
                pass

    def kopie_excel(self):
        return sorted(glob.glob(os.path.join(
            self.folder, "188_LA_TH_2026 - ILAJ*.xlsx")))

    def swiadectwa(self):
        return sorted(glob.glob(os.path.join(
            self.folder, "*_188_LA_TH_2026.docx")))

    def test_konczy_sie_sukcesem(self):
        self.assertEqual(self.kod, 0, self.log[-4000:])

    def test_powstaje_dokladnie_jedna_kopia(self):
        self.assertEqual(len(self.kopie_excel()), 1,
                         [os.path.basename(p) for p in self.kopie_excel()])

    def test_powstaje_dokladnie_jedno_swiadectwo(self):
        self.assertEqual(len(self.swiadectwa()), 1,
                         [os.path.basename(p) for p in self.swiadectwa()])

    def test_zostal_wlasciwy_przyrzad(self):
        self.assertIn("37025720", os.path.basename(self.kopie_excel()[0]))

    def test_nazwa_zachowuje_numer_z_protokolu(self):
        """Czwarty przyrzad zostaje czwartym — nazwa musi zgadzac sie z Strona 2."""
        self.assertIn(" - 4 - ", os.path.basename(self.kopie_excel()[0]))

    def test_log_wymienia_pominiete_przyrzady(self):
        self.assertIn("Pomijam 3", self.log)

    def test_kopia_nie_jest_pusta(self):
        """Pozostala kopia ma miec zakladki punktow, nie sam arkusz Wyniki."""
        wb = openpyxl.load_workbook(self.kopie_excel()[0], read_only=True)
        nazwy = wb.sheetnames
        wb.close()
        self.assertIn("Wyniki", nazwy)
        self.assertGreater(len(nazwy), 1, nazwy)

    def test_swiadectwo_ma_numer_poczatkowy(self):
        """Numeracja startuje od GEN_NR_SW, a nie od numeru pozycji przyrzadu."""
        nazwa = os.path.basename(self.swiadectwa()[0])
        self.assertTrue(nazwa.startswith(NR_SW_TESTOWY), nazwa)


if __name__ == "__main__":
    unittest.main(verbosity=2)
