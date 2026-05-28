# -*- coding: utf-8 -*-
import openpyxl, os

FOLDER = os.path.dirname(os.path.abspath(__file__))
PROTOKOL = os.path.join(FOLDER, "133; 148_LA_TH_2026 - protokół CC.xlsx")

wb = openpyxl.load_workbook(PROTOKOL, read_only=True, data_only=True)
ws = wb["Strona 3"]

print("=== Strona 3 — blok 1 (wiersze 20-24) ===")
for row in range(20, 25):
    vals = {col: ws.cell(row=row, column=col).value
            for col in [1, 2, 3, 12, 13, 17, 18, 19, 20]}
    print(f"  row {row}: A={vals[1]!r}  B={vals[2]!r}  C={vals[3]!r} | "
          f"L={vals[12]!r}  M={vals[13]!r} | Q={vals[17]!r}  R={vals[18]!r}  S={vals[19]!r}  T={vals[20]!r}")

print("\n=== Strona 3 — blok 2 (wiersze 25-29) ===")
for row in range(25, 30):
    vals = {col: ws.cell(row=row, column=col).value
            for col in [1, 12, 13, 17, 18]}
    print(f"  row {row}: A={vals[1]!r}  L={vals[12]!r}  M={vals[13]!r}  Q={vals[17]!r}  R={vals[18]!r}")

wb.close()
print("\nDone.")
