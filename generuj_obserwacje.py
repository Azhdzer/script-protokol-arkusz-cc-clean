# -*- coding: utf-8 -*-
"""
generuj_obserwacje.py

Tworzy arkusz obserwacji CC na podstawie pliku TXT z danymi pomiarowymi.

1. Nazwa pliku TXT  → numer pomiaru (po dacie i czasie) → nazwa kopii szablonu
2. Kopia szablonu: xxx → numer pomiaru
3. Parsowanie pliku TXT:
       A8  – czujnik wzorcowy (tekst po ':')
       A15+ – dane pomiarowe (sep: Tab / Średnik / Przecinek), kolumny A:L
4. Dane A15:L(ostatni) → kopia szablonu A2:L(ostatni)
5. J1  – podmiana wzorca "Pt100-XX" na czujnik z pliku
6. N92 – aktualna data
7. O92 – podpis
8. Analiza stabilności:
       - szuka okna po 1.5h stabilnym B+C (rozgrzewka)
       - filtruje wiersze wg kryteriów K (punkt rosy) i L (temperatura)
       - wybiera 5-minutowe okno z najmniejszym rozrzutem K
       - koloruje okno na zielono (#E2EFDA)
       - wybiera 5 wierszy reprezentacyjnych (1/min, blisko granicy minutowej)
         i oznacza je ciemniejszą zielenią (#A9D08E) + pogrubienie
"""

import os
import re
import math
import bisect
import shutil
import zipfile
import datetime
import statistics
from copy import copy as _copy_obj
import openpyxl
from openpyxl.styles import PatternFill, Font, Border
from openpyxl.utils import get_column_letter, column_index_from_string

import pz_dane   # wspolny modul: dane przyrzadow z PZ + Zestawienie
import cc_config as C   # rejestr ustawien + odczyt zmiennych srodowiskowych z panelu

# =============================================================================
# KONFIGURACJA
#
# Wszystkie wartosci ponizej ustawia sie w PANELU (app_gui.py) — to, co widzisz
# w kodzie, to tylko wartosci DOMYSLNE uzywane przy recznym uruchomieniu skryptu.
# Panel podaje je przez zmienne srodowiskowe (nazwy w nawiasach przy C.*).
# =============================================================================

FOLDER           = os.environ.get("CC_FOLDER") or \
                   r"C:\Users\artisom.azhdzer\Desktop\Script protokoł - arkusz CC"
TXT_FILENAME     = ""
# Wiele plikow TXT jednego (przerwanego) pomiaru — zostana rozparsowane i
# sklejone w jeden ciag. Pusta lista => uzywany jest pojedynczy TXT_FILENAME.
# Panel GUI podaje liste przez zmienna OBS_TXT_FILES (rozdzielona ';').
TXT_FILENAMES    = []
TEMPLATE         = C.tekst("OBS_TEMPLATE",   "xxx_LA_TH_2026 - obserwacje CC.xlsx")
CC04_TEMPLATE    = C.tekst("OBS_CC04_TEMPLATE", "szablon_LA_TH_2026 - obserwacje.xlsx")
PROTOKOL_CC_TEMPLATE   = C.tekst("OBS_PROT_CC",   "xxx_LA_TH_2026 - protokół CC.xlsx")
PROTOKOL_CC04_TEMPLATE = C.tekst("OBS_PROT_CC04", "xxx_LA_TH_2026 - protokół CC-04.xlsx")
PODPIS           = C.tekst("OBS_PODPIS", "Artsiom Azhdzer")      # 'Pomiary wykonal(a)'
PODPIS_SPRAWDZIL = C.tekst("OBS_PODPIS_SPR", "Marek Szpakowski") # 'Protokol sprawdzil(a)'
# rozgrzewka: gdy odczyty nie wejda w widelki, okno od tego czasu od poczatku punktu
STABILIZACJA_MIN  = C.minuty("OBS_STAB_MIN", datetime.timedelta(hours=2))

# Dobor okna analizy: liczymy od momentu, gdy ODCZYTY komory wejda w widelki wokol nastaw:
#   • temperatura: |Todczytana - Tzadana| <= PROG_WEJSCIA_TEMP  (°C, bezwzglednie),
#   • wilgotnosc  : |RHodczytana - RHzadana| w granicach PROG_WEJSCIA_RH_PROC %  (wzglednie),
# i odliczamy STABILIZACJA_PO_RH (2h). Okno = od tego czasu do konca punktu (tam 5 reprezentantow).
# WAZNE: gdy punkt trzymany jest ~2h, samo 2h "zjadloby" caly punkt — dlatego start
# jest cofany tak, by na koncu punktu zostal zawsze ogon pomiarowy MIN_OKNO_ANALIZY.
# Punkt tylko-temperatura: liczy sie samo wejscie temperatury w widelki.
STABILIZACJA_PO_RH    = C.minuty("OBS_STAB_PO_RH", datetime.timedelta(hours=2))
PROG_WEJSCIA_TEMP     = C.liczba("OBS_PROG_T", 0.4)   # +-0.4 °C (bezwzglednie) od Tzadana
PROG_WEJSCIA_RH_PROC  = C.liczba("OBS_PROG_RH", 3.0)  # +-3% (wzglednie) od RHzadana
# gwarantowany ogon pomiarowy na koncu punktu (nigdy 0)
MIN_OKNO_ANALIZY      = C.minuty("OBS_MIN_OKNO", datetime.timedelta(minutes=15))
# Odstep 5 reprezentantow od KONCA punktu (zmiany nastawy). Na styku komora
# zaczyna juz przechodzic do kolejnego punktu, a 15-minutowe rozrzuty lapia
# probki zza granicy — odczyty wychodza rozmazane.
ODSTEP_OD_KONCA_PUNKTU = C.minuty("OBS_ODSTEP_KONIEC", datetime.timedelta(minutes=2))
# zakres T [°C] charakterystyczny dla suszenia
SUSZENIE_T_ZAKRES = (C.liczba("OBS_SUSZ_T_MIN", 21.0), C.liczba("OBS_SUSZ_T_MAX", 27.0))
SUSZENIE_RH_MAX   = C.liczba("OBS_SUSZ_RH_MAX", 50.0)  # max RH [%] — ponizej tej wartosci punkt moze byc suszeniem

# Wybor punktow wg PZ ('Zakres wzorcowania'). Jeden wsad komory obsluguje czesto KILKA
# zlecen, wiec w obserwacji sa tez punkty z innych PZ. Gdy znamy liste punktow zamowionych,
# bierzemy do protokolu WYLACZNIE je (z powtorzeniami — np. drugi punkt 50 % na histereze),
# a reszte pomijamy. Ponizsze tolerancje dopasowuja nastawy komory do wartosci z PZ
# (np. PZ '25 C / 30 %' -> nastawa komory 25,0 / 28,0).
WYBIERAJ_PUNKTY_WG_PZ = C.flaga("OBS_PZ_PUNKTY", True)
TOL_PUNKT_T  = C.liczba("OBS_TOL_PUNKT_T", 1.5)    # [st.C] dopuszczalna roznica nastawy od punktu z PZ
TOL_PUNKT_RH = C.liczba("OBS_TOL_PUNKT_RH", 4.0)   # [%RH]  dopuszczalna roznica nastawy od punktu z PZ

# Folder z zunifikowanymi plikami wynikow (z analizuj_excele.py)
WYNIKI_FOLDER = C.sciezka("ANL_OUTPUT", "wyniki", FOLDER)

# Dane przyrzadow: PDFy "Potwierdzenie zamowienia" (PZ/) + Zestawienie rozdzielczosci.
PZ_FOLDER        = C.sciezka("CC_PZ_FOLDER", "PZ", FOLDER)
ZESTAWIENIE_PLIK = C.sciezka("CC_ZESTAWIENIE",
                             "Zestawienie wzorcowanych przyrządów.xlsx", FOLDER)

# --- ZDJECIA (foto) punktow pomiarowych --------------------------------------
# KOPIUJ_FOTO=True: dla KAZDEGO wybranego punktu kopiujemy zdjecia odpowiadajace
# czasom jego 5 wierszy reprezentacyjnych (te podswietlone w obserwacji).
# Czas zdjecia czytany z NAZWY pliku (np. '2026-07-23_16.12.51.jpg'); gdy nazwa nie
# zawiera czasu — z daty modyfikacji pliku.
KOPIUJ_FOTO      = C.flaga("OBS_FOTO", False)
FOTO_ZRODLO      = C.tekst("OBS_FOTO_ZRODLO", r"\\83b\Zdjęcia")  # folder ZRODLOWY ze zdjeciami
FOTO_FOLDER      = C.sciezka("OBS_FOTO_CEL", "foto", FOLDER)  # folder DOCELOWY
FOTO_TOLERANCJA  = C.minuty("OBS_FOTO_TOL", datetime.timedelta(minutes=1))  # max odchylka czasu zdjecia od wiersza
FOTO_ROZSZERZENIA = ('.jpg', '.jpeg', '.png', '.bmp')

# Kolumna startowa dla danych srodowiskowych z wynikow w Strona 3
#   CC:   Q = 17  (temperatura=Q, wilgotnosc=R)
#   CC04: S = 19  (temperatura=S, wilgotnosc=T)
WYNIKI_START_COL_CC   = 17   # Q
WYNIKI_START_COL_CC04 = 19   # S

# Pierwszy wiersz tabeli przyrzadow na Stronie 2 (przyrzad i -> wiersz 11+i).
# Naglowek pary kolumn w Stronie 3 (wiersz 12) pokazuje nr fabryczny tego przyrzadu:
#   para 1 -> ='Strona 2'!$E$11,  para 2 -> $E$12, ...
WIERSZ_1_PRZYRZADU_S2 = 11

# Maksymalna tolerancja przy dopasowaniu timestampow [minuty; ulamki dozwolone,
# np. 0.5 = 30 s]. To PROG ODRZUCENIA, nie okno wyszukiwania: algorytm zawsze
# bierze najblizszy rekord loggera, a ta wartosc decyduje, kiedy uznac, ze plik
# nie pasuje do tego wzorcowania. Faktyczna odchylke widac w logu.
#
# Loggery ustawiane sa na zapis co 1 min, wiec typowa odchylka to sekundy.
# Zapas do 3 min pokrywa przerwy w zapisie (Aranet / Efento potrafia pominac
# 2-3 probki pod rzad). Wczesniej bylo tu 30 min — przy punkcie trwajacym 2 h
# pozwalalo to po cichu wziac odczyt z zupelnie innej fazy punktu.
WYNIKI_TOLERANCJA_MIN = C.liczba("OBS_TOL", 3.0)

# Kolumny bez zadnych danych (nieuzyte kanaly multimetru) nie sa w ogole
# wpisywane do arkusza obserwacji — tabela jest ciagla, bez dziur. Analizy ani
# protokolu to nie dotyczy: obie pracuja na sparsowanych wierszach w pamieci,
# a nie na kolumnach arkusza. Odwolania wykresow sa przeliczane po przesunieciu.
POMIJAJ_PUSTE_KOLUMNY = C.flaga("OBS_POMIJAJ_PUSTE_KOL", True)

# Maksymalna roznica miedzy odczytem przyrzadu a nastawa komory [st.C].
# Sam czas NIE wystarcza do przypisania pliku wynikow do pomiaru: w tej samej dobie
# moze trwac inne wzorcowanie w DRUGIEJ komorze (inne zlecenie). Przyrzad lezacy w
# NASZEJ komorze musi pokazywac mniej wiecej jej nastawe — pliki, ktore tego nie
# spelniaja, sa odrzucane (inaczej do protokolu trafiaja obce dane).
MAX_ROZNICA_PRZYRZAD_C = C.liczba("OBS_MAX_ROZN_PRZYRZAD", 5.0)

# Korekta ZEGARA loggera. Tanie loggery (np. Tempmate) bywaja rozjechane w czasie o
# godziny — wtedy dane trafialyby do zlego punktu. Gdy odczyty nie zgadzaja sie z
# nastawami, skrypt porownuje PROFIL temperatury loggera z profilem komory, wykrywa
# przesuniecie i koryguje je (glosno raportujac w logu). Wartosci pozostaja oryginalne
# — korygowany jest wylacznie czas dopasowania.
KOREKTA_ZEGARA          = C.flaga("OBS_KOREKTA_ZEGARA", True)
KOREKTA_ZEGARA_MAX_MIN  = C.calk("OBS_KZ_MAX", 360)   # maksymalne szukane przesuniecie [min] (+/-)
KOREKTA_ZEGARA_KROK_MIN = C.calk("OBS_KZ_KROK", 5)    # rozdzielczosc szukania [min]

# Filtr zgodnosci NASTAWY z ODCZYTEM komory:
#   B = Tzadana  vs  D = Todczytana   (temperatura)
#   C = RHzadana vs  E = RHodczytana  (wilgotnosc)
# Gdy komora NIE osiagnela nastawy, to nie jest punkt pomiarowy tylko przejscie/
# suszenie (np. 10h schladzania 35->24 st.C przy nastawie 23) — taki segment jest
# POMIJANY (nie trafia ani do obserwacji jako punkt, ani do protokolu).
# Uwaga: sam prog wzgledny [%] jest zbyt ostry przy malych nastawach (np. RH 8%:
# 10% = zaledwie 0,8 %RH), dlatego odczyt uznajemy za zgodny, gdy miesci sie w
# progu WZGLEDNYM **albo** w tolerancji BEZWZGLEDNEJ ponizej.
FILTR_NASTAWA_ODCZYT = C.flaga("OBS_FILTR", True)
MAX_ROZNICA_PROCENT  = C.liczba("OBS_PROG", 10.0)   # dozwolona wzgledna roznica |nastawa-odczyt|/nastawa * 100
TOL_ABS_TEMP = C.liczba("OBS_TOL_ABS_T", 1.0)   # [st.C]  odczyt w tych granicach od nastawy = zgodny
TOL_ABS_RH   = C.liczba("OBS_TOL_ABS_RH", 2.0)  # [%RH]   odczyt w tych granicach od nastawy = zgodny

# =============================================================================

FILL_LIGHT = PatternFill(fill_type='solid', fgColor='E2EFDA')  # jasna zieleń
FILL_DARK  = PatternFill(fill_type='solid', fgColor='A9D08E')  # ciemna zieleń
# Punkt, ktory NIE przeszedl kryterium (wybrany awaryjnie „najblizej po czasie"):
FILL_WARN_LIGHT = PatternFill(fill_type='solid', fgColor='FCE4D6')  # jasny pomaranczowy (blok)
FILL_WARN_DARK  = PatternFill(fill_type='solid', fgColor='F4B183')  # pomaranczowy (reprezentanci)


def parse_measurement_id(filename: str) -> str:
    """Zwraca część nazwy pliku po 'YYYY-MM-DD HH.MM ' (np. '119_141_147_149_151_164')."""
    base = os.path.splitext(filename)[0]
    m = re.search(r'\d{4}-\d{2}-\d{2} \d{2}\.\d{2}[ _](.+)', base)
    if not m:
        raise ValueError(f"Nierozpoznany format nazwy pliku: {filename!r}")
    return m.group(1)


def open_txt(path: str):
    """Otwiera plik TXT próbując kolejnych kodowań."""
    for enc in ('cp1250', 'utf-8-sig', 'utf-8', 'latin-1'):
        try:
            with open(path, encoding=enc) as f:
                return f.readlines()
        except UnicodeDecodeError:
            continue
    with open(path, encoding='utf-8', errors='replace') as f:
        return f.readlines()


def parse_txt(path: str):
    """
    Zwraca (sensor_name, rows):
      sensor_name – tekst po ':' z linii 8 (indeks 7)
      rows        – lista list[str] z linii 15+ (indeks 14+), pierwsze 12 kolumn (A:L)
    """
    lines = open_txt(path)

    # Linia 8 (indeks 7): "Czujnik wzorcowy: Pt100-11"
    if len(lines) > 7:
        sensor_line = lines[7].strip()
        sensor_name = sensor_line.split(':', 1)[1].strip() if ':' in sensor_line else sensor_line
    else:
        sensor_name = ''

    # Linia 15 i niżej (indeks 14+): dane pomiarowe
    rows = []
    for line in lines[14:]:
        line = line.strip()
        if not line:
            continue
        parts = re.split(r'[;\t,]', line)
        row = parts[:12]
        while len(row) < 12:
            row.append('')
        rows.append(row)

    return sensor_name, rows


def detect_file_type(lines) -> str:
    """Zwraca 'CC04' jesli linia 5 zawiera 'CC-04', inaczej 'CC'."""
    if len(lines) > 4 and 'CC-04' in lines[4]:
        return 'CC04'
    return 'CC'


# --- CC-04: kanaly czujnikow wzorcowych ------------------------------------
# Multimetr rejestruje np. 8 kanalow (Ch101..Ch108). W OBSERWACJI pokazujemy WSZYSTKIE
# (pelna struktura pliku, kolejnosc jak w pliku), a do ANALIZY i PROTOKOLU bierzemy tylko
# GLOWNE (typowo „co drugi": 101,103,105,107). Numery Pt100 (09/13/01/18) czytane sa
# z pliku PO NUMERZE KANALU — nie wpisujemy ich recznie.
CC04_KANALY_WSZYSTKIE = [101, 102, 103, 104, 105, 106, 107, 108]   # kolejnosc jak w pliku
CC04_KANALY_GLOWNE    = [101, 103, 105, 107]                       # -> analiza + protokol
CC04_KANALY_ZAPASOWE  = [c for c in CC04_KANALY_WSZYSTKIE if c not in CC04_KANALY_GLOWNE]

# Pelny uklad kolumn CC-04 — DOKLADNIE jak w pliku multimetru (wszystkie kanaly po kolei:
# najpierw odczyty ChNNN, potem tempChNNN, na koncu roztempChNNN). Kolumny dobierane
# PO NAZWIE z naglowka pliku, wiec odporne na dodatkowe/inne kanaly.
CC04_KOLUMNY = (
    ['Data Czas', 'Tzadana', 'RHzadana', 'Todczytana', 'RHodczytana']
    + [f'Ch{ch}' for ch in CC04_KANALY_WSZYSTKIE]
    + ['tdp', 'thigro', '%rh']
    + [f'tempCh{ch}' for ch in CC04_KANALY_WSZYSTKIE]
    + ['roztdp(15min)']
    + [f'roztempCh{ch}' for ch in CC04_KANALY_WSZYSTKIE]
)


def _kolumny_z_danymi(rows, n_kol, chronione=5):
    """
    Indeksy (0-based) kolumn, w ktorych JEST cokolwiek — patrzac od 2. wiersza w dol.

    Multimetr zapisuje komplet kanalow, a pracuje sie czesto na dwoch. Kolumny
    nieuzytych kanalow byly wpisywane do arkusza puste, z samym naglowkiem.

    Pierwsze `chronione` kolumn (czas, nastawy, odczyty komory) zostaja zawsze:
    to ramka arkusza, a pusta RHzadana przy pomiarze tylko-temperatura jest
    normalna. Sprawdzamy tylko kolumny objete ukladem (dlugosc `n_kol`) — dalej,
    gdzie naglowki sie koncza, nie zagladamy.
    """
    uzyte = set(range(min(chronione, n_kol)))
    do_sprawdzenia = set(range(chronione, n_kol))
    for row in rows:
        if not do_sprawdzenia:
            break                          # wszystko juz ma dane — nie ma po co czytac dalej
        znalezione = [i for i in do_sprawdzenia
                      if i < len(row) and str(row[i]).strip() not in ("", "None")]
        for i in znalezione:
            uzyte.add(i)
            do_sprawdzenia.discard(i)
    return sorted(uzyte)


def _wyczysc_pozostalosci_szablonu(ws, pierwsza_pusta, ostatnia_kolumna,
                                   wiersze_podpisow=()):
    """
    Czysci to, co zostalo po szablonie na prawo od zapisanych danych.

    Szablon ma naglowki dla PELNEGO ukladu (wszystkie kanaly) i wlasny blok
    podpisow. Dopoki zapisywalismy komplet kolumn, dane po prostu je nadpisywaly.
    Gdy puste kanaly wypadaja, po prawej zostawaly osierocone naglowki
    ('RozrzutTemperatura_Pt100(15min)' nad pusta kolumna) i drugi, stary komplet
    podpisow. Usuwamy je, zeby arkusz konczyl sie tam, gdzie koncza sie dane.
    """
    wyczyszczone = 0
    for kol in range(pierwsza_pusta, ostatnia_kolumna + 1):
        komorka = ws.cell(row=1, column=kol)
        if komorka.value is not None:
            komorka.value = None
            wyczyszczone += 1
    # Podpisy z szablonu (np. X92/X93) — tylko gdy leza juz poza danymi.
    for wiersz, kol in wiersze_podpisow:
        if kol >= pierwsza_pusta:
            komorka = ws.cell(row=wiersz, column=kol)
            if komorka.value is not None:
                komorka.value = None
                komorka.border = Border()
                komorka.number_format = "General"
                wyczyszczone += 1
    return wyczyszczone


def _formaty_kolumn_szablonu(ws, n_kol, wiersz=2):
    """
    Format liczbowy kazdej kolumny szablonu (z pierwszego wiersza danych).

    Czytamy PRZED zapisem danych — pozniej kolumny sa juz nadpisane.
    """
    return [ws.cell(row=wiersz, column=i).number_format
            for i in range(1, n_kol + 1)]


def _przenies_formaty_kolumn(ws, kolumny, formaty, ile_wierszy):
    """
    Przenosi format liczbowy razem z danymi po przesunieciu kolumn.

    Kazda kolumna szablonu ma wlasny format (np. 'Wskazania multimetru' ->
    '0.0000', reszta -> 'General'). Gdy pominiemy nieuzyte kanaly, dane wjezdzaja
    w kolumny o CUDZYM formacie i liczby dostaja obce miejsca po przecinku —
    'TPunktuRosy' 16,07 pokazywalo sie jako 16,0700, bo wyladowalo w kolumnie
    sformatowanej dla wskazan multimetru.

    Format bierzemy z kolumny ZRODLOWEJ, czyli tej, z ktorej pochodza dane.
    """
    if not formaty or ile_wierszy <= 0:
        return 0

    przeniesione = 0
    for nowy, stary in enumerate(kolumny):
        if stary >= len(formaty) or nowy >= len(formaty):
            continue
        format_zrodla = formaty[stary]
        if format_zrodla == formaty[nowy]:
            continue                      # kolumna nie zmienila miejsca albo format ten sam
        for wiersz in range(2, ile_wierszy + 2):
            ws.cell(row=wiersz, column=1 + nowy).number_format = format_zrodla
        przeniesione += 1

    if przeniesione:
        print(f"  Przeniesiono format liczbowy dla {przeniesione} przesunietych kolumn.")
    return przeniesione


def _raport_pominietych_kolumn(kolumny, naglowki, n_kol):
    """
    Wypisuje, ktore kolumny wypadly, i zwraca mape 'stary indeks -> nowy indeks'
    (0-based). Mapa sluzy pozniej do przeliczenia odwolan w wykresach.
    """
    mapa = {stary: nowy for nowy, stary in enumerate(kolumny)}
    pominiete = [i for i in range(n_kol) if i not in mapa]
    if pominiete:
        opis = ", ".join(
            f"{get_column_letter(i + 1)} ({naglowki[i]})" if i < len(naglowki)
            else get_column_letter(i + 1)
            for i in pominiete[:6])
        print(f"  Pominieto {len(pominiete)} pustych kolumn (nieuzyte kanaly): {opis}"
              + (" …" if len(pominiete) > 6 else ""))
        print(f"  Arkusz ma {len(kolumny)} kolumn zamiast {n_kol}.")
    return mapa



def _przelicz_odwolania_wykresu(xml, mapa_kolumn):
    """
    Przepisuje odwolania do kolumn w XML wykresu wg mapy 'stary -> nowy' (0-based).

    Wykresy w szablonie celuja w konkretne litery kolumn (np. $J = Ch105). Gdy
    puste kanaly wypadaja z arkusza, dane przesuwaja sie w lewo i bez przeliczenia
    wykres pokazywalby sasiednia kolumne. Podmieniamy tylko fragmenty '$LITERA$',
    jednym przebiegiem — dzieki temu zamiany nie nakladaja sie na siebie.
    """
    if not mapa_kolumn:
        return xml, []

    zgubione = []

    def zamien(m):
        litera = m.group(1)
        stary = column_index_from_string(litera) - 1
        nowy = mapa_kolumn.get(stary)
        if nowy is None:
            zgubione.append(litera)
            return m.group(0)
        return f"${get_column_letter(nowy + 1)}$"

    return re.sub(r"\$([A-Z]{1,3})\$", zamien, xml), sorted(set(zgubione))


def _przywroc_wykresy_z_szablonu(sciezka_szablonu, sciezka_wyniku, mapa_kolumn=None):
    """
    Przywraca w zapisanym arkuszu obserwacji definicje wykresow z szablonu.

    openpyxl czyta wykresy tylko czesciowo: po cyklu wczytaj-zapisz z pliku
    znikaja WSZYSTKIE serie danych (<c:ser>) i odwolania do kolumn (<c:f>).
    W arkuszu zostaja same ramki z osiami — wykresy przestaja cokolwiek
    pokazywac. Dotyczy to kazdego wygenerowanego arkusza obserwacji, niezaleznie
    od pozostalych ustawien.

    Naprawa jest chirurgiczna: po zapisie podmieniamy w gotowym .xlsx (to zwykle
    archiwum ZIP) czesci 'xl/charts/*' na oryginalne z szablonu. Reszta pliku —
    dane, kolorowanie, kotwice wykresow — zostaje taka, jaka zapisal openpyxl.

    Zwraca liczbe przywroconych czesci; 0 gdy nie bylo czego przywracac.
    Bledy sa wylapywane: brak wykresow nie moze zablokowac generowania arkusza.
    """
    try:
        with zipfile.ZipFile(sciezka_szablonu) as zs:
            czesci = {n: zs.read(n) for n in zs.namelist()
                      if n.startswith("xl/charts/")}
        if not czesci:
            return 0

        # Kolumny mogly sie przesunac (pominiete puste kanaly) — odwolania
        # wykresow trzeba przeliczyc, inaczej celowalyby w sasiednie dane.
        if mapa_kolumn:
            zgubione_lacznie = set()
            for nazwa, dane in list(czesci.items()):
                if not nazwa.endswith(".xml"):
                    continue
                xml = dane.decode("utf-8", "replace")
                xml, zgubione = _przelicz_odwolania_wykresu(xml, mapa_kolumn)
                czesci[nazwa] = xml.encode("utf-8")
                zgubione_lacznie.update(zgubione)
            if zgubione_lacznie:
                print(f"  [UWAGA] Wykres odwoluje sie do kolumn, ktorych nie ma "
                      f"w tym pomiarze: {', '.join(sorted(zgubione_lacznie))} — "
                      f"ta seria bedzie pusta.")

        with zipfile.ZipFile(sciezka_wyniku) as zw:
            zawartosc = [(i, zw.read(i.filename)) for i in zw.infolist()]
        obecne = {i.filename for i, _d in zawartosc}

        tymczasowy = sciezka_wyniku + ".tmp"
        with zipfile.ZipFile(tymczasowy, "w", zipfile.ZIP_DEFLATED) as out:
            for info, dane in zawartosc:
                out.writestr(info, czesci.get(info.filename, dane))
            for nazwa, dane in czesci.items():
                if nazwa not in obecne:
                    out.writestr(nazwa, dane)
        os.replace(tymczasowy, sciezka_wyniku)

        print(f"  Przywrocono wykresy z szablonu ({len(czesci)} czesci) — "
              f"openpyxl gubi serie danych przy zapisie.")
        return len(czesci)
    except Exception as exc:                      # noqa: BLE001 — wykresy sa dodatkiem
        print(f"  [UWAGA] Nie udalo sie przywrocic wykresow: "
              f"{type(exc).__name__}: {exc}")
        return 0



def _znajdz_naglowek(lines):
    """
    Znajduje wiersz naglowka kolumn (zaczyna sie od 'Data Czas') niezaleznie od
    dlugosci bloku metadanych. Zwraca (index_naglowka, mapa nazwa->pozycja) albo
    (None, None) gdy nie znaleziono.
    """
    for idx, line in enumerate(lines):
        if line.strip().lower().startswith('data czas'):
            names = [c.strip() for c in re.split(r'[;\t,]', line.strip())]
            return idx, {n: i for i, n in enumerate(names) if n}
    return None, None


def _mapa_kanal_pt(lines):
    """
    Mapa numer_kanalu -> 'Pt100-XX' z linii naglowka pliku CC-04, np.:
      "Czujnik wzorcowy: Pt100-09; Wejscie pomiarowe Ch: 101"  ->  {101: 'Pt100-09'}
    Uzywana do przypisania POPRAWNYCH numerow czujnikow do kanalow (glownych i zapasowych).
    """
    mapa = {}
    for line in lines:
        m = re.search(r'(Pt100-\d+).*?Ch:\s*0*(\d+)', line)
        if m:
            mapa[int(m.group(2))] = m.group(1)
    return mapa


def _naglowki_cc04(kanal_pt):
    """
    Naglowki (rzad 1) obserwacji CC-04 dla PELNEGO ukladu CC04_KOLUMNY — z nazwami
    czujnikow (Pt100-XX) zamiast surowych 'ChNNN'. Kolejnosc dokladnie jak CC04_KOLUMNY.
    """
    def et(ch):
        return kanal_pt.get(ch) or f'Ch{ch}'
    return (
        ['Data Czas', 'Tzadana', 'RHzadana', 'Todczytana', 'RHodczytana']
        + [f'Wskazania multimetru {et(ch)}' for ch in CC04_KANALY_WSZYSTKIE]
        + ['TPunktuRosy', 'Temperatura', '%RH']
        + [f'Temperatura {et(ch)}' for ch in CC04_KANALY_WSZYSTKIE]
        + ['RozrzutTPunktuRosy(15min)']
        + [f'RozrzutTemperatura_{et(ch)}(15min)' for ch in CC04_KANALY_WSZYSTKIE]
    )


def parse_txt_cc04(path: str):
    """
    Format CC-04:
      Linie 8-11 (ind. 7-10) : 4 czujniki, np. 'Czujnik wzorcowy: Pt100-09; Wejscie...'
      Naglowek kolumn        : wiersz zaczynajacy sie od 'Data Czas' (pozycja zmienna!)
      Dane pomiarowe         : wiersze ponizej naglowka

    Kolumny dobierane sa PO NAZWIE do ukladu kanonicznego CC04_KOLUMNY (21 kol.),
    co czyni parser odpornym na pliki z dodatkowymi kanalami (np. Przetwornik U) —
    kluczowe przy sklejaniu wielu plikow o roznej liczbie kolumn.
    Zwraca (sensor_names: list[str], rows: list[list[str]]).
    """
    lines = open_txt(path)

    # Numer Pt100 kazdego czujnika bierzemy PO NUMERZE KANALU (nie po kolejnosci linii,
    # bo dane sa z „co drugiego" kanalu). Nazwy do protokolu/obserwacji = kanaly GLOWNE
    # (101,103,105,107 -> Pt100-09,-13,-01,-18).
    kanal_do_pt = _mapa_kanal_pt(lines)
    sensor_names = [kanal_do_pt.get(ch, '') for ch in CC04_KANALY_GLOWNE]
    while len(sensor_names) < len(CC04_KANALY_GLOWNE):
        sensor_names.append('')

    n_kol = len(CC04_KOLUMNY)
    hdr_idx, colmap = _znajdz_naglowek(lines)
    rows = []

    if colmap is not None:
        # Mapowanie po nazwach kolumn — pelny uklad pliku (wszystkie kanaly po kolei).
        for line in lines[hdr_idx + 1:]:
            line = line.strip()
            if not line:
                continue
            parts = re.split(r'[;\t,]', line)
            row = []
            for name in CC04_KOLUMNY:
                ci = colmap.get(name)
                row.append(parts[ci] if (ci is not None and ci < len(parts)) else '')
            rows.append(row)
    else:
        # Fallback (stary tryb staloindeksowy) — gdy nie rozpoznano naglowka.
        for line in lines[17:]:
            line = line.strip()
            if not line:
                continue
            parts = re.split(r'[;\t,]', line)
            row = parts[:n_kol]
            while len(row) < n_kol:
                row.append('')
            rows.append(row)

    return sensor_names, rows


def _dostepne_txt():
    """Lista plikow .txt w FOLDER (bez plikow tymczasowych)."""
    try:
        return sorted(f for f in os.listdir(FOLDER)
                      if f.lower().endswith('.txt') and not f.startswith('~$'))
    except OSError:
        return []


def _znajdz_txt(nazwa):
    """
    Znajduje plik TXT tolerancyjnie (typowe pomylki w konfiguracji):
      - dokladna nazwa, z dodanym '.txt', ze spacja przed '.txt',
      - dopasowanie po nazwie bez rozszerzenia i nadmiaru spacji (rowne lub 'zaczyna sie od').
    Zwraca sciezke albo None.
    """
    if os.path.isabs(nazwa):
        return nazwa if os.path.exists(nazwa) else None
    baza = nazwa.strip()
    for k in (baza, baza + '.txt', baza.rstrip() + ' .txt'):
        p = os.path.join(FOLDER, k)
        if os.path.exists(p):
            return p

    def _norm(s):
        return re.sub(r'\s+', ' ', os.path.splitext(s)[0]).strip().lower()

    cel = _norm(baza)
    if cel:
        for f in _dostepne_txt():
            if _norm(f) == cel or _norm(f).startswith(cel):
                return os.path.join(FOLDER, f)
    return None


def resolve_txt_files():
    """
    Zwraca liste sciezek plikow TXT wejsciowych (posortowana chronologicznie).

    Zrodlo (wg priorytetu):
      1) zmienna OBS_TXT_FILES  — nazwy rozdzielone ';' (z panelu GUI),
      2) lista TXT_FILENAMES    — jesli niepusta,
      3) pojedynczy TXT_FILENAME.

    Nazwy wzgledne sa rozwiazywane wzgledem FOLDER. Sortowanie po nazwie pliku
    daje kolejnosc chronologiczna (nazwa zaczyna sie od 'YYYY-MM-DD HH.MM').
    """
    env = os.environ.get("OBS_TXT_FILES")
    if env:
        names = [n.strip() for n in env.split(';') if n.strip()]
    else:
        # TXT_FILENAMES (jesli zdefiniowane i niepuste) ma pierwszenstwo, w
        # przeciwnym razie TXT_FILENAME. globals().get() sprawia, ze zakomentowanie
        # TXT_FILENAMES nie wywala skryptu. Oba moga byc stringiem (jeden plik)
        # albo lista (wiele plikow) — normalizujemy do listy.
        src = globals().get("TXT_FILENAMES") or TXT_FILENAME
        names = [src] if isinstance(src, str) else list(src)
    names = [str(n).strip() for n in names if str(n).strip()]
    # Gdy lista pusta (np. TXT_FILENAMES = ["", ""]) — wroc do pojedynczego TXT_FILENAME.
    if not names and isinstance(TXT_FILENAME, str) and TXT_FILENAME.strip():
        names = [TXT_FILENAME.strip()]

    dostepne = _dostepne_txt()
    lista_txt = "\n".join(f"    • {f}" for f in dostepne) or "    (brak plikow .txt w folderze)"

    if not names:
        raise FileNotFoundError(
            "Nie podano pliku TXT — ustaw TXT_FILENAME = \"nazwa.txt\" (albo TXT_FILENAMES / "
            "OBS_TXT_FILES).\n  Dostepne pliki .txt:\n" + lista_txt)

    paths = []
    for n in names:
        p = _znajdz_txt(n)
        if p is None:
            raise FileNotFoundError(
                f"Nie znaleziono pliku TXT: '{n}'.\n  Dostepne pliki .txt:\n" + lista_txt)
        paths.append(p)

    paths.sort(key=lambda p: os.path.basename(p).lower())
    return paths


def combine_txt(paths, parse_one):
    """
    Parsuje wiele plikow TXT (przerwany pomiar) i skleja je w jeden ciag.

    - naglowek (czujnik/czujniki) bierzemy z PIERWSZEGO pliku,
    - wiersze laczymy w kolejnosci plikow (juz posortowanych chronologicznie),
    - duplikaty po znaczniku czasu (kolumna A) sa pomijane — usuwa to
      nakladajace sie probki na styku dwoch plikow.

    parse_one to parse_txt (CC) albo parse_txt_cc04 (CC-04); obie zwracaja
    (header, rows), wiec ta funkcja dziala dla obu formatow.
    """
    header = None
    combined = []
    seen_ts = set()
    for i, p in enumerate(paths):
        h, rows = parse_one(p)
        if i == 0:
            header = h
        added = 0
        for row in rows:
            ts = row[0].strip() if row and row[0] else ''
            if ts and ts in seen_ts:
                continue          # duplikat czasu (styk plikow) — pomijamy
            if ts:
                seen_ts.add(ts)
            combined.append(row)
            added += 1
        if len(paths) > 1:
            print(f"  + {os.path.basename(p)}: {len(rows)} wierszy, dodano {added}")
    return header, combined


_ILLEGAL_CHARS_RE = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f]')

def to_value(s: str):
    """Konwertuje string na odpowiedni typ Pythona (dla zapisu do Excela).
    Usuwa znaki kontrolne niedozwolone przez openpyxl."""
    s = _ILLEGAL_CHARS_RE.sub('', str(s).strip())
    if not s:
        return None
    try:
        return datetime.datetime.strptime(s, '%Y-%m-%d %H:%M:%S')
    except ValueError:
        pass
    try:
        return int(s)
    except ValueError:
        pass
    try:
        return float(s.replace(',', '.'))
    except ValueError:
        pass
    return s


# =============================================================================
# ANALIZA I KOLOROWANIE
# =============================================================================

def _s_to_float(s):
    """Konwertuje string na float; None jeśli niemożliwe (np. 'brak')."""
    try:
        return float(str(s).strip().replace(',', '.'))
    except (ValueError, AttributeError):
        return None


_DT_FORMATS = ('%Y-%m-%d %H:%M:%S', '%d.%m.%Y %H:%M:%S', '%d.%m.%Y %H:%M')

def _s_to_dt(s):
    """Parsuje string daty na datetime; obsługuje kilka formatów i podwójne spacje."""
    text = re.sub(r'\s+', ' ', str(s).strip())
    for fmt in _DT_FORMATS:
        try:
            return datetime.datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def get_k_threshold(temp, rh):
    """
    Zwraca maksymalne dozwolone K (roztdp) dla zadanych T i RH.
      0.2  — warunki dobrze kontrolowane (srodkowa wilgotnosc w typowym zakresie T),
      0.45 — wartosc MAKSYMALNA / domyslna dla wszystkich pozostalych przypadkow
             (skrajna wilgotnosc, wysoka/nietypowa temperatura itd.),
      None — tylko gdy brak danych T lub RH.

    Zaostrzony prog 0.2 obowiazuje dla:
      ~10 C (5-15)  : 40 < RH < 75
      ~23 C (15-30) : 30 <= RH <= 75
      ~35 C (30-45) : 40 <= RH <= 75
    """
    if temp is None or rh is None:
        return None

    if 5 <= temp <= 15:          # ~10 C
        if 40 < rh < 75:
            return 0.2
    elif 15 < temp <= 30:        # ~23 C
        if 30 <= rh <= 75:
            return 0.2
    elif 30 < temp <= 45:        # ~35 C
        if 40 <= rh <= 75:
            return 0.2

    return 0.45   # domyslnie — jesli nic innego nie pasuje (wartosc maksymalna)


def _prepare_data(rows):
    """
    Przetwarza wiersze str na listę krotek (dt, temp, rh, k, l, d, e).
    Indeks na liście = Excel_row - 2.
    d, e = odczyt komory (Todczytana, RHodczytana) — do filtru zgodnosci.
    """
    result = []
    for row in rows:
        result.append((
            _s_to_dt(row[0]),       # A – czas
            _s_to_float(row[1]),    # B – Tzadana
            _s_to_float(row[2]),    # C – RHzadana
            _s_to_float(row[10]),   # K – roztdp(15min)
            _s_to_float(row[11]),   # L – roztempCh001
            _s_to_float(row[3]),    # D – Todczytana (odczyt komory)
            _s_to_float(row[4]),    # E – RHodczytana (odczyt komory)
        ))
    return result


def _prepare_data_cc04(rows):
    """
    CC-04: krotka (dt, Tzad, RHzad, roztdp, l1, l2, l3, l4, Todcz, RHodcz)
    gdzie l1..l4 = rozrzuty 15-min czterech kanalow GLOWNYCH (do kryterium L).
    Indeksy w `rows` dobierane PO NAZWIE z CC04_KOLUMNY (pelny uklad pliku), by dzialalo
    niezaleznie od liczby/kolejnosci kanalow (obserwacja pokazuje wszystkie, analiza — glowne).
    """
    idx = {name: i for i, name in enumerate(CC04_KOLUMNY)}
    i_dt   = idx['Data Czas']
    i_tz   = idx['Tzadana']
    i_rhz  = idx['RHzadana']
    i_tod  = idx['Todczytana']
    i_rho  = idx['RHodczytana']
    i_rtdp = idx['roztdp(15min)']
    i_rl   = [idx[f'roztempCh{ch}'] for ch in CC04_KANALY_GLOWNE]   # 4 kanaly glowne

    def g(row, i):
        return row[i] if 0 <= i < len(row) else None

    result = []
    for row in rows:
        result.append((
            _s_to_dt(g(row, i_dt)),
            _s_to_float(g(row, i_tz)),
            _s_to_float(g(row, i_rhz)),
            _s_to_float(g(row, i_rtdp)),
            _s_to_float(g(row, i_rl[0])),
            _s_to_float(g(row, i_rl[1])),
            _s_to_float(g(row, i_rl[2])),
            _s_to_float(g(row, i_rl[3])),
            _s_to_float(g(row, i_tod)),   # Todczytana (odczyt komory)
            _s_to_float(g(row, i_rho)),   # RHodczytana (odczyt komory)
        ))
    return result


def _oblicz_start_okna(data, i, j, idx_t_odcz, idx_rh_odcz, min_stable, min_po,
                       prog_temp, prog_rh):
    """
    Zwraca start_idx okna analizy w segmencie stabilnym data[i:j] albo None.

    Zasada:
      1. Szukamy chwili wejscia ODCZYTOW komory w widelki wokol nastaw:
           • temperatura: |Todczytana - Tzadana| <= prog_temp (°C, bezwzglednie),
           • wilgotnosc (punkt z RH): |RHodczytana - RHzadana| w granicach prog_rh %.
         Punkt tylko-temperatura: liczy sie samo wejscie temperatury.
      2. Okno startuje min_po (2h) PO wejsciu. Gdy odczyty nie weszly w widelki —
         start liczony od min_stable (rozgrzewka) od poczatku punktu.
      3. Punkt trzymany jest zwykle ~2h, wiec samo "2h po wejsciu" czesto wypadaloby
         dokladnie na koncu punktu (okno = 0). Dlatego start jest cofany tak, by na
         koncu punktu zostal ZAWSZE ogon pomiarowy MIN_OKNO_ANALIZY (min. 5 minut do
         wyboru reprezentantow). To odtwarza dawne zachowanie (pomiar na koncu punktu).
      4. Punkty krotsze niz min_stable (przejsciowe/przerwane) sa pomijane (None).
    """
    seg_dt0    = data[i][0]
    seg_end_dt = data[j - 1][0]
    seg_t      = data[i][1]                   # Tzadana (nastawa temperatury)
    seg_c      = data[i][2]                   # RHzadana (nastawa wilgotnosci)
    if seg_dt0 is None or seg_end_dt is None:
        return None
    if seg_end_dt - seg_dt0 < min_stable:
        return None                           # punkt zbyt krotki — pomijamy

    temp_only = (seg_c is None or seg_c == 0.0)

    entry_dt = None
    for ki in range(i, j):
        t_odcz = data[ki][idx_t_odcz]         # Todczytana (odczyt komory)
        if seg_t is None or t_odcz is None or abs(seg_t - t_odcz) > prog_temp:
            continue                          # temperatura poza widelkami
        if temp_only:
            entry_dt = data[ki][0]            # sama temperatura w widelkach
            break
        rh = data[ki][idx_rh_odcz]            # RHodczytana (odczyt komory)
        if rh is not None and not _rozne_procent(seg_c, rh, prog_rh):
            entry_dt = data[ki][0]            # OBA (T i RH) w widelkach
            break

    if entry_dt is not None:
        start_time = entry_dt + min_po        # 2h po wejsciu w widelki
    else:
        start_time = seg_dt0 + min_stable     # brak wejscia — rozgrzewka od poczatku

    # Nie "zjadaj" calego punktu: zostaw ogon pomiarowy na koncu.
    latest = seg_end_dt - MIN_OKNO_ANALIZY
    if start_time > latest:
        start_time = latest
    if start_time < seg_dt0:
        start_time = seg_dt0

    for ki in range(i, j):
        if data[ki][0] and data[ki][0] >= start_time:
            return ki
    return None


def _find_all_analysis_windows(data, min_stable, idx_t_odcz=5, idx_rh_odcz=6,
                               min_po=None, prog_temp=None, prog_rh=None):
    """
    Dla każdego segmentu stabilnego B+C wyznacza okno analizy (start_idx, end_idx),
    gdzie end_idx = pierwszy wiersz ze zmianą B lub C (wyłącznie).

    Start okna liczy _oblicz_start_okna: okno rusza min_po (2h) od wejścia odczytów w
    widełki (T: |Todczytana-Tzadana| <= prog_temp °C; RH: prog_rh % od RHzadana). Dla
    tylko-temp liczy się sama temperatura. Fallback: po min_stable od początku.
    idx_t_odcz / idx_rh_odcz = indeksy Todczytana / RHodczytana w krotce (5/6 dla CC, 8/9 dla CC-04).
    """
    if min_po is None:
        min_po = STABILIZACJA_PO_RH
    if prog_temp is None:
        prog_temp = PROG_WEJSCIA_TEMP
    if prog_rh is None:
        prog_rh = PROG_WEJSCIA_RH_PROC

    windows = []
    n = len(data)
    i = 0
    while i < n:
        seg_b, seg_c, seg_dt0 = data[i][1], data[i][2], data[i][0]
        if seg_b is None or seg_c is None or seg_dt0 is None:
            i += 1
            continue

        # Koniec bieżącego segmentu stabilnego (stałe B i C)
        j = i + 1
        while j < n:
            b, c = data[j][1], data[j][2]
            if b != seg_b or c != seg_c:
                break
            j += 1

        start_idx = _oblicz_start_okna(data, i, j, idx_t_odcz, idx_rh_odcz,
                                       min_stable, min_po, prog_temp, prog_rh)
        if start_idx is not None:
            windows.append((start_idx, j))

        i = j  # następny segment

    return windows


def _find_best_minute_reps(data, valid_indices, start_idx, end_idx, score_fn):
    """
    Bezposrednia optymalizacja po wyniku score_fn.

    score_fn(i) -> float|None: funkcja zwracajaca metric do minimalizacji dla wiersza i.
    Przyklady:
      CC normalny:     lambda i: data[i][3]            (K – roztdp)
      CC temp-only:    lambda i: data[i][4]            (L – roztemp)
      CC-04 normalny:  lambda i: data[i][3]            (Q – roztdp)
      CC-04 temp-only: lambda i: max(data[i][4:8])    (max 4xL)

    1. Dla kazdej pelnej minuty szuka najblizszego waznego wiersza (max 30s od :00).
    2. Wybiera 5 kolejnych minut z minimalna srednia score_fn — przy rownym wyniku
       preferuje pozniejsze okno (porownanie <=).

    Zwraca liste 5 indeksow (0-based do `data`) lub None.
    """
    MAX_SEC = 30
    ONE_MIN = datetime.timedelta(minutes=1)

    if not valid_indices:
        return None

    dt_start = data[start_idx][0]
    dt_end   = data[end_idx - 1][0]
    if dt_start is None or dt_end is None:
        return None

    # Posortowane wazne wiersze: (datetime, score, data_idx)
    vrows = sorted(
        [(data[i][0], score_fn(i), i) for i in valid_indices
         if data[i][0] is not None and score_fn(i) is not None],
        key=lambda x: x[0]
    )
    vtimes = [vr[0] for vr in vrows]

    # Dla kazdej pelnej minuty szukamy najblizszego waznego wiersza
    cur = dt_start.replace(second=0, microsecond=0)
    minute_reps = {}   # minute_dt -> (score, data_idx)

    while cur <= dt_end + ONE_MIN:
        lo = bisect.bisect_left(vtimes,  cur - datetime.timedelta(seconds=MAX_SEC))
        hi = bisect.bisect_right(vtimes, cur + datetime.timedelta(seconds=MAX_SEC))
        if lo < hi:
            best = min(vrows[lo:hi], key=lambda x: abs((x[0] - cur).total_seconds()))
            minute_reps[cur] = (best[1], best[2])
        cur += ONE_MIN

    # Wybieramy 5 kolejnych minut z minimalna srednia score
    all_minutes = sorted(minute_reps.keys())
    if len(all_minutes) < 5:
        return None

    # Odstep od konca punktu. Przy plaskich odczytach wszystkie okna maja te sama
    # srednia, a remis rozstrzygany jest na korzysc POZNIEJSZEGO okna — reprezentanci
    # ladowali wiec tuz przed zmiana nastawy. Na takim styku komora zaczyna juz
    # przechodzic do kolejnego punktu, a 15-minutowe rozrzuty lapia probki zza
    # granicy — odczyty wychodza "rozmazane". Cofamy dozwolony koniec okna.
    granica = dt_end - ODSTEP_OD_KONCA_PUNKTU

    best_mean = float('inf')
    best_5    = None

    for pos in range(len(all_minutes) - 4):
        group = [all_minutes[pos + j] for j in range(5)]
        if any(group[j + 1] - group[j] != ONE_MIN for j in range(4)):
            continue
        if group[4] > granica:
            continue          # okno siega zbyt blisko zmiany nastawy
        vals = [minute_reps[m][0] for m in group]
        mean_val = sum(vals) / 5
        if mean_val <= best_mean:
            best_mean = mean_val
            best_5    = [minute_reps[m][1] for m in group]

    if best_5 is None and ODSTEP_OD_KONCA_PUNKTU:
        # Punkt za krotki, by zachowac odstep — bierzemy najlepsze okno mimo
        # wszystko, ale glosno o tym mowimy (odczyty moga byc z okolic styku).
        print(f"      [UWAGA] Punkt jest za krotki na odstep "
              f"{int(ODSTEP_OD_KONCA_PUNKTU.total_seconds() // 60)} min od zmiany "
              f"nastawy — reprezentanci wybrani az do konca punktu.")
        for pos in range(len(all_minutes) - 4):
            group = [all_minutes[pos + j] for j in range(5)]
            if any(group[j + 1] - group[j] != ONE_MIN for j in range(4)):
                continue
            vals = [minute_reps[m][0] for m in group]
            mean_val = sum(vals) / 5
            if mean_val <= best_mean:
                best_mean = mean_val
                best_5    = [minute_reps[m][1] for m in group]

    return best_5


def _szerokosc_danych(ws, domyslna):
    """
    Ile kolumn arkusz NAPRAWDE ma — liczone po naglowkach w wierszu 1.

    Nie mozna tu uzyc ani stalej z ukladu, ani ws.max_column:
      • stala (np. 33 kolumny CC-04) nie wie o kolumnach pominietych, bo kanal
        nie byl uzywany — kolorowanie ciagnelo sie wtedy daleko poza dane,
        a znacznik numeru punktu ladowal kilkanascie kolumn na prawo,
      • ws.max_column zwraca szerokosc SZABLONU, bo wyczyszczone komorki nadal
        istnieja (maja jedynie wartosc None).
    """
    szerokosc = 0
    for kol in range(1, (ws.max_column or domyslna) + 1):
        if ws.cell(row=1, column=kol).value in (None, ""):
            break
        szerokosc = kol
    return szerokosc or domyslna


def _apply_fill(ws, excel_row, fill, n_cols=12):
    for col in range(1, n_cols + 1):
        ws.cell(row=excel_row, column=col).fill = fill


def _apply_bold(ws, excel_row, n_cols=12):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=excel_row, column=col)
        f = cell.font
        cell.font = Font(
            name=f.name, size=f.size, italic=f.italic,
            underline=f.underline, strike=f.strike, color=f.color,
            bold=True,
        )


def _rozne_procent(nastawa, odczyt, prog_proc):
    """
    True, gdy odczyt komory rozni sie od nastawy o wiecej niz prog_proc [%].
    Brak ktorejkolwiek wartosci -> False (nie filtrujemy). Nastawa ~0 -> porownanie
    bezwzgledne (prog traktowany jako wartosc), by uniknac dzielenia przez zero.
    """
    if nastawa is None or odczyt is None:
        return False
    baza = abs(nastawa)
    if baza < 1e-9:
        return abs(odczyt) > prog_proc
    return abs(nastawa - odczyt) / baza * 100.0 > prog_proc


def _poza_nastawa(nastawa, odczyt, prog_proc, tol_abs):
    """
    True, gdy komora NIE osiagnela nastawy: odczyt jest poza progiem WZGLEDNYM
    (prog_proc [%]) I JEDNOCZESNIE poza tolerancja BEZWZGLEDNA (tol_abs).
    Podwojny warunek chroni przed falszywym odrzuceniem przy malych nastawach
    (np. RH 8% vs 7,1% to 11% wzglednie, ale tylko 0,9 %RH — punkt jest poprawny).
    """
    if nastawa is None or odczyt is None:
        return False
    if abs(nastawa - odczyt) <= tol_abs:
        return False
    return _rozne_procent(nastawa, odczyt, prog_proc)


def _reps_ostatnie_minuty(data, start_idx, end_idx, k=5):
    """
    Awaryjny wybor reprezentantow: po jednym wierszu na minute, k OSTATNICH minut
    okna (najblizej konca punktu). Uzywane gdy nie ma 5 kolejnych minut spelniajacych
    kryteria — bierzemy po prostu najblizsze po czasie (koniec punktu = najstabilniej).
    """
    by_min = {}
    for i in range(start_idx, end_idx):
        dt = data[i][0]
        if dt is None:
            continue
        m = dt.replace(second=0, microsecond=0)
        cur = by_min.get(m)
        if cur is None or dt.second < data[cur][0].second:
            by_min[m] = i
    minuty = sorted(by_min)
    if not minuty:
        return None
    return [by_min[m] for m in minuty[-k:]]


def _fallback_reprezentanci(data, start_idx, end_idx, score_fn):
    """
    Wybor reprezentantow BEZ twardych kryteriow (K/L, filtr nastawy) — dla punktow,
    ktore kryterium NIE przeszly, ale i tak chcemy je pokazac/wpisac (na pomaranczowo).
    Najpierw najlepsze 5 kolejnych minut wg score_fn; gdy sie nie da — 5 ostatnich minut.
    """
    valid = [i for i in range(start_idx, end_idx)
             if data[i][0] is not None and score_fn(i) is not None]
    reps = _find_best_minute_reps(data, valid, start_idx, end_idx, score_fn)
    if reps:
        return reps
    return _reps_ostatnie_minuty(data, start_idx, end_idx, 5)


def _oznacz_blok(ws, data, start_idx, end_idx, rep_indices, n_cols,
                 fill_light, fill_dark, powod=None):
    """
    Koloruje 5-minutowy blok (fill_light) + reprezentantow (fill_dark, pogrubienie).
    Gdy podano `powod` (punkt nie przeszedl kryterium) — dopisuje komentarz w komorce
    NA PRAWO od bloku (kolumna n_cols+2, wiersz pierwszego reprezentanta).
    Zwraca liste wierszy jasnego bloku.
    """
    rep_times  = [data[i][0] for i in rep_indices]
    blok_start = rep_times[0].replace(second=0, microsecond=0)
    blok_end   = blok_start + datetime.timedelta(minutes=5)
    blok = [i for i in range(start_idx, end_idx)
            if data[i][0] is not None and blok_start <= data[i][0] <= blok_end]
    for i in blok:
        _apply_fill(ws, 2 + i, fill_light, n_cols)
    for i in rep_indices:
        _apply_fill(ws, 2 + i, fill_dark, n_cols)
        _apply_bold(ws, 2 + i, n_cols)
    if powod:
        kom = ws.cell(row=2 + rep_indices[0], column=n_cols + 2)   # +1 zajmuje znacznik nr punktu
        kom.value = f"UWAGA (nie na zielono): {powod}. Wybrano 5 wierszy najblizszych czasowo."
        kom.fill  = fill_dark
        fo = kom.font
        kom.font = Font(name=fo.name, size=fo.size, bold=True, italic=True, color='9C5700')
    return blok


def _process_segment(ws, data, start_idx, end_idx, seg_num, file_type='CC',
                     punkt_z_pz=False):
    """
    Analizuje jeden segment i koloruje go w arkuszu obserwacji:
      • komora NIE osiagnela nastawy  -> POMIJAMY (to nie punkt, tylko przejscie/suszenie),
      • kryterium stabilnosci spelnione   -> ZIELONY (jasny blok + ciemni reprezentanci),
      • kryterium stabilnosci niespelnione -> POMARANCZOWY + komentarz z POWODEM
        (punkt jest realny, wiec go NIE gubimy — bierzemy najblizsze po czasie).

    `punkt_z_pz=True` — segment zostal DOPASOWANY do punktu zamowionego w PZ. Wtedy nie
    zgadujemy juz, czy to punkt pomiarowy: PZ o tym przesadza. Rozbieznosc nastawy z
    odczytem komory (np. nastawa 7,5 %RH, a komora trzyma ~13,7 % — przy tak niskiej
    wilgotnosci to normalne) daje wtedy POMARANCZOWY z uwaga, a nie pominiecie punktu.

    Zwraca krotke (rep_indices, powod):
      powod = None  -> punkt zielony (OK),
      powod = tekst -> punkt pomaranczowy (przyczyna),
      (None, None)  -> segment pominiety (nie jest punktem pomiarowym).
    """
    t0 = data[start_idx][0]
    t1 = data[end_idx - 1][0]
    b  = data[start_idx][1]
    c  = data[start_idx][2]
    print(f"  Segment {seg_num}: Excel row {2+start_idx}–{2+end_idx-1}  "
          f"T={b} RH={c}  ({t0} – {t1})")

    is_cc04    = (file_type == 'CC04')
    temp_only  = (c is not None and c == 0.0)
    n_cols     = _szerokosc_danych(ws, len(CC04_KOLUMNY) if is_cc04 else 12)
    n_l        = '4L' if is_cc04 else 'L'
    crit_label = n_l if temp_only else f"K+{n_l}"   # skrot techniczny (do konsoli)
    # Opis warunku stabilnosci po polsku (do komentarza w arkuszu — zrozumialy dla operatora)
    if temp_only:
        warunek_txt = "rozrzut temperatury czujnikow w 15 min <= 0,1 st.C"
    else:
        warunek_txt = "rozrzut temperatury czujnikow w 15 min <= 0,1 st.C oraz rozrzut punktu rosy w normie"

    # Kanaly czujnikow FAKTYCZNIE uzyte w tym pomiarze. Multimetr zapisuje kolumny tylko
    # dla podlaczonych czujnikow (np. 2 z 4: Ch105 i Ch107), reszta jest pusta. Kryterium
    # rozrzutu wolno sprawdzac WYLACZNIE na uzytych kanalach — inaczej brak danych z
    # niepodlaczonego wejscia dyskwalifikuje kazdy wiersz i punkt zawsze wychodzi
    # 'niestabilny', mimo ze rozrzut realnych czujnikow wynosi 0.
    if is_cc04:
        idx_l = [j for j in range(4, 8)
                 if any(data[i][j] is not None for i in range(start_idx, end_idx))]
    else:
        idx_l = [4]

    # Score function (wspolna dla sciezki normalnej i awaryjnej)
    if temp_only:
        if is_cc04:
            def score_fn(i):
                vals = [data[i][j] for j in idx_l if data[i][j] is not None]
                return max(vals) if vals else None
        else:
            score_fn = lambda i: data[i][4]
    else:
        score_fn = lambda i: data[i][3]   # K (roztdp) dla CC i CC-04

    powod       = None    # None => zielony (OK); tekst => pomaranczowy (przyczyna)
    rep_indices = None

    # 1) Czy komora w ogole osiagnela nastawe? Jesli NIE — to nie jest punkt pomiarowy,
    #    tylko przejscie/suszenie (np. wielogodzinne schladzanie) → POMIJAMY segment.
    if FILTR_NASTAWA_ODCZYT:
        d_idx, e_idx = (8, 9) if is_cc04 else (5, 6)
        d_vals = [data[i][d_idx] for i in range(start_idx, end_idx) if data[i][d_idx] is not None]
        e_vals = [data[i][e_idx] for i in range(start_idx, end_idx) if data[i][e_idx] is not None]
        d_med  = statistics.median(d_vals) if d_vals else None
        e_med  = statistics.median(e_vals) if e_vals else None
        odrzuc_t  = _poza_nastawa(b, d_med, MAX_ROZNICA_PROCENT, TOL_ABS_TEMP)
        odrzuc_rh = (not temp_only) and _poza_nastawa(c, e_med, MAX_ROZNICA_PROCENT, TOL_ABS_RH)
        if odrzuc_t or odrzuc_rh:
            czesci = []
            if odrzuc_t:
                czesci.append(f"T: nastawa {b} st.C, komora ~{round(d_med, 2) if d_med is not None else '-'} st.C")
            if odrzuc_rh:
                czesci.append(f"RH: nastawa {c}%, komora ~{round(e_med, 2) if e_med is not None else '-'}%")
            opis = "; ".join(czesci)

            if punkt_z_pz:
                # Punkt zamowiony w PZ — nie pomijamy go. Komora bywa fizycznie
                # niezdolna utrzymac nastawy (zwlaszcza przy bardzo niskiej wilgotnosci),
                # co jest informacja do protokolu, a nie powodem do skasowania punktu.
                powod = f"komora nie utrzymala nastawy — {opis}"
                print(f"    [UWAGA] {powod} — punkt z PZ, wiec zostaje (pomaranczowy).")
            else:
                print(f"    [POMINIETO — nie punkt pomiarowy] komora nie osiagnela nastawy: "
                      + opis + " (przejscie/suszenie).")
                return None, None

    # 2) Wiersze spelniajace kryteria K/L (tylko jesli filtr nie odrzucil)
    if powod is None:
        valid_indices = []
        for i in range(start_idx, end_idx):
            row = data[i]
            dt, temp, rh, k = row[0], row[1], row[2], row[3]
            ls = [row[j] for j in idx_l]     # rozrzut tylko z UZYTYCH kanalow
            if None in (dt, temp, rh):
                continue
            if not ls or any(l is None for l in ls):
                continue
            if any(l > 0.1 for l in ls):
                continue
            if not temp_only:
                if k is None:
                    continue
                threshold = get_k_threshold(temp, rh)
                if threshold is None or k > threshold:
                    continue
            valid_indices.append(i)

        print(f"    Wiersze spelniajace kryteria {crit_label} "
              f"(kanaly uzyte: {len(idx_l)}): {len(valid_indices)}")
        if not idx_l:
            powod = ("brak danych rozrzutu z jakiegokolwiek czujnika — nie da sie ocenic "
                     "stabilnosci punktu")
        elif not valid_indices:
            powod = f"punkt niestabilny — w zadnym momencie nie spelniono warunku: {warunek_txt}"
        else:
            rep_indices = _find_best_minute_reps(data, valid_indices, start_idx, end_idx, score_fn)
            if rep_indices is None:
                powod = f"komora za malo stabilna — brak 5 kolejnych minut, gdzie {warunek_txt}"

    # 3a) ZIELONY — kryteria spelnione
    if powod is None:
        rep_times = [data[i][0] for i in rep_indices]
        print(f"    Wiersze reprezentacyjne: {[str(t) for t in rep_times]}")
        if not temp_only:
            rep_k = [round(data[i][3], 4) for i in rep_indices]
            print(f"    K (roztdp)  : {rep_k}  srednia={round(sum(rep_k)/5, 4)}")
        if is_cc04:
            for off, lbl in enumerate(['R', 'S', 'T', 'U']):
                vals = [round(data[i][4 + off], 4) for i in rep_indices
                        if data[i][4 + off] is not None]
                if vals:
                    print(f"    L{lbl} (roztemp): {vals}  srednia={round(sum(vals)/len(vals),4)}")
        else:
            rep_l = [round(data[i][4], 4) for i in rep_indices]
            print(f"    L (roztemp) : {rep_l}  srednia={round(sum(rep_l)/5, 4)}")

        blok = _oznacz_blok(ws, data, start_idx, end_idx, rep_indices, n_cols,
                            FILL_LIGHT, FILL_DARK, powod=None)
        if blok:
            print(f"    Wiersze jasno-zielone: {len(blok)} (Excel {2+blok[0]}–{2+blok[-1]})")
        return rep_indices, None

    # 3b) POMARANCZOWY — kryterium niespelnione: nie pomijamy, wybieramy najblizsze po czasie
    fb = _fallback_reprezentanci(data, start_idx, end_idx, score_fn)
    if not fb:
        print(f"    [POMINIETO] {powod} — brak nawet danych zastepczych.")
        return None, None
    _oznacz_blok(ws, data, start_idx, end_idx, fb, n_cols,
                 FILL_WARN_LIGHT, FILL_WARN_DARK, powod=powod)
    print(f"    [POMARANCZOWY] {powod} — oznaczono {len(fb)} reprezentantow (komentarz w kol. {n_cols+2}).")
    return fb, powod


def _oznacz_numery_punktow(ws, segments, file_type):
    """
    Znaczniki nawigacyjne: NUMER punktu (1,2,3...) w kolumnie z prawej, przy pierwszym
    wierszu okna kazdego punktu. Kolumna zawiera liczbe tylko w punktach, wiec w Excelu
    mozna szybko skakac miedzy nimi (Ctrl+strzalka w dol) albo znalezc po numerze.
    Numer = pozycja punktu w protokole. Kolor = stan (zielony / pomaranczowy).
    """
    n_cols_mark = _szerokosc_danych(ws, len(CC04_KOLUMNY) if file_type == 'CC04' else 12)
    marker_col  = n_cols_mark + 1
    ws.cell(row=1, column=marker_col).value = "Nr"
    for i, (rep_idx, powod, *_) in enumerate(segments, 1):
        cell = ws.cell(row=2 + rep_idx[0], column=marker_col)
        cell.value = i
        cell.fill  = FILL_WARN_DARK if powod else FILL_DARK
        fo = cell.font
        cell.font = Font(name=fo.name, size=fo.size, bold=True)


def _wybierz_okna_wg_pz(data, windows, punkty_pz):
    """
    Wybiera z wykrytych okien te, ktore odpowiadaja punktom ZAMOWIONYM w PZ.

    Dla kazdego punktu z PZ (po kolei, razem z powtorzeniami) bierzemy NAJWCZESNIEJSZE
    jeszcze niewykorzystane okno o zgodnych nastawach (tolerancje TOL_PUNKT_T/RH).
    Dzieki temu:
      • punkt powtorzony na histereze (np. drugi raz 50 %) NIE jest gubiony — PZ go zamawia,
        wiec nie stosujemy tu reguly „suszenie/powtorzenie",
      • okna z INNEGO zlecenia (ten sam wsad komory) nie trafiaja do protokolu.
    Zwraca liste okien w kolejnosci chronologicznej.
    """
    uzyte = set()
    dopasowane = []
    for t_exp, rh_exp in punkty_pz:
        etykieta = (f"{t_exp:g} st.C / {rh_exp:g} %RH" if rh_exp is not None
                    else f"{t_exp:g} st.C (tylko temperatura)")
        # Wybieramy okno NAJBLIZSZE nastawa (a nie pierwsze pasujace w tolerancji) —
        # przy sasiednich punktach (np. 4 i 5 st.C przy nastawach 4,3 i 5,3) reguła
        # „pierwsze pasujace" krzyzowala przypisania.
        wybrany, najlepszy_dyst = None, None
        for idx, (si, _ei) in enumerate(windows):
            if idx in uzyte:
                continue
            b, c = data[si][1], data[si][2]
            if b is None or abs(b - t_exp) > TOL_PUNKT_T:
                continue
            if rh_exp is not None and (c is None or abs(c - rh_exp) > TOL_PUNKT_RH):
                continue
            dyst = abs(b - t_exp)
            if rh_exp is not None and c is not None:
                dyst += abs(c - rh_exp) / 10.0     # RH wazona slabiej (inna skala)
            if najlepszy_dyst is None or dyst < najlepszy_dyst - 1e-9:
                wybrany, najlepszy_dyst = idx, dyst
        if wybrany is None:
            print(f"  [PZ] Punkt {etykieta}: BRAK pasujacego segmentu w obserwacji "
                  f"(sprawdz nastawy komory albo tolerancje TOL_PUNKT_*).")
            continue
        uzyte.add(wybrany)
        si = windows[wybrany][0]
        print(f"  [PZ] Punkt {etykieta} -> segment {wybrany + 1} "
              f"(nastawa {data[si][1]}/{data[si][2]}, start {data[si][0]})")
        dopasowane.append((wybrany, (t_exp, rh_exp)))

    pominiete = [i for i in range(len(windows)) if i not in uzyte]
    for i in pominiete:
        si = windows[i][0]
        print(f"  [PZ] Segment {i + 1} (nastawa {data[si][1]}/{data[si][2]}, "
              f"start {data[si][0]}) — spoza zamowienia, pomijam.")

    dopasowane.sort(key=lambda x: x[0])
    # (okno, punkt_z_PZ) — punkt jest potrzebny pozniej, by wiedziec, ktore przyrzady
    # faktycznie zamowily ten punkt (reszta odczytow idzie na szaro).
    return [(windows[i], punkt) for i, punkt in dopasowane]


def analyze_and_highlight(ws, rows, file_type='CC', punkty_pz=None):
    """
    Dla każdego segmentu stabilnego B+C (>= STABILIZACJA_MIN):
      - koloruje najlepsze 5-minutowe okno na #E2EFDA
      - oznacza 5 wierszy reprezentacyjnych (#A9D08E + bold)
    """
    data = _prepare_data_cc04(rows) if file_type == 'CC04' else _prepare_data(rows)

    # pozycje odczytow komory w krotce: Todczytana / RHodczytana
    idx_t_odcz  = 8 if file_type == 'CC04' else 5
    idx_rh_odcz = 9 if file_type == 'CC04' else 6
    windows = _find_all_analysis_windows(data, STABILIZACJA_MIN, idx_t_odcz, idx_rh_odcz)
    if not windows:
        print(f"  ANALIZA: brak segmentow stabilnych >= {STABILIZACJA_MIN} — pomijam kolorowanie.")
        return

    # Gdy znamy punkty zamowione w PZ — one decyduja, ktore segmenty ida do protokolu.
    # Regula „suszenie/powtorzenie" jest wtedy zbedna (a wrecz szkodliwa: gubila punkt
    # powtorzony na histereze, ktory PZ jawnie zamawia).
    if WYBIERAJ_PUNKTY_WG_PZ and punkty_pz:
        print(f"\n  Wybor punktow wg PZ ({len(punkty_pz)} zamowionych, "
              f"{len(windows)} segmentow w obserwacji):")
        filtered = _wybierz_okna_wg_pz(data, windows, punkty_pz)
        print(f"  Do protokolu: {len(filtered)} punktow.")
        segments = []
        found = ostrzezenia = 0
        for num, ((start_idx, end_idx), punkt_pz) in enumerate(filtered, 1):
            rep_idx, powod = _process_segment(ws, data, start_idx, end_idx, num, file_type,
                                              punkt_z_pz=True)
            if rep_idx is not None:
                found += 1
                if powod:
                    ostrzezenia += 1
                segments.append((rep_idx, powod, punkt_pz))
        print(f"  Oznaczono segmentow: {found}/{len(filtered)} "
              f"(zielonych: {found - ostrzezenia}, pomaranczowych: {ostrzezenia})")
        _oznacz_numery_punktow(ws, segments, file_type)
        return segments

    # Wykrywanie gisterezy:
    # Jesli (T, RH) pojawia sie powторnie ORAZ miedzy pierwszym a obecnym wystаpieniem
    # byl segment z tym samym T ale innym RH → powrót z gisterezy.
    # Segment NASTEPUJACY po takim powrocie jest zawsze suszeniem lub punktem tylko-temp.
    # Ta sama temperatura moze sie roznic o maks. ~1 stopien (poprawka czujnika),
    # dlatego T porownujemy z tolerancja < 1.0 C przy wykrywaniu gisterezy.
    # Segment C=0 (tylko-temp) po gisterezie NIE jest pomijany — mierzymy temperature.
    def _t_key(t_val, t_map):
        """Zwraca klucz z t_map najblizszy t_val (tolerancja < 1.0 C), lub t_val jesli brak."""
        for k in t_map:
            if abs(k - t_val) < 1.0:
                return k
        return t_val

    last_rh_for_t: dict = {}   # t_key -> ostatnio widziana RH dla tego T
    seen_tr:       set  = set() # (t_key, RH) widziane co najmniej raz

    post_histereza: set = set()
    for i, (si, _) in enumerate(windows):
        b = data[si][1]
        c = data[si][2]
        t  = round(b, 1) if b is not None else None
        rh = round(c, 1) if c is not None else None
        if t is not None and rh is not None:
            # Segment tylko-temperatura (RH=0): po nim komora wraca do normy,
            # nastepny segment jest suszeniem/powrotem — tak samo jak po gisterezie.
            if rh == 0:
                if i + 1 < len(windows):
                    post_histereza.add(i + 1)
                print(f"  [TYLKO-TEMP] T~{t} RH=0 (seg {i+1}), "
                      f"nastepny seg {i+2} jest po-suszeniowy")

            tk = _t_key(t, last_rh_for_t)   # kanoniczny klucz T (tolerancja < 1 C)
            prev_rh = last_rh_for_t.get(tk)
            if (tk, rh) in seen_tr and prev_rh is not None and prev_rh != rh:
                # Powrot do (T_group, RH) po segmencie z innym RH → gistereza
                if i + 1 < len(windows):
                    post_histereza.add(i + 1)
                print(f"  [HISTEREZA] T~{tk}: ...-> {prev_rh} -> {rh} (seg {i+1}), "
                      f"nastepny seg {i+2} jest po-histerezowy")
            seen_tr.add((tk, rh))
            last_rh_for_t[tk] = rh

    # Filtrowanie suszenia:
    # Pomijamy jezeli: T w SUSZENIE_T_ZAKRES, RH w (0, SUSZENIE_RH_MAX]
    # ORAZ (ta sama para (T,RH) juz wczesniej wystapila LUB segment jest po-histerezowy).
    t_min, t_max = SUSZENIE_T_ZAKRES
    seen_bc: set = set()
    filtered = []

    for idx, (start_idx, end_idx) in enumerate(windows):
        b = data[start_idx][1]
        c = data[start_idx][2]
        bc = (round(b, 1) if b is not None else None,
              round(c, 1) if c is not None else None)

        is_suszenie = (
            b is not None and c is not None
            and c > 0                      # RH=0 = brak sterowania wilgotnoscia, nie suszenie
            and t_min <= b <= t_max
            and c <= SUSZENIE_RH_MAX
            and (bc in seen_bc             # powtorzony segment
                 or idx in post_histereza) # lub pierwszy po gisterezie
        )

        if is_suszenie:
            t0 = data[start_idx][0]
            reason = 'po-histereza' if idx in post_histereza and bc not in seen_bc else 'powtorzenie'
            print(f"  [SUSZENIE/{reason}] Pomijam segment T={b} RH={c}  (start: {t0})")
        else:
            filtered.append((start_idx, end_idx))

        seen_bc.add(bc)

    print(f"  Segmentow lacznie: {len(windows)},  po odfiltrowaniu suszenia: {len(filtered)}")

    segments = []
    found = 0
    ostrzezenia = 0
    for num, (start_idx, end_idx) in enumerate(filtered, 1):
        rep_idx, powod = _process_segment(ws, data, start_idx, end_idx, num, file_type)
        if rep_idx is not None:
            found += 1
            if powod:
                ostrzezenia += 1
            segments.append((rep_idx, powod, None))   # bez PZ nie znamy punktu zamowionego
    print(f"  Oznaczono segmentow: {found}/{len(filtered)} "
          f"(zielonych: {found - ostrzezenia}, pomaranczowych: {ostrzezenia})")

    _oznacz_numery_punktow(ws, segments, file_type)
    return segments


# =============================================================================
# DANE ŚRODOWISKOWE Z WYNIKÓW (wyniki/*.xlsx → Strona 3)
# =============================================================================

def _zaladuj_wyniki_xlsx(path):
    """
    Wczytuje zunifikowany plik wynikow. Kolumny wykrywane po NAGLOWKU (wiersz 1):
      - Czas         : kolumna 1,
      - Temperatura  : pierwsza kolumna z 'Temperatura' (dla xTHERM = wewnetrzna),
      - Temperatura2 : druga kolumna z 'Temperatura' (xTHERM = zewnetrzna; inaczej None),
      - Wilgotnosc   : kolumna z 'Wilgotn'/'%rh'.
    Gdy naglowkow brak — fallback pozycyjny (kol. 2 = temp, kol. 3 = RH),
    zgodny ze starym 3-kolumnowym formatem.
    Zwraca liste krotek (dt, temp, rh, row_1based, temp2).
    """
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active

        header = [str(c.value).lower() if c.value is not None else '' for c in ws[1]]
        temp_cols = [i for i, h in enumerate(header) if 'temperatura' in h]
        temp_idx  = temp_cols[0] if temp_cols else 1
        temp2_idx = temp_cols[1] if len(temp_cols) > 1 else None
        rh_idx    = next((i for i, h in enumerate(header) if 'wilgot' in h or '%rh' in h), 2)

        result = []
        for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or row[0] is None:
                continue
            dt_val   = row[0]
            temp_val = row[temp_idx] if len(row) > temp_idx else None
            rh_val   = row[rh_idx]   if len(row) > rh_idx else None
            temp2_val = row[temp2_idx] if (temp2_idx is not None and len(row) > temp2_idx) else None

            if isinstance(dt_val, datetime.datetime):
                dt = dt_val
            else:
                dt = _s_to_dt(str(dt_val))
            if dt is None:
                continue

            def _to_f(v):
                try:
                    return float(str(v).replace(',', '.')) if v is not None else None
                except (ValueError, TypeError):
                    return None

            result.append((dt, _to_f(temp_val), _to_f(rh_val), r_idx, _to_f(temp2_val)))
        wb.close()
        return result
    except Exception as exc:
        print(f"  [WYNIKI] Blad odczytu '{os.path.basename(path)}': {exc}")
        return []


_CACHE_WYNIKI = {}   # (sciezka, mtime) -> dane; pliki wynikow czytamy raz na uruchomienie

# Zbiorcze zestawienie (z analizuj_excele.py) lezy w tym samym folderze co pliki wynikow,
# ale NIE jest przyrzadem — trzeba je wykluczyc, inaczej zajmuje pare kolumn w protokole.
ZESTAWIENIE_WYNIKI_NAZWA = "zestawienie_pomiarow.xlsx"


def _pliki_wynikow():
    """
    Pliki wynikow POJEDYNCZYCH przyrzadow z WYNIKI_FOLDER (posortowane).
    Pomija zbiorcze 'zestawienie_pomiarow.xlsx' i pliki tymczasowe Excela ('~$').
    """
    if not os.path.isdir(WYNIKI_FOLDER):
        return []
    return sorted(
        f for f in os.listdir(WYNIKI_FOLDER)
        if f.lower().endswith('.xlsx')
        and not f.startswith('~$')
        and f.lower() != ZESTAWIENIE_WYNIKI_NAZWA.lower()
    )


def _zaladuj_wyniki_xlsx_cache(path):
    """_zaladuj_wyniki_xlsx z pamiecia (ten sam plik czytany raz)."""
    try:
        klucz = (path, os.path.getmtime(path))
    except OSError:
        return _zaladuj_wyniki_xlsx(path)
    if klucz not in _CACHE_WYNIKI:
        _CACHE_WYNIKI[klucz] = _zaladuj_wyniki_xlsx(path)
    return _CACHE_WYNIKI[klucz]


def _policz_potrzebne_przyrzady(rep_groups, rows_obs):
    """
    Ile PAR kolumn przyrzadow bedzie potrzebnych w Stronie 3 — liczone tak samo jak
    w _wypelnij_wyniki_srodowiskowe (tylko pliki pasujace czasowo do punktow; plik
    xTHERM z temperatura zewnetrzna liczy sie jako DWA przyrzady).
    Wywolywane PRZED otwarciem protokolu, zeby w razie potrzeby dolozyc kolumny.
    """
    pliki = _pliki_wynikow()
    if not pliki:
        return 0

    punkty = []
    nastawy = []
    for rep_indices, _powod, *_ in rep_groups:
        tts = [_s_to_dt(rows_obs[i][0]) for i in rep_indices]
        punkty.append(None if any(t is None for t in tts) else tts)
        r0 = rows_obs[rep_indices[0]]
        nastawy.append(_s_to_float(r0[1]) if len(r0) > 1 else None)
    tol_s = datetime.timedelta(minutes=WYNIKI_TOLERANCJA_MIN).total_seconds()

    ile = 0
    for fname in pliki:
        dane = _zaladuj_wyniki_xlsx_cache(os.path.join(WYNIKI_FOLDER, fname))
        if not dane:
            continue
        # Ta sama logika co przy wypelnianiu (czas + zgodnosc z nastawa + korekta zegara),
        # tylko po cichu — tu liczymy wylacznie ILE par kolumn bedzie potrzebnych.
        dop, _shift = _dopasuj_plik_do_punktow(dane, punkty, nastawy, tol_s,
                                               fname=fname, cicho=True)
        if not dop:
            continue
        ile += 2 if any(row[4] is not None for row in dane) else 1   # xTHERM = 2 przyrzady
    return ile


def _wykryj_przesuniecie_zegara(dane, punkty, nastawy):
    """
    Wykrywa przesuniecie ZEGARA loggera wzgledem multimetru, porownujac jego profil
    temperatury z nastawami komory. Dla kazdego kandydata przesuniecia liczy medianowe
    |odczyt - nastawa| w srodkach punktow.

    Punkty trwaja godzinami, wiec „dobrych" przesuniec jest zwykle caly przedzial.
    Wybieramy jego SRODEK (mediane) — dopasowanie trafia wtedy w srodek plateau punktu,
    a nie na krawedz, gdzie temperatura jeszcze sie zmienia (tam odczyt bywa zanizony).
    Zwraca (timedelta, odchylka) albo (None, None).
    """
    czasy = [d[0] for d in dane]
    temps = [d[1] for d in dane]
    if not czasy:
        return None, None
    cele = [(tts[len(tts) // 2], nastawy[i])
            for i, tts in enumerate(punkty)
            if tts and i < len(nastawy) and nastawy[i] is not None]
    if not cele:
        return None, None

    wyniki = []
    for mins in range(-KOREKTA_ZEGARA_MAX_MIN, KOREKTA_ZEGARA_MAX_MIN + 1,
                      KOREKTA_ZEGARA_KROK_MIN):
        sh = datetime.timedelta(minutes=mins)
        odch = []
        for t_cel, nast in cele:
            cel = t_cel + sh
            poz = bisect.bisect_left(czasy, cel)
            kand = [i for i in (poz - 1, poz) if 0 <= i < len(czasy)]
            if not kand:
                continue
            i_best = min(kand, key=lambda i: abs((czasy[i] - cel).total_seconds()))
            if temps[i_best] is not None:
                odch.append(abs(temps[i_best] - nast))
        if odch:
            wyniki.append((statistics.median(odch), abs(mins), mins))
    if not wyniki:
        return None, None

    min_odch = min(w[0] for w in wyniki)
    prog = min_odch + 0.3          # rownie dobre = w granicach 0,3 st.C od najlepszego
    dobre = sorted(w[2] for w in wyniki if w[0] <= prog)
    if not dobre:
        return None, None
    wybor = int(statistics.median(dobre))
    # zaokraglij mediane do siatki szukania i odczytaj jej faktyczna odchylke
    wybor = min(dobre, key=lambda m: abs(m - wybor))
    odch = next(w[0] for w in wyniki if w[2] == wybor)
    return datetime.timedelta(minutes=wybor), odch


def _czas_txt(sekundy):
    """'12 s' / '3,5 min' — odchylka czasu w czytelnej postaci."""
    if sekundy is None:
        return "?"
    if sekundy < 90:
        return f"{sekundy:.0f} s"
    return f"{sekundy / 60:.1f} min".replace(".", ",")


def _odchylki_dopasowania(dopasowania, punkty, shift=None):
    """
    Zwraca (srednia, maksymalna) odchylke czasu [s] dopasowanych wierszy.

    Sluzy do POKAZANIA, jak blisko naprawde trafilismy. Sam algorytm zawsze
    bierze NAJBLIZSZY rekord loggera, a tolerancja jest tylko progiem odrzucenia
    — bez tej liczby w logu nie widac roznicy miedzy trafieniem co do sekundy
    a wzieciem odczytu sprzed kwadransa.
    """
    przesun = shift or datetime.timedelta(0)
    odchylki = []
    for idx, matched in dopasowania.items():
        cele = punkty[idx] if idx < len(punkty) else None
        if not cele:
            continue
        for i, krotka in enumerate(matched):
            if i < len(cele) and krotka[0] is not None:
                odchylki.append(abs(((cele[i] + przesun) - krotka[0]).total_seconds()))
    if not odchylki:
        return None, None
    return sum(odchylki) / len(odchylki), max(odchylki)


def _ostrzez_o_odchylce(fname, odch_max, tol_s):
    """
    Glosno ostrzega, gdy najgorsze dopasowanie jest daleko od czasu pomiaru.

    Punkt pomiarowy trwa godzinami, wiec odczyt wziety kilkanascie minut obok
    moze pochodzic z zupelnie innej fazy punktu. Zwykle oznacza to rzadkie
    probkowanie loggera albo dziure w jego zapisie.
    """
    if odch_max is None:
        return
    prog = min(120.0, tol_s / 4) if tol_s else 120.0
    if odch_max > prog:
        _u = _czas_txt(odch_max)
        print(f"      [UWAGA] '{fname}': najdalszy dopasowany odczyt jest o {_u} "
              f"od czasu pomiaru. Sprawdz gestosc zapisu loggera albo zmniejsz "
              f"tolerancje dopasowania czasu.")


def _dopasuj_plik_do_punktow(dane, punkty, nastawy, tol_s, fname="", cicho=False):
    """
    Dopasowuje jeden plik wynikow do punktow pomiarowych.

    Dwa warunki: zgodnosc CZASU (5 najblizszych wierszy w tolerancji) oraz zgodnosc
    ODCZYTU z nastawa komory (przyrzad lezacy w komorze musi pokazywac to, co komora).
    Gdy odczyty nie pasuja, a KOREKTA_ZEGARA=True — szuka przesuniecia zegara loggera.

    Zwraca (dopasowania, przesuniecie) albo (None, None) gdy plik jest z innego wzorcowania.
    """
    def _dop(shift):
        dop = {}
        for idx, tts in enumerate(punkty):
            if tts is None:
                continue
            tt = [t + shift for t in tts] if shift else tts
            m = _znajdz_5_wierszy(dane, tt)
            if m is None:
                continue
            avg = sum(abs((tt[i] - m[i][0]).total_seconds()) for i in range(5)) / 5
            if avg <= tol_s:
                dop[idx] = m
        return dop

    def _odch(dop):
        o = []
        for idx, m in dop.items():
            nast = nastawy[idx] if idx < len(nastawy) else None
            vals = [k[1] for k in m if k[1] is not None]
            if nast is not None and vals:
                o.append(abs(statistics.median(vals) - nast))
        return statistics.median(o) if o else None

    zero = datetime.timedelta(0)
    dop0 = _dop(zero)
    od0  = _odch(dop0)
    if dop0 and (od0 is None or od0 <= MAX_ROZNICA_PRZYRZAD_C):
        return dop0, zero      # wszystko gra bez korekty

    if not KOREKTA_ZEGARA:
        if not cicho and dop0:
            print(f"  [WYNIKI] '{fname}' — ODRZUCONY: odczyty odbiegaja od nastaw komory "
                  f"o ~{od0:.1f} st.C (limit {MAX_ROZNICA_PRZYRZAD_C:g}).")
        return None, None

    shift, odch_sh = _wykryj_przesuniecie_zegara(dane, punkty, nastawy)
    if shift is None or odch_sh is None or odch_sh > MAX_ROZNICA_PRZYRZAD_C:
        if not cicho and dop0:
            print(f"  [WYNIKI] '{fname}' — ODRZUCONY: odczyty odbiegaja od nastaw komory "
                  f"o ~{od0:.1f} st.C i nie znaleziono sensownego przesuniecia zegara. "
                  f"To plik z INNEGO wzorcowania — usun go z folderu 'wyniki'.")
        return None, None

    dop = _dop(shift)
    od  = _odch(dop)
    if not dop or od is None or od > MAX_ROZNICA_PRZYRZAD_C:
        if not cicho and dop0:
            print(f"  [WYNIKI] '{fname}' — ODRZUCONY (po probie korekty zegara).")
        return None, None

    if not cicho:
        godz = shift.total_seconds() / 3600.0
        print(f"  [WYNIKI] '{fname}' — WYKRYTO PRZESUNIECIE ZEGARA loggera: "
              f"{godz:+.2f} h. Skorygowano czasy dopasowania "
              f"(odchylka {od0:.1f} -> {od:.1f} st.C). Wartosci pomiarowe bez zmian.")
    return dop, shift


def _miejsca_po_przecinku(res):
    """Ile miejsc po przecinku wynika z rozdzielczosci: 1->0, 0.1->1, 0.01->2, 0.001->3."""
    if not res or res <= 0:
        return None
    for dec, krok in ((0, 1.0), (1, 0.1), (2, 0.01), (3, 0.001)):
        if abs(res - krok) < 1e-9:
            return dec
    return min(4, max(0, int(round(-math.log10(res)))))


def _zaokr_do_rozdz(val, res):
    """
    Zaokragla wartosc do rozdzielczosci przyrzadu (np. 0.1 -> jedno miejsce po przecinku).
    Bez rozdzielczosci — awaryjnie 4 miejsca (jak dotad).
    """
    dec = _miejsca_po_przecinku(res)
    return round(val, 4 if dec is None else dec)


def _format_rozdz(res):
    """Format liczbowy Excela wg rozdzielczosci: 0.1 -> '0.0', 0.01 -> '0.00', 1 -> '0'."""
    dec = _miejsca_po_przecinku(res)
    if dec is None:
        return None
    return '0' if dec == 0 else '0.' + '0' * dec


def _znajdz_5_wierszy(wyniki_rows, target_times):
    """
    Dla listy 5 docelowych dat/czasow zwraca liste 5 najblizszych wpisow
    z wyniki_rows [(dt, temp, rh, row_idx, temp2)].
    Zwraca liste 5 krotek lub None jesli wyniki_rows jest puste.
    """
    if not wyniki_rows:
        return None
    matched = []
    for t_target in target_times:
        best = min(wyniki_rows, key=lambda e: abs((e[0] - t_target).total_seconds()))
        matched.append(best)
    return matched


def _oznacz_wyniki_xlsx(path, row_indices):
    """Koloruje podane wiersze w pliku wynikow na ciemno-zielono + bold."""
    try:
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        n_cols = ws.max_column
        for r_idx in row_indices:
            for col in range(1, n_cols + 1):
                cell = ws.cell(row=r_idx, column=col)
                cell.fill = FILL_DARK
                f = cell.font
                cell.font = Font(
                    name=f.name, size=f.size, italic=f.italic,
                    underline=f.underline, strike=f.strike, color=f.color,
                    bold=True,
                )
        wb.save(path)
        print(f"    [WYNIKI] Oznaczono {len(row_indices)} wierszy: {os.path.basename(path)}")
    except Exception as exc:
        print(f"    [WYNIKI] Blad oznaczania '{os.path.basename(path)}': {exc}")


def _wypelnij_wyniki_srodowiskowe(proto_ws, rep_groups, rows_obs, obs_type):
    """
    Wpisuje dane srodowiskowe (Temperatura+Wilgotnosc) z plikow WYNIKI_FOLDER do
    Strona 3.

    UKLAD: kazdy przyrzad (logger) = osobna PARA kolumn idaca w PRAWO —
      CC   : Q/R, S/T, U/V, ...      (start_col = 17)
      CC04 : S/T, U/V, W/X, Y/Z, AA/AB (start_col = 19, max 5 par)
    Wiersze (bloki po 5) = kolejne punkty pomiarowe.

    Pliki przypisywane do przyrzadow po KOLEJNOSCI NAZW; brane sa tylko te, ktore
    pasuja czasowo do punktow (stare/niepasujace pliki pomijamy automatycznie).

    PRZYPADEK SPECJALNY (rzadki) — plik z 4 kolumnami (xTHERM: Czas | Temp wewn |
    Temp zewn | Wilgotnosc) liczy sie jako DWA przyrzady:
      - para N   : temperatura WEWNETRZNA + wilgotnosc,
      - para N+1 : temperatura ZEWNETRZNA, BEZ wilgotnosci (kolumna RH pusta).
    """
    if not os.path.exists(WYNIKI_FOLDER):
        print(f"  [WYNIKI] Brak folderu '{WYNIKI_FOLDER}' — pomijam dane srodowiskowe.")
        return

    pliki = _pliki_wynikow()
    if not pliki:
        print(f"  [WYNIKI] Brak plikow .xlsx w '{WYNIKI_FOLDER}' — pomijam.")
        return

    print(f"  [WYNIKI] Wczytuje {len(pliki)} plik(ow) wynikow...")
    baza = {}
    for fname in pliki:
        dane = _zaladuj_wyniki_xlsx_cache(os.path.join(WYNIKI_FOLDER, fname))
        if dane:
            baza[fname] = dane
    if not baza:
        print(f"  [WYNIKI] Zadne pliki wynikow nie zawieraja danych — pomijam.")
        return

    start_col = WYNIKI_START_COL_CC04 if obs_type == 'CC04' else WYNIKI_START_COL_CC
    BLOCK_START_ROW = 20
    BLOCK_SIZE      = 5
    tol_s = datetime.timedelta(minutes=WYNIKI_TOLERANCJA_MIN).total_seconds()

    # Liczba par kolumn przyrzadow — wykrywana z naglowka szablonu (wiersz 10 ma
    # numery 1,2,3,... co 2 kolumny: S=1, U=2, W=3, ...). Dzieki temu nie zależymy
    # od stalej — szablon CC04 ma ich 10 (S/T..AK/AL), dalej sa kolumny 'rozrzut'.
    MAX_PRZYRZADY = 0
    _c = start_col
    while _c <= proto_ws.max_column:
        _v = proto_ws.cell(row=10, column=_c).value
        if isinstance(_v, (int, float)) and int(_v) == MAX_PRZYRZADY + 1:
            MAX_PRZYRZADY += 1
            _c += 2
        else:
            break
    if MAX_PRZYRZADY == 0:
        MAX_PRZYRZADY = 5   # fallback, gdy nie udalo sie odczytac naglowka
    print(f"  [WYNIKI] Szablon ma {MAX_PRZYRZADY} par kolumn przyrzadow (od "
          f"{get_column_letter(start_col)}).")

    # Docelowe czasy dla kazdego punktu (None gdy brak timestampow) + nastawa temperatury
    # punktu (kolumna B) — sluzy do sprawdzenia, czy plik wynikow pochodzi z NASZEJ komory.
    punkty = []
    nastawy = []
    for rep_indices, _powod, *_ in rep_groups:
        tts = [_s_to_dt(rows_obs[i][0]) for i in rep_indices]
        punkty.append(None if any(t is None for t in tts) else tts)
        r0 = rows_obs[rep_indices[0]]
        nastawy.append(_s_to_float(r0[1]) if len(r0) > 1 else None)

    oznaczenia_per_plik = {}   # fname -> set(row_idx)
    dev = 0                    # numer przyrzadu (0-based) -> para kolumn w prawo
    uzyte = []                 # per przyrzad (w kolejnosci dev): (serial, temps, rhs)

    def _wpisz(col, matched_per_punkt, val_idx, z_wilgotnoscia, res_t=None, res_rh=None):
        """
        Wpisuje jedna kolumne (temp z val_idx) + opc. wilgotnosc do blokow punktow.
        Wartosci sa ZAOKRAGLANE do rozdzielczosci przyrzadu (res_t/res_rh, np. 0.1) i
        dostaja odpowiedni format liczbowy ('0.0'), zeby nie pokazywac zer na setnych.
        """
        for punkt_idx, matched in matched_per_punkt.items():
            r0 = BLOCK_START_ROW + punkt_idx * BLOCK_SIZE
            for off, krotka in enumerate(matched):
                er = r0 + off
                temp_v = krotka[val_idx]
                rh_v   = krotka[2]
                row_idx = krotka[3]
                if temp_v is not None:
                    c = proto_ws.cell(row=er, column=col)
                    c.value = _zaokr_do_rozdz(temp_v, res_t)
                    fmt = _format_rozdz(res_t)
                    if fmt:
                        c.number_format = fmt
                if z_wilgotnoscia and rh_v is not None:
                    c = proto_ws.cell(row=er, column=col + 1)
                    c.value = _zaokr_do_rozdz(rh_v, res_rh)
                    fmt = _format_rozdz(res_rh)
                    if fmt:
                        c.number_format = fmt
                oznaczenia_per_plik.setdefault(fname, set()).add(row_idx)

    for fname in pliki:
        if fname not in baza:
            continue
        dane = baza[fname]

        # Dopasowanie pliku do punktow: zgodnosc CZASU + zgodnosc odczytow z nastawami
        # komory (chroni przed doklejeniem plikow z innego wzorcowania), a gdy zegar
        # loggera jest rozjechany — automatyczna korekta przesuniecia.
        dopasowania, _shift = _dopasuj_plik_do_punktow(dane, punkty, nastawy, tol_s,
                                                       fname=fname)
        if not dopasowania:
            continue

        odch_sr, odch_max = _odchylki_dopasowania(dopasowania, punkty, _shift)

        ma_zewn = any(row[4] is not None for row in dane)   # temp2 -> xTHERM = 2 przyrzady

        if dev >= MAX_PRZYRZADY:
            print(f"  [WYNIKI] '{fname}' — brak wolnej kolumny przyrzadu "
                  f"(max {MAX_PRZYRZADY}) — pomijam.")
            continue

        # Rozdzielczosc przyrzadu z JEGO danych (np. 0.1) — do zaokraglenia i formatu
        # komorek. Liczona z calego pliku, wiec odporna na pojedyncze rowne wartosci.
        res_t  = pz_dane.rozdzielczosc_z_kolumny([r[1] for r in dane])
        res_rh = pz_dane.rozdzielczosc_z_kolumny([r[2] for r in dane])
        res_t2 = pz_dane.rozdzielczosc_z_kolumny([r[4] for r in dane]) if ma_zewn else None

        # Przyrzad: temperatura (wewnetrzna) + wilgotnosc
        col = start_col + dev * 2
        punkty_txt = ', '.join(str(p + 1) for p in sorted(dopasowania))
        print(f"  [WYNIKI] Przyrzad {dev+1} ({get_column_letter(col)}/{get_column_letter(col+1)})"
              f" <- '{fname}'  (punkty: {punkty_txt}; rozdz. t={res_t} rh={res_rh};"
              f" odchylka czasu sr. {_czas_txt(odch_sr)}, max {_czas_txt(odch_max)})")
        _ostrzez_o_odchylce(fname, odch_max, tol_s)
        _wpisz(col, dopasowania, val_idx=1, z_wilgotnoscia=True, res_t=res_t, res_rh=res_rh)
        temps = [k[1] for m in dopasowania.values() for k in m if k[1] is not None]
        rhs   = [k[2] for m in dopasowania.values() for k in m if k[2] is not None]
        uzyte.append((_serial_z_wyniku(fname), temps, rhs))
        dev += 1

        # Przyrzad 2 (tylko xTHERM): temperatura zewnetrzna, BEZ wilgotnosci
        if ma_zewn:
            if dev >= MAX_PRZYRZADY:
                print(f"  [WYNIKI] '{fname}' (zewn.) — brak wolnej kolumny — pomijam czesc zewn.")
            else:
                col = start_col + dev * 2
                print(f"  [WYNIKI] Przyrzad {dev+1} ({get_column_letter(col)}, temp zewn., bez RH)"
                      f" <- '{fname}'")
                _wpisz(col, dopasowania, val_idx=4, z_wilgotnoscia=False, res_t=res_t2)
                temps2 = [k[4] for m in dopasowania.values() for k in m if k[4] is not None]
                uzyte.append((_serial_z_wyniku(fname), temps2, []))
                dev += 1

    if dev == 0:
        print(f"  [WYNIKI] Zaden plik nie pasowal czasowo do punktow — brak danych srodowiskowych.")

    # Oznacz wiersze w plikach wynikow
    for fname, row_set in oznaczenia_per_plik.items():
        _oznacz_wyniki_xlsx(os.path.join(WYNIKI_FOLDER, fname), sorted(row_set))

    return uzyte


# Data/godzina doklejona do nazwy pliku przez program logujacy (np. LogSoft):
#   '1970325 2026-07-31 12.19.00_wynik.xlsx'  ->  serial '1970325'
_RE_TS_W_NAZWIE = re.compile(
    r'[\s_-]*\d{4}-\d{2}-\d{2}[\s_T-]+\d{1,2}[.:_-]\d{2}(?:[.:_-]\d{2})?\s*$')


def _serial_z_wyniku(fname):
    """
    Nr fabryczny przyrzadu z nazwy pliku wynikow (klucz do dopasowania z PZ):
      'TMM230200349_wynik.xlsx'                 -> 'TMM230200349'
      '1970325 2026-07-31 12.19.00_wynik.xlsx'  -> '1970325'
    Odcinamy sufiks '_wynik' oraz date/godzine, ktora niektore programy (LogSoft)
    dokleja do nazwy pliku — bez tego klucz nie pasowal do PZ i tabela przyrzadow
    na Stronie 2 zostawala pusta.
    """
    base = os.path.splitext(fname)[0]
    base = re.sub(r'_wynik$', '', base, flags=re.I).strip()
    base = _RE_TS_W_NAZWIE.sub('', base).strip()
    # Koncowka '_2', '_3'... to numer KOLEJNEGO pomiaru tego samego przyrzadu
    # (np. '37025105_2' = drugie wzorcowanie), a nie czesc numeru fabrycznego —
    # bez odciecia przyrzad nie zostalby znaleziony w PZ.
    bez_powtorki = re.sub(r'_\d{1,2}$', '', base)
    return bez_powtorki or base


# Kolumny Strony 2 (tabela przyrzadow, od wiersza 11):
#  B=2 wytworca(obiekt) C=3 wytworca(czujnik) D=4 typ E=5 nr fabr F=6 nr ewid
#  G=7 adres H=8 typ(czujnik) I=9 nr fabr(czujnik) J=10 nr ewid(czujnik)
#  K=11 rozdz. t  L=12 rozdz. RH  O=15 nr zlecenia
STRONA2_PIERWSZY_WIERSZ = 11

# Blok podpisow po PRAWEJ stronie Strony 2 (T = data, V = podpis; komorki scalone):
#   wiersz 4  — '4. Pomiary wykonal(a)'      (wiersze 5-8 to miejsce na kolejne osoby)
#   wiersz 10 — '5. Protokol sprawdzil(a)'
STRONA2_KOL_DATA   = 20   # T
STRONA2_KOL_PODPIS = 22   # V
STRONA2_W_WYKONAL  = 4
STRONA2_W_SPRAWDZIL = 10


def wypelnij_podpisy_strona2(ws2, data_pomiaru=None):
    """
    Wpisuje date i podpis w bloku po prawej stronie Strony 2:
      'Pomiary wykonal(a)'    -> data POMIARU (z ostatniego punktu) + PODPIS,
      'Protokol sprawdzil(a)' -> data dzisiejsza + PODPIS_SPRAWDZIL.
    Szablon ma tam '-' w komorkach; komorki sa scalone, wiec piszemy przez _ustaw_komorke.
    """
    dzis = datetime.date.today().strftime('%d.%m.%Y')
    data_wyk = data_pomiaru.strftime('%d.%m.%Y') if data_pomiaru else dzis

    _ustaw_komorke(ws2, STRONA2_W_WYKONAL,   STRONA2_KOL_DATA,   data_wyk)
    _ustaw_komorke(ws2, STRONA2_W_WYKONAL,   STRONA2_KOL_PODPIS, PODPIS)
    _ustaw_komorke(ws2, STRONA2_W_SPRAWDZIL, STRONA2_KOL_DATA,   dzis)
    _ustaw_komorke(ws2, STRONA2_W_SPRAWDZIL, STRONA2_KOL_PODPIS, PODPIS_SPRAWDZIL)
    print(f"    podpisy: wykonal '{PODPIS}' ({data_wyk}), "
          f"sprawdzil '{PODPIS_SPRAWDZIL}' ({dzis})")


def wypelnij_strone2_z_pz(ws2, uzyte, pz_mapa, zest):
    """
    Wypelnia tabele przyrzadow na Stronie 2 protokolu na podstawie PZ.

    `uzyte` = lista (serial, temps, rhs) w kolejnosci kolumn pomiarowych Strony 3
    (i-ty przyrzad -> wiersz 11+i). Dopasowanie przyrzadu z PZ po nr fabrycznym
    (serial z nazwy pliku wyniku). Rozdzielczosc: z Zestawienia (po producencie+typie),
    a gdy brak — z wahania cyfr po przecinku w danych pomiarowych.
    Przyrzady bez dopasowania w PZ (np. mierniki reczne bez pliku logera) zostaja
    do recznego uzupelnienia (log ostrzegawczy).
    """
    if not uzyte:
        return
    print("\n  Wypelnianie tabeli przyrzadow (Strona 2) z PZ...")
    for i, (serial, temps, rhs) in enumerate(uzyte):
        w = STRONA2_PIERWSZY_WIERSZ + i
        dev = None
        if pz_mapa:
            # Dopasowanie po pelnym kluczu, a gdy brak — po pierwszym czlonie nazwy
            # (pliki z niektorych programow maja w nazwie dodatki po nr seryjnym).
            for kandydat in (serial, serial.split()[0] if serial.split() else ""):
                if not kandydat:
                    continue
                dev = pz_mapa.get(pz_dane.normalizuj_serial(kandydat))
                if dev is not None:
                    break
        if dev is None:
            print(f"    wiersz {w}: brak dopasowania w PZ (serial '{serial}') — uzupelnij recznie.")
            continue

        # Rozdzielczosc: Zestawienie -> fallback z danych
        t_res, rh_res = pz_dane.rozdzielczosc_zestawienie(zest, dev.wytworca, dev.typ)
        zrodlo = "Zestawienie"
        if t_res is None:
            t_res = pz_dane.rozdzielczosc_z_kolumny(temps); zrodlo = "dane"
        if rh_res is None:
            rh_res = pz_dane.rozdzielczosc_z_kolumny(rhs)
            if rh_res is None:       # przyrzad tylko-temperatura — brak danych RH
                rh_res = t_res

        def _set(col, val):
            ws2.cell(row=w, column=col).value = val if (val not in (None, "")) else "-"

        # Ten protokol dotyczy wzorcowania w KOMORZE KLIMATYCZNEJ — czujnik wewnetrzny
        # (kanal 1), wiec kolumny 'Czujnik pomiarowy' zostaja puste ('-'). Gdy przyrzad
        # trafil z pozycji PZ wzorcowanej w termostacie (czujnik zewnetrzny 0572 1001),
        # tych danych NIE przepisujemy — naleza do innego protokolu.
        w_komorze = getattr(dev, 'komora', True)
        czuj_w  = dev.czuj_wytworca if w_komorze else ""
        czuj_t  = dev.czuj_typ      if w_komorze else ""
        czuj_nf = dev.czuj_nr_fabr  if w_komorze else ""

        _set(2, dev.wytworca)            # B
        _set(3, czuj_w)                  # C
        _set(4, dev.typ)                 # D
        _set(5, dev.nr_fabr)             # E
        _set(6, dev.nr_ewid)             # F
        _set(7, "")                      # G (adres) — brak w PZ
        _set(8, czuj_t)                  # H
        _set(9, czuj_nf)                 # I
        _set(10, "")                     # J (nr ewid czujnika) — brak w PZ
        ws2.cell(row=w, column=11).value = t_res    # K
        ws2.cell(row=w, column=12).value = rh_res   # L
        ws2.cell(row=w, column=15).value = dev.nr_zlecenia   # O
        print(f"    wiersz {w}: {dev.wytworca} / {dev.typ} / {dev.nr_fabr} "
              f"(zlec {dev.nr_zlecenia}, K={t_res} L={rh_res} [{zrodlo}])")


# =============================================================================
# GENEROWANIE PROTOKOŁU
# =============================================================================

def _wyszarz_punkty_spoza_zamowienia(ws3, rep_groups, uzyte, obs_type,
                                     block_start_row, block_size, fill_grey):
    """
    Wyszarza odczyty przyrzadu w punktach, ktorych NIE zamowiono wlasnie dla niego.

    Jeden wsad komory obsluguje kilka zlecen, a w PZ kazda pozycja (przyrzad) ma wlasny
    zakres, np.:
        3) (2; 8) °C     -> 37025101, 37025108
        6) (24; 25; 26) °C -> 37025105
    Protokol zawiera SUME punktow, wiec kazdy przyrzad ma w nim rowniez punkty, na ktore
    nie byl zamowiony. Takie komorki dostaja SZARE tlo, ale ODCZYTY W NICH ZOSTAJA:
    przyrzad lezal w komorze przez caly wsad, wiec dane sa prawdziwe i przydatne
    informacyjnie — szare tlo mowi jedynie, ze punkt nie wchodzi do jego swiadectwa.

    Wymaga: punktu z PZ przypisanego do bloku (3. element rep_groups) oraz mapy
    przyrzad -> punkty. Gdy ktoregos brak, nic nie zmieniamy.
    """
    if not uzyte or not rep_groups:
        return
    mapa_punktow = pz_dane.wczytaj_punkty_przyrzadow(PZ_FOLDER)
    if not mapa_punktow:
        return

    start_col = WYNIKI_START_COL_CC04 if obs_type == 'CC04' else WYNIKI_START_COL_CC
    razem = 0
    for dev_idx, wpis in enumerate(uzyte):
        serial = wpis[0] if wpis else None
        if not serial:
            continue
        punkty_dev = (mapa_punktow.get(pz_dane.normalizuj_serial(serial))
                      or mapa_punktow.get(pz_dane.normalizuj_serial(str(serial).split()[0])))
        if not punkty_dev:
            continue                       # nie wiemy, co zamowiono — nie ruszamy
        zbedne = []
        col = start_col + dev_idx * 2
        for punkt_idx, wiersz in enumerate(rep_groups):
            punkt_pz = wiersz[2] if len(wiersz) > 2 else None
            if punkt_pz is None:
                continue                   # brak przypisania punktu z PZ — zostawiamy
            if punkt_pz in punkty_dev:
                continue                   # ten punkt byl zamowiony dla tego przyrzadu
            r0 = block_start_row + punkt_idx * block_size
            for row_off in range(block_size):
                for c in (col, col + 1):
                    # Odczyty ZOSTAJA — przyrzad byl w komorze, wiec dane istnieja i sa
                    # informacja pomocnicza. Szare tlo oznacza tylko, ze punkt nie byl
                    # zamowiony dla tego przyrzadu i nie wchodzi do jego swiadectwa.
                    ws3.cell(row=r0 + row_off, column=c).fill = fill_grey
            zbedne.append(punkt_idx + 1)
            razem += 1
        if zbedne:
            print(f"    [PZ] '{serial}' ({get_column_letter(col)}/{get_column_letter(col+1)}): "
                  f"punkty spoza zamowienia -> szare: {zbedne} "
                  f"(zamowiono {pz_dane._opis_punktow(punkty_dev)})")
    if razem:
        print(f"    [PZ] Wyszarzono {razem} blokow odczytow spoza zamowienia.")


def _delete_rows_via_excel(filepath, sheet_name, row_from, row_count):
    """
    Usuwa row_count wierszy od row_from przez Excel COM (przesuwa reszte w gore —
    odpowiednik zaznaczenia wierszy + Delete / Przesuniecie w gore).
    Zwraca True przy sukcesie, False jesli COM niedostepny.
    """
    try:
        import win32com.client
    except ImportError:
        return False

    xl = wb = None
    try:
        xl = win32com.client.Dispatch('Excel.Application')
        xl.Visible        = False
        xl.DisplayAlerts  = False
        xl.ScreenUpdating = False

        wb = xl.Workbooks.Open(os.path.abspath(filepath))
        ws = None
        for idx in range(1, wb.Sheets.Count + 1):
            if wb.Sheets(idx).Name == sheet_name:
                ws = wb.Sheets(idx)
                break
        if ws is None:
            return False

        addr = f'{row_from}:{row_from + row_count - 1}'
        ws.Rows(addr).Delete(Shift=-4162)   # xlShiftUp

        wb.Save()
        return True

    except Exception as exc:
        print(f'  [Excel COM] Blad usuwania wierszy: {exc}')
        return False
    finally:
        if wb is not None:
            try: wb.Close(False)
            except: pass
        if xl is not None:
            try: xl.Quit()
            except: pass


def _insert_blocks_via_excel(filepath, sheet_name, block_start, block_size, n_extra):
    """
    Otwiera plik w Excelu przez COM i wykonuje n_extra razy operacje
    'Kopiuj wiersze block_start..block_start+block_size-1 → Wstaw skopiowane komorki'
    (odpowiednik Ctrl+C / Wstaw z przesuniciem w dol).
    Zapisuje i zamyka plik.  Zwraca True przy sukcesie, False jesli COM niedostepny.
    """
    try:
        import win32com.client
    except ImportError:
        print('  [Excel COM] Brak pywin32 (win32com) — nie moge wstawic blokow.')
        return False

    xl = wb = None
    try:
        # DEDYKOWANA, nowa instancja Excela (DispatchEx) — NIE doczepiamy sie do
        # ewentualnego otwartego Excela uzytkownika (zajety/modalny -> wyjatek i tylko
        # 6 punktow w protokole). Fallback do Dispatch, gdyby DispatchEx zawiodl.
        try:
            xl = win32com.client.DispatchEx('Excel.Application')
        except Exception:
            xl = win32com.client.Dispatch('Excel.Application')
        xl.Visible         = False
        xl.DisplayAlerts   = False
        xl.ScreenUpdating  = False
        try:
            xl.AskToUpdateLinks = False
            xl.EnableEvents     = False
        except Exception:
            pass

        # UpdateLinks=0: nie aktualizuj linkow zewn. przy otwarciu (unika siegania do sieci).
        wb  = xl.Workbooks.Open(os.path.abspath(filepath), UpdateLinks=0, ReadOnly=False)
        ws  = None
        for idx in range(1, wb.Sheets.Count + 1):
            if wb.Sheets(idx).Name == sheet_name:
                ws = wb.Sheets(idx)
                break
        if ws is None:
            print(f'  [Excel COM] Brak arkusza "{sheet_name}" — nie wstawiam blokow.')
            return False

        addr = f'{block_start}:{block_start + block_size - 1}'
        for _ in range(n_extra):
            ws.Rows(addr).Copy()
            ws.Rows(addr).Insert(Shift=-4121)   # xlShiftDown
            xl.CutCopyMode = False

        wb.Save()
        return True

    except Exception as exc:
        print(f'  [Excel COM] Blad wstawiania blokow: {type(exc).__name__}: {exc}')
        return False
    finally:
        if wb is not None:
            try: wb.Close(False)
            except: pass
        if xl is not None:
            try: xl.Quit()
            except: pass


def _rozszerz_kolumny_przyrzadow(filepath, sheet_name, start_col, potrzebne):
    """
    Dokłada PARY kolumn przyrzadow w Stronie 3, gdy przyrzadow jest wiecej niz w szablonie
    (szablon CC ma 10 par: Q/R … AI/AJ). Kopiuje OSTATNIA pare (np. AI:AJ) i wstawia ja
    z przesunieciem w PRAWO — tak jak robi to operator recznie, wiec zachowane sa scalenia,
    formatowanie, obramowania i naglowek 'Przyrzady wzorcowane' (rozciaga sie automatycznie).

    Po wstawieniu ustawia w kazdej nowej parze:
      - wiersz 10: numer przyrzadu (11, 12, ...),
      - wiersz 12: formule ='Strona 2'!$E$<kolejny wiersz przyrzadu> (wzor z ostatniej pary).
    Formuly w wierszu 18 (=CONCATENATE("t",AI10)) sa relatywne — przeliczaja sie same.

    Zwraca liczbe dodanych par (0 gdy nie bylo potrzeby lub COM niedostepny).
    """
    if potrzebne <= 0:
        return 0
    try:
        import win32com.client
    except ImportError:
        print('  [Excel COM] Brak pywin32 — nie moge dolozyc kolumn przyrzadow.')
        return 0

    xl = wb = None
    try:
        try:
            xl = win32com.client.DispatchEx('Excel.Application')
        except Exception:
            xl = win32com.client.Dispatch('Excel.Application')
        xl.Visible = False
        xl.DisplayAlerts = False
        xl.ScreenUpdating = False
        try:
            xl.AskToUpdateLinks = False
            xl.EnableEvents = False
        except Exception:
            pass

        wb = xl.Workbooks.Open(os.path.abspath(filepath), UpdateLinks=0, ReadOnly=False)
        ws = None
        for idx in range(1, wb.Sheets.Count + 1):
            if wb.Sheets(idx).Name == sheet_name:
                ws = wb.Sheets(idx)
                break
        if ws is None:
            print(f'  [Excel COM] Brak arkusza "{sheet_name}" — nie dokladam kolumn.')
            return 0

        # Ile par jest teraz? (wiersz 10: numery 1,2,3... co 2 kolumny od start_col)
        obecne = 0
        c = start_col
        while True:
            v = ws.Cells(10, c).Value
            if isinstance(v, (int, float)) and int(v) == obecne + 1:
                obecne += 1
                c += 2
            else:
                break
        if obecne == 0:
            print('  [Excel COM] Nie rozpoznano par kolumn przyrzadow — nie dokladam.')
            return 0

        # 1) Dokladanie brakujacych par: kopiuj OSTATNIA pare i wstaw w prawo.
        dodane = 0
        c_last = start_col + (obecne - 1) * 2
        for nr in range(obecne + 1, potrzebne + 1):
            src = f"{get_column_letter(c_last)}:{get_column_letter(c_last + 1)}"
            dst = f"{get_column_letter(c_last + 2)}:{get_column_letter(c_last + 3)}"
            ws.Columns(src).Copy()
            ws.Columns(dst).Insert(Shift=-4161)      # xlShiftToRight
            xl.CutCopyMode = False
            c_new = c_last + 2
            ws.Cells(10, c_new).Value = nr           # numer przyrzadu w naglowku
            c_last = c_new
            dodane += 1

        # 2) Naglowek pary (wiersz 12) = nr fabryczny przyrzadu z tabeli na Stronie 2.
        #    Szablon ma te komorki puste, wiec wpisujemy je dla WSZYSTKICH uzywanych par
        #    (par nadmiarowych nie ruszamy, zeby nie pokazywac zer przy pustych przyrzadach).
        for nr in range(1, potrzebne + 1):
            cc = start_col + (nr - 1) * 2
            ws.Cells(12, cc).Formula = f"='Strona 2'!$E${WIERSZ_1_PRZYRZADU_S2 + nr - 1}"

        wb.Save()
        if dodane:
            print(f"  Dolozono {dodane} par kolumn przyrzadow ({obecne} -> {potrzebne}); "
                  f"ostatnia para: {get_column_letter(c_last)}/{get_column_letter(c_last + 1)}.")
        print(f"  Naglowki przyrzadow (wiersz 12) ustawione dla {potrzebne} par "
              f"(= nr fabryczne ze Strony 2).")
        return dodane

    except Exception as exc:
        print(f'  [Excel COM] Blad dokladania kolumn przyrzadow: {type(exc).__name__}: {exc}')
        return 0
    finally:
        if wb is not None:
            try: wb.Close(False)
            except: pass
        if xl is not None:
            try: xl.Quit()
            except: pass


def _ustaw_komorke(ws, row, col, value):
    """
    Wpisuje wartosc, uwzgledniajac SCALENIA: gdy (row,col) lezy w scalonym zakresie,
    zapis idzie do jego lewego-gornego rogu (openpyxl nie pozwala pisac do MergedCell).
    """
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            row, col = rng.min_row, rng.min_col
            break
    ws.cell(row=row, column=col).value = value
    return row, col


def _kolumna_roku_s3(ws3, wiersz=5):
    """
    Kolumna na ROK w naglowku Strony 3 = pierwsza kolumna ZA etykieta '/LA/TH/'.
    Szablony maja rozny uklad (CC: O5 -> rok w P5; CC-04: Q5 -> rok w R5), wiec
    szukamy po etykiecie zamiast wpisywac stala kolumne.
    """
    for c in range(1, (ws3.max_column or 40) + 1):
        v = ws3.cell(row=wiersz, column=c).value
        if isinstance(v, str) and 'LA/TH' in v.replace(' ', '').upper():
            koniec = c
            for rng in ws3.merged_cells.ranges:
                if rng.min_row <= wiersz <= rng.max_row and rng.min_col <= c <= rng.max_col:
                    koniec = rng.max_col
                    break
            return koniec + 1
    return None


def _zapisz_bezpiecznie(wb, path, opis="plik"):
    """
    Zapisuje skoroszyt; gdy plik jest OTWARTY w Excelu (zablokowany) — daje czytelny
    komunikat zamiast surowego PermissionError. To najczestsza przyczyna „braku danych":
    plik otwarty w Excelu w trakcie pracy skryptu przerywa zapis.
    """
    try:
        wb.save(path)
    except PermissionError:
        raise PermissionError(
            f"\n  Nie moge zapisac ({opis}):\n    {path}\n"
            f"  >>> Plik jest OTWARTY w Excelu. ZAMKNIJ go i uruchom skrypt ponownie. <<<\n"
            f"  (Otwarcie pliku w trakcie pracy skryptu przerywa zapis danych — stad „pusty\" protokol.)"
        )


def _round_hm(dt):
    """Zaokrągla datetime do pelnej minuty i zwraca string HH:MM."""
    if dt is None:
        return None
    if dt.second >= 30:
        dt = dt + datetime.timedelta(minutes=1)
    return dt.strftime('%H:%M')


def generuj_protokol(rep_groups, rows, measurement_id, obs_type, sensor_names=None,
                     pz_mapa=None, zest=None, pz_lista=None):
    """
    Tworzy plik protokolu na podstawie reprezentacyjnych wierszy z obserwacji.

    rep_groups : lista krotek (list[5 int], powod) – indeksy reprezentantow (0-based
                 do `rows`) dla kazdego punktu; powod=None => zielony (OK),
                 powod=tekst => pomaranczowy (kryterium niespelnione, wpisany mimo to)
    rows       : surowe wiersze danych z pliku TXT (list of list[str])
    obs_type   : 'CC' lub 'CC04'
    """
    TEMPLATE_BLOCKS = 6     # liczba blokow (punktow) w szablonie
    BLOCK_START_ROW = 20    # pierwszy wiersz Excel pierwszego bloku
    BLOCK_SIZE      = 5     # wierszy na jeden punkt

    if obs_type == 'CC04':
        tmpl_name   = PROTOKOL_CC04_TEMPLATE
        # Do protokolu ida TYLKO kanaly GLOWNE — indeksy odczytow ChNNN + tdp w rows[]
        # (dobierane po nazwie z pelnego ukladu CC04_KOLUMNY). K,L,M,N,O = 4 kanaly + tdp.
        _idx        = {name: i for i, name in enumerate(CC04_KOLUMNY)}
        src_indices = [_idx[f'Ch{ch}'] for ch in CC04_KANALY_GLOWNE] + [_idx['tdp']]
        dst_cols    = [11, 12, 13, 14, 15]  # K,L,M,N,O (1-based w Excel)
    else:
        tmpl_name   = PROTOKOL_CC_TEMPLATE
        src_indices = [5, 6]                # F,G
        dst_cols    = [12, 13]              # L,M

    tmpl_path = os.path.join(FOLDER, tmpl_name)
    out_name  = tmpl_name.replace('xxx', measurement_id, 1)
    out_path  = os.path.join(FOLDER, out_name)

    if not os.path.exists(tmpl_path):
        print(f"\n  PROTOKOL: Brak szablonu '{tmpl_name}' — pomijam generowanie protokolu.")
        return

    N = len(rep_groups)
    if N == 0:
        print("  PROTOKOL: Brak punktow pomiarowych — pomijam.")
        return

    shutil.copy2(tmpl_path, out_path)
    print(f"\nSkopiowano szablon protokolu: {tmpl_name}")

    # ── Dopasowanie liczby bloków przez Excel COM (przed zaladowaniem openpyxl) ─
    if N > TEMPLATE_BLOCKS:
        extra = N - TEMPLATE_BLOCKS
        print(f"  Wstawiam {extra} dodatkowych blokow przez Excel COM...")
        ok = _insert_blocks_via_excel(out_path, 'Strona 3', BLOCK_START_ROW, BLOCK_SIZE, extra)
        if ok:
            print(f"  Wstawiono {extra} blokow.")
        else:
            print(f"  UWAGA: Excel COM niedostepny (pip install pywin32). "
                  f"Ograniezam do {TEMPLATE_BLOCKS} punktow.")
            N          = TEMPLATE_BLOCKS
            rep_groups = rep_groups[:N]

    elif N < TEMPLATE_BLOCKS:
        extra     = TEMPLATE_BLOCKS - N
        del_from  = BLOCK_START_ROW + N * BLOCK_SIZE
        del_count = extra * BLOCK_SIZE
        print(f"  Usuwam {extra} nadmiarowych blokow przez Excel COM...")
        ok = _delete_rows_via_excel(out_path, 'Strona 3', del_from, del_count)
        if ok:
            print(f"  Usunieto {extra} blokow.")
        else:
            print(f"  UWAGA: Excel COM niedostepny — nadmiarowe bloki pozostana w pliku.")

    # ── Dolozenie PAR KOLUMN przyrzadow, gdy przyrzadow jest wiecej niz w szablonie ──
    # (szablon ma 10 par Q/R…AI/AJ; przy 12 przyrzadach dokladamy AK/AL i AM/AN)
    _start_col = WYNIKI_START_COL_CC04 if obs_type == 'CC04' else WYNIKI_START_COL_CC
    _potrzebne = _policz_potrzebne_przyrzady(rep_groups, rows)
    if _potrzebne:
        _rozszerz_kolumny_przyrzadow(out_path, 'Strona 3', _start_col, _potrzebne)

    # ── Ładujemy plik (już z prawidłową liczbą bloków) z openpyxl ────────────
    proto_wb = openpyxl.load_workbook(out_path)

    if 'Strona 3' not in proto_wb.sheetnames:
        print(f"  PROTOKOL: Brak arkusza 'Strona 3' w '{tmpl_name}' — pomijam.")
        proto_wb.close()
        return
    ws3 = proto_wb['Strona 3']

    # ── Naglowek: P5 = rok wzorcowania, J6 = numer protokolu ─────────────────
    # Rok bierzemy z daty pomiaru (a nie z zegara systemowego) — protokol dotyczy
    # konkretnego wzorcowania; gdy brak daty, awaryjnie rok biezacy.
    _rok = None
    _dt_pierwszy = _s_to_dt(rows[rep_groups[0][0][0]][0]) if rep_groups else None
    if _dt_pierwszy is not None:
        _rok = _dt_pierwszy.year
    _rok = _rok or datetime.date.today().year
    _kol_rok = _kolumna_roku_s3(ws3)          # CC -> P5, CC-04 -> R5
    if _kol_rok:
        _w, _k = _ustaw_komorke(ws3, 5, _kol_rok, _rok)
        print(f"  Naglowek: rok = {_rok} ({get_column_letter(_k)}{_w}), nr protokolu = 1 (J6)")
    else:
        print(f"  [UWAGA] Nie znalazlem etykiety '/LA/TH/' w wierszu 5 — rok niewpisany.")
    _ustaw_komorke(ws3, 6, 10, 1)             # J6 (nr protokolu)

    # ── Zapis danych dla każdego punktu ───────────────────────────────────────
    FILL_GREY = PatternFill(fill_type='solid', fgColor='BFBFBF')

    for punkt_idx, (rep_indices, powod, *_reszta) in enumerate(rep_groups):
        r0 = BLOCK_START_ROW + punkt_idx * BLOCK_SIZE

        # Numer punktu w kolumnie A (A20=1, A25=2, ...). Wstawianie blokow kopiuje
        # szablonowy „1" do kazdego bloku — nadpisujemy poprawnym numerem kolejnym.
        ws3.cell(row=r0, column=1).value = punkt_idx + 1

        first_i   = rep_indices[0]
        last_i    = rep_indices[-1]
        row_first = rows[first_i]

        # Dane pomiarowe (F,G[,H,I,J] z obserwacji) → kolumny docelowe protokolu
        # Jesli wartosc to "brak" — komorke zostawiamy pusta i zaznaczamy szarym
        for row_off, data_i in enumerate(rep_indices):
            r = rows[data_i]
            for src_ci, dst_col in zip(src_indices, dst_cols):
                raw  = r[src_ci] if src_ci < len(r) else ''
                cell = ws3.cell(row=r0 + row_off, column=dst_col)
                if str(raw).strip().lower() == 'brak':
                    cell.value = None
                    cell.fill  = FILL_GREY
                else:
                    cell.value = to_value(raw) if raw else None

        # D i E z pierwszego wiersza punktu → I(r0) i J(r0)
        ws3.cell(row=r0, column=9).value  = (to_value(row_first[3])
                                              if len(row_first) > 3 else None)
        ws3.cell(row=r0, column=10).value = (to_value(row_first[4])
                                              if len(row_first) > 4 else None)

        # B (Tzadana) i C (RHzadana) z pierwszego wiersza → B(r0) i C(r0)
        rh_val  = _s_to_float(row_first[2]) if len(row_first) > 2 else None
        rh_zero = (rh_val is not None and rh_val == 0.0)

        ws3.cell(row=r0, column=2).value = (to_value(row_first[1])
                                             if len(row_first) > 1 else None)
        c_cell = ws3.cell(row=r0, column=3)
        if rh_zero:
            c_cell.value = '-'
            c_cell.fill  = FILL_GREY
        else:
            c_cell.value = rh_val

        # Jesli RH=0 (tryb tylko temperatura): kolumna O na szaro + "-" w kolumnach bocznych
        if rh_zero:
            for row_off in range(BLOCK_SIZE):
                ws3.cell(row=r0 + row_off, column=15).fill = FILL_GREY  # O

            # Kolumny WILGOTNOSCI przyrzadow (druga kolumna kazdej pary: R, T, V, ...) —
            # w punkcie tylko-temperaturowym nie ma odczytu RH, wiec zaznaczamy je na szaro
            # (nie zostawiamy pustych bialych komorek).
            _par = _potrzebne or 0
            for _k in range(_par):
                _col_rh = _start_col + _k * 2 + 1
                for row_off in range(BLOCK_SIZE):
                    ws3.cell(row=r0 + row_off, column=_col_rh).fill = FILL_GREY
            if obs_type == 'CC04':
                ws3.cell(row=r0, column=17).value = '-'   # Q
                ws3.cell(row=r0, column=18).value = '-'   # R
            else:
                ws3.cell(row=r0, column=15).value = '-'   # O  (CC)
                ws3.cell(row=r0, column=16).value = '-'   # P  (CC)

        # Data i czasy z kolumny A obserwacji
        start_dt = _s_to_dt(row_first[0])
        end_dt   = _s_to_dt(rows[last_i][0])

        if start_dt:
            ws3.cell(row=r0 + 1, column=5).value = start_dt.strftime('%d.%m.%Y')
            ws3.cell(row=r0 + 2, column=5).value = _round_hm(start_dt)
        if end_dt:
            ws3.cell(row=r0 + 3, column=5).value = _round_hm(end_dt)

        # Punkt NIE przeszedl kryterium (w obserwacji pomaranczowy) — oznacz TAK SAMO
        # w protokole: kolumna B (nastawa T) calego bloku na pomaranczowo + nota z powodem
        # w kolumnie poza obszarem druku (widoczna operatorowi w Excelu).
        if powod:
            for row_off in range(BLOCK_SIZE):
                ws3.cell(row=r0 + row_off, column=2).fill = FILL_WARN_DARK
            nota = ws3.cell(row=r0, column=20)
            nota.value = f"UWAGA (nie na zielono): {powod}"
            fo = nota.font
            nota.font = Font(name=fo.name, size=fo.size, bold=True, italic=True, color='9C5700')
            print(f"    [PROTOKOL] Punkt {punkt_idx+1} oznaczony na pomaranczowo: {powod}")

    # ── Podmiana Pt100-XX w naglowkach na Stronie 3 ──────────────────────────
    def _subst(cell, sname):
        """Podmienia Pt100-XX w wartosci komorki lub wpisuje nazwe jesli brak wzorca."""
        if not sname:
            return
        if cell.value and re.search(r'Pt100-\d+', str(cell.value)):
            cell.value = re.sub(r'Pt100-\d+', sname, str(cell.value))
        else:
            cell.value = sname

    if sensor_names:
        if obs_type == 'CC04':
            # K9, L9, M9, N9 → czujniki 1-4
            for col_off, sname in enumerate(sensor_names[:4]):
                _subst(ws3.cell(row=9, column=11 + col_off), sname)
        else:
            # CC: L9 → jeden czujnik
            _subst(ws3.cell(row=9, column=12), sensor_names[0])

    # ── Strona 1: F10 = "1-N" oraz czujniki wzorcowe ─────────────────────────
    if 'Strona 1' in proto_wb.sheetnames:
        ws1 = proto_wb['Strona 1']
        ws1.cell(row=10, column=6).value = f"1-{N}"

        if sensor_names:
            if obs_type == 'CC04':
                # G:H36-39 → czujniki 1-4  (komorki scalone G:H, piszemy do G=7)
                for row_off, sname in enumerate(sensor_names[:4]):
                    _subst(ws1.cell(row=36 + row_off, column=7), sname)
            else:
                # CC: H:I43 → jeden czujnik  (piszemy do H=8)
                _subst(ws1.cell(row=43, column=8), sensor_names[0])

    # ── Dane srodowiskowe z wynikow (wyniki/*.xlsx → Q/S kolumny Strona 3) ──────
    print("\n  Szukam danych srodowiskowych w wynikach...")
    uzyte = _wypelnij_wyniki_srodowiskowe(ws3, rep_groups, rows, obs_type)

    # Przyrzady BEZ pliku logera (np. reczny termohigrometr odczytywany multimetrem)
    # nie maja jak trafic do 'uzyte' — wtedy tabele przyrzadow budujemy WPROST z PZ:
    # bierzemy pozycje wzorcowane w komorze klimatycznej, w kolejnosci z PZ.
    if not uzyte and pz_lista:
        widziane, z_pz = set(), []
        for p in pz_lista:
            if not getattr(p, 'komora', False) or not p.nr_fabr:
                continue
            klucz = pz_dane.normalizuj_serial(p.nr_fabr)
            if klucz in widziane:
                continue          # ten sam przyrzad na kilku pozycjach — raz wystarczy
            widziane.add(klucz)
            z_pz.append((p.nr_fabr, [], []))
        if z_pz:
            uzyte = z_pz
            print(f"  [PZ] Brak plikow wynikow — tabele przyrzadow buduje z PZ "
                  f"({len(uzyte)} przyrzad(ow) z komory klimatycznej).")

    # Odczyty przyrzadu w punktach, ktorych NIE zamowiono wlasnie dla niego, sa zbedne —
    # protokol zawiera SUME punktow wszystkich zlecen, a kazdy przyrzad ma swoj zakres
    # (PZ: pozycja przyrzadu -> jego punkty). Takie komorki wyszarzamy.
    _wyszarz_punkty_spoza_zamowienia(ws3, rep_groups, uzyte, obs_type,
                                     BLOCK_START_ROW, BLOCK_SIZE, FILL_GREY)

    # Tabela przyrzadow (Strona 2) z PZ — dopasowanie po nr fabrycznym do kolumn Strony 3.
    if 'Strona 2' in proto_wb.sheetnames:
        wypelnij_strone2_z_pz(proto_wb['Strona 2'], uzyte, pz_mapa, zest)
        # Data pomiaru = dzien OSTATNIEGO punktu (wtedy pomiary zostaly zakonczone).
        _dt_ost = _s_to_dt(rows[rep_groups[-1][0][-1]][0]) if rep_groups else None
        wypelnij_podpisy_strona2(proto_wb['Strona 2'],
                                 _dt_ost.date() if _dt_ost else None)
    else:
        print("  [PZ] Brak arkusza 'Strona 2' — pomijam tabele przyrzadow.")

    _zapisz_bezpiecznie(proto_wb, out_path, "protokol")
    print(f"Zapisano protokol: {out_name}")


def oznacz_zestawienie_punkty(rep_groups, rows):
    """
    W pliku wyniki/zestawienie_pomiarow.xlsx (wszystkie przyrzady na wspolnej osi czasu)
    zaznacza okna WYBRANYCH punktow: koloruje wiersze mieszczace sie w oknie punktu i
    wpisuje NUMER punktu (ten sam co w protokole) w kolumnie z prawej — do szybkiej
    nawigacji w Excelu (Ctrl+strzalka w dol / wyszukanie numeru).
    Dopasowanie po kolumnie 'Czas' do okna [pierwszy..ostatni reprezentant] punktu.
    """
    zest_path = os.path.join(WYNIKI_FOLDER, "zestawienie_pomiarow.xlsx")
    if not os.path.exists(zest_path):
        print(f"  [Zestawienie] Brak {os.path.basename(zest_path)} — pomijam znaczniki punktow.")
        return
    try:
        wb = openpyxl.load_workbook(zest_path)
    except Exception as e:
        print(f"  [Zestawienie] Nie moge otworzyc pliku: {type(e).__name__}: {e}")
        return
    ws = wb.active
    n_col = ws.max_column

    # (czas, wiersz) z kolumny 1 ('Czas')
    czas_rows = []
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        dt = v if isinstance(v, datetime.datetime) else (_s_to_dt(str(v)) if v else None)
        if dt is not None:
            czas_rows.append((dt, r))
    if not czas_rows:
        print("  [Zestawienie] Nie znaleziono kolumny czasu — pomijam.")
        return

    marker_col = n_col + 2            # kolumna na numer punktu (1 kolumna odstepu)
    ws.cell(row=1, column=marker_col).value = "Nr punktu"
    czasy_zest = [dt for dt, _ in czas_rows]
    oznaczone = 0
    for i, (rep_idx, powod, *_) in enumerate(rep_groups, 1):
        # Dla KAZDEGO z 5 wierszy punktu bierzemy NAJBLIZSZY czasowo wiersz zestawienia.
        # (Dopasowanie „po zakresie czasu" dawalo mniej wierszy niz 5, bo siatka czasu
        # zestawienia jest przesunieta wzgledem probek multimetru.)
        matched = []
        for idx_r in rep_idx:
            cel = _s_to_dt(rows[idx_r][0])
            if cel is None:
                continue
            poz = bisect.bisect_left(czasy_zest, cel)
            kand = []
            if poz > 0:
                kand.append(czas_rows[poz - 1])
            if poz < len(czas_rows):
                kand.append(czas_rows[poz])
            if not kand:
                continue
            _dt, r = min(kand, key=lambda x: abs(x[0] - cel))
            if r not in matched:
                matched.append(r)
        if not matched:
            continue
        matched.sort()
        fill_l = FILL_WARN_LIGHT if powod else FILL_LIGHT
        fill_d = FILL_WARN_DARK  if powod else FILL_DARK
        for r in matched:
            for c in range(1, n_col + 1):
                ws.cell(row=r, column=c).fill = fill_l
        mcell = ws.cell(row=matched[0], column=marker_col)
        mcell.value = i
        mcell.fill  = fill_d
        fo = mcell.font
        mcell.font = Font(name=fo.name, size=fo.size, bold=True)
        oznaczone += 1

    _zapisz_bezpiecznie(wb, zest_path, "zestawienie")
    print(f"  [Zestawienie] Oznaczono {oznaczone}/{len(rep_groups)} punktow "
          f"({os.path.basename(zest_path)}).")


# =============================================================================
# ZDJECIA PUNKTOW (foto/)
# =============================================================================

_FOTO_WZORCE_CZASU = (
    # 2026-07-23_16.12.51 / 2026-07-23 16.12.51 / 2026-07-23T16:12:51
    re.compile(r'(\d{4})-(\d{2})-(\d{2})[ _T]+(\d{2})[.:_-](\d{2})[.:_-](\d{2})'),
    # 20260723_161251
    re.compile(r'(\d{4})(\d{2})(\d{2})[ _T-]+(\d{2})(\d{2})(\d{2})'),
)


def _czas_z_nazwy_foto(nazwa):
    """Czas zdjecia z NAZWY pliku (np. '2026-07-23_16.12.51.jpg'). None gdy brak."""
    for wz in _FOTO_WZORCE_CZASU:
        m = wz.search(nazwa)
        if m:
            try:
                y, mo, d, h, mi, s = (int(g) for g in m.groups())
                return datetime.datetime(y, mo, d, h, mi, s)
            except ValueError:
                continue
    return None


def _indeks_zdjec(folder):
    """
    Indeks zdjec: posortowana lista (czas, sciezka). Czas z nazwy pliku, a gdy nazwa
    go nie zawiera — z daty modyfikacji pliku (mtime).
    """
    idx = []
    bez_czasu_w_nazwie = 0
    try:
        pliki = os.listdir(folder)
    except OSError as e:
        print(f"  [FOTO] Nie moge odczytac folderu zrodlowego:\n    {folder}\n    {e}")
        return []
    for nazwa in pliki:
        if not nazwa.lower().endswith(FOTO_ROZSZERZENIA):
            continue
        sciezka = os.path.join(folder, nazwa)
        dt = _czas_z_nazwy_foto(nazwa)
        if dt is None:
            try:
                dt = datetime.datetime.fromtimestamp(os.path.getmtime(sciezka))
                bez_czasu_w_nazwie += 1
            except OSError:
                continue
        idx.append((dt, sciezka))
    idx.sort(key=lambda x: x[0])
    if bez_czasu_w_nazwie:
        print(f"  [FOTO] {bez_czasu_w_nazwie} plikow bez czasu w nazwie — uzyto daty modyfikacji.")
    return idx


def _najblizsze_zdjecie(idx, czasy, cel):
    """Zwraca (czas, sciezka, odchylka) zdjecia najblizszego czasowi `cel`."""
    poz = bisect.bisect_left(czasy, cel)
    kandydaci = []
    if poz > 0:
        kandydaci.append(idx[poz - 1])
    if poz < len(idx):
        kandydaci.append(idx[poz])
    if not kandydaci:
        return None
    dt, sciezka = min(kandydaci, key=lambda x: abs(x[0] - cel))
    return dt, sciezka, abs(dt - cel)


def _nazwa_folderu_punktu(nr, row_first):
    """Nazwa podfolderu punktu, np. 'punkt_01_50,0C' albo 'punkt_07_23,0C_60RH'."""
    t = _s_to_float(row_first[1]) if len(row_first) > 1 else None
    rh = _s_to_float(row_first[2]) if len(row_first) > 2 else None
    czesci = [f"punkt_{nr:02d}"]
    if t is not None:
        czesci.append(f"{t:.1f}C".replace('.', ','))
    if rh is not None and rh > 0:
        czesci.append(f"{rh:.0f}RH")
    nazwa = "_".join(czesci)
    return re.sub(r'[\\/:*?"<>|]', '_', nazwa)


def kopiuj_foto_punktow(rep_groups, rows):
    """
    Kopiuje zdjecia odpowiadajace WYBRANYM punktom: dla kazdego z 5 wierszy
    reprezentacyjnych punktu szuka zdjecia najblizszego czasowo i kopiuje je do
    FOTO_FOLDER/punkt_NN_<nastawa>/ (oryginalne nazwy plikow).
    Zdjecia dalej niz FOTO_TOLERANCJA od wiersza sa pomijane (z ostrzezeniem).
    """
    print(f"\nZdjecia punktow (foto):")
    print(f"  Zrodlo : {FOTO_ZRODLO}")
    print(f"  Cel    : {FOTO_FOLDER}")
    if not os.path.isdir(FOTO_ZRODLO):
        print(f"  [FOTO] Folder zrodlowy nie istnieje — pomijam kopiowanie.")
        return

    idx = _indeks_zdjec(FOTO_ZRODLO)
    if not idx:
        print("  [FOTO] Brak zdjec w folderze zrodlowym — pomijam.")
        return
    czasy = [dt for dt, _ in idx]
    print(f"  Znaleziono {len(idx)} zdjec ({czasy[0]} … {czasy[-1]}).")

    os.makedirs(FOTO_FOLDER, exist_ok=True)
    razem = pominiete = 0
    max_odchylka = datetime.timedelta(0)

    for nr, (rep_idx, _powod, *_) in enumerate(rep_groups, 1):
        row_first = rows[rep_idx[0]]
        pod = os.path.join(FOTO_FOLDER, _nazwa_folderu_punktu(nr, row_first))
        skopiowane = set()
        braki = 0
        for i in rep_idx:
            cel = _s_to_dt(rows[i][0])
            if cel is None:
                continue
            trafienie = _najblizsze_zdjecie(idx, czasy, cel)
            if trafienie is None:
                braki += 1
                continue
            _dt, sciezka, odchylka = trafienie
            if odchylka > FOTO_TOLERANCJA:
                braki += 1
                continue
            if sciezka in skopiowane:
                continue            # dwa wiersze trafily w to samo zdjecie
            os.makedirs(pod, exist_ok=True)
            try:
                shutil.copy2(sciezka, os.path.join(pod, os.path.basename(sciezka)))
            except OSError as e:
                print(f"    [FOTO] Nie moge skopiowac {os.path.basename(sciezka)}: {e}")
                continue
            skopiowane.add(sciezka)
            max_odchylka = max(max_odchylka, odchylka)
        razem += len(skopiowane)
        pominiete += braki
        status = f"{len(skopiowane)} zdjec"
        if braki:
            status += f"  [UWAGA] {braki} wierszy bez zdjecia w tolerancji {FOTO_TOLERANCJA}"
        print(f"  punkt {nr:>2}: {status}  -> {os.path.basename(pod)}")

    print(f"  Skopiowano lacznie {razem} zdjec dla {len(rep_groups)} punktow "
          f"(max odchylka czasu: {max_odchylka}).")
    if pominiete:
        print(f"  [FOTO] {pominiete} wierszy bez pasujacego zdjecia — sprawdz zgodnosc "
              f"zegara aparatu z multimetrem albo zwieksz FOTO_TOLERANCJA.")


# =============================================================================
# MAIN
# =============================================================================

def main():
    txt_files = resolve_txt_files()
    first_name = os.path.basename(txt_files[0])
    measurement_id = parse_measurement_id(first_name)

    # Typ pliku i naglowek wyznaczamy z pierwszego pliku (przerwany pomiar to
    # ten sam typ i ten sam czujnik we wszystkich czesciach).
    raw_lines = open_txt(txt_files[0])
    file_type = detect_file_type(raw_lines)

    # Dane przyrzadow z PZ wczytujemy NA STARCIE (przed analiza), zeby byly gotowe
    # do wypelnienia Strony 2 przy budowie protokolu.
    print("Wczytywanie danych przyrzadow (PZ + Zestawienie)...")
    pz_mapa, _pz_lista = pz_dane.wczytaj_pz(PZ_FOLDER)
    # Punkty ZAMOWIONE w PZ ('Zakres wzorcowania') — decyduja, ktore segmenty obserwacji
    # trafia do protokolu (jeden wsad komory obsluguje czesto kilka zlecen).
    punkty_pz = pz_dane.wczytaj_punkty(PZ_FOLDER) if WYBIERAJ_PUNKTY_WG_PZ else []
    zest = pz_dane.wczytaj_zestawienie(ZESTAWIENIE_PLIK)

    print(f"Numer pomiaru : {measurement_id}")
    if len(txt_files) == 1:
        print(f"Plik wejsciowy: {first_name}")
    else:
        print(f"Pliki wejsciowe ({len(txt_files)}):")
        for p in txt_files:
            print(f"    • {os.path.basename(p)}")
    print(f"Typ pliku     : {file_type}")

    today        = datetime.date.today()
    sensor_names = None   # wypelniane w galezi CC04

    if file_type == 'CC04':
        # ── CC-04 ──────────────────────────────────────────────────────────────
        template_path = os.path.join(FOLDER, CC04_TEMPLATE)
        output_name   = CC04_TEMPLATE.replace('szablon', measurement_id, 1)
        output_path   = os.path.join(FOLDER, output_name)

        print(f"Szablon       : {CC04_TEMPLATE}")
        print(f"Plik wyjsciowy: {output_name}")

        if not os.path.exists(template_path):
            raise FileNotFoundError(f"Nie znaleziono szablonu: {template_path}")
        shutil.copy2(template_path, output_path)
        print("Skopiowano szablon.")

        sensor_names, rows = combine_txt(txt_files, parse_txt_cc04)
        print(f"Czujniki wzorcowe: {sensor_names}")
        print(f"Wierszy danych   : {len(rows)}")

        wb = openpyxl.load_workbook(output_path)
        ws = wb.active

        # PELNY uklad jak w pliku multimetru: WSZYSTKIE kanaly po kolei (A..AG = 33 kol).
        # Analiza i protokol biora tylko kanaly GLOWNE (po nazwie); reszta jest do wgladu.
        N_KOL = len(CC04_KOLUMNY)
        kanal_pt = _mapa_kanal_pt(raw_lines)

        # Podpisy z szablonu (X92/X93) + ich STYL (ramka, format daty) czytamy PRZED
        # nadpisaniem danymi. W pelnym ukladzie kol. 24-25 to juz dane, wiec podpisy,
        # ramke i date przenosimy na PRAWO od danych, a stare komorki czyscimy (inaczej
        # zostaje pusta ramka i data pokazuje np. „19.02.1900" na miejscu danych).
        sig1 = ws.cell(row=92, column=24).value
        sig2 = ws.cell(row=93, column=24).value
        sig_border = _copy_obj(ws.cell(row=92, column=24).border)
        date_nf    = ws.cell(row=92, column=25).number_format or 'yyyy-mm-dd'

        # Kolumny nieuzytych kanalow (pomiar np. na 2 czujnikach zamiast 4) nie
        # trafiaja do arkusza — inaczej zostawaly puste, z samym naglowkiem.
        naglowki = _naglowki_cc04(kanal_pt)
        # Formaty liczbowe czytamy PRZED zapisem — potem kolumny sa nadpisane.
        formaty_szablonu = _formaty_kolumn_szablonu(ws, N_KOL)
        kolumny = _kolumny_z_danymi(rows, N_KOL) if POMIJAJ_PUSTE_KOLUMNY \
            else list(range(N_KOL))
        mapa_kolumn = _raport_pominietych_kolumn(kolumny, naglowki, N_KOL)

        # Dane pomiarowe — tylko kolumny, w ktorych cokolwiek jest
        for r_i, row in enumerate(rows):
            for nowy, stary in enumerate(kolumny):
                if stary < len(row):
                    ws.cell(row=2 + r_i, column=1 + nowy).value = to_value(row[stary])

        # Naglowki (rzad 1) — z nazwami czujnikow (Pt100-XX) zamiast surowych 'ChNNN'
        for nowy, stary in enumerate(kolumny):
            if stary < len(naglowki):
                ws.cell(row=1, column=1 + nowy).value = naglowki[stary]
        print(f"  Czujniki glowne : {sensor_names}")
        _przenies_formaty_kolumn(ws, kolumny, formaty_szablonu, len(rows))
        print(f"  Czujniki zapasowe: {[kanal_pt.get(c) or c for c in CC04_KANALY_ZAPASOWE]}")

        # Osierocone naglowki i stary blok podpisow z szablonu — na prawo od danych.
        _wyczysc_pozostalosci_szablonu(
            ws, len(kolumny) + 1, max(N_KOL, ws.max_column),
            wiersze_podpisow=((92, 24), (93, 24), (92, 25), (93, 25)))

        # Podpisy + data PRZENIESIONE na prawo od danych (z ramka jak w oryginale)
        SIG_COL = len(kolumny) + 3
        for i, nazwisko in enumerate((sig1, sig2)):
            cn = ws.cell(row=92 + i, column=SIG_COL)
            cn.value  = nazwisko
            cn.border = _copy_obj(sig_border)
            cd = ws.cell(row=92 + i, column=SIG_COL + 1)
            cd.value  = today
            cd.border = _copy_obj(sig_border)
            cd.number_format = date_nf

        # Wyczysc STARE komorki podpisow (kol 24-25, w. 92-93) — teraz to zwykle dane:
        # bez ramki i bez formatu daty.
        _pusta_ramka = Border()
        for _r in (92, 93):
            for _col in (24, 25):
                _oc = ws.cell(row=_r, column=_col)
                _oc.border = _pusta_ramka
                _oc.number_format = 'General'

        print("\nAnaliza stabilnosci...")
        rep_groups = analyze_and_highlight(ws, rows, file_type='CC04', punkty_pz=punkty_pz)

    else:
        # ── CC (oryginalny) ────────────────────────────────────────────────────
        template_path = os.path.join(FOLDER, TEMPLATE)
        output_name   = TEMPLATE.replace('xxx', measurement_id, 1)
        output_path   = os.path.join(FOLDER, output_name)

        print(f"Szablon       : {TEMPLATE}")
        print(f"Plik wyjsciowy: {output_name}")

        if not os.path.exists(template_path):
            raise FileNotFoundError(f"Nie znaleziono szablonu: {template_path}")
        shutil.copy2(template_path, output_path)
        print("Skopiowano szablon.")

        sensor_name, rows = combine_txt(txt_files, parse_txt)
        sensor_names = [sensor_name]   # ujednolicamy z formatem CC04
        print(f"Czujnik wzorcowy: {sensor_name}")
        print(f"Wierszy danych  : {len(rows)}")

        wb = openpyxl.load_workbook(output_path)
        ws = wb.active

        # A2:L(ostatni) – dane pomiarowe, z pominieciem kolumn calkiem pustych
        n_kol_cc = max((len(r) for r in rows), default=0)
        kolumny = _kolumny_z_danymi(rows, n_kol_cc) if POMIJAJ_PUSTE_KOLUMNY \
            else list(range(n_kol_cc))
        naglowki_cc = [ws.cell(row=1, column=1 + i).value for i in range(n_kol_cc)]
        formaty_szablonu = _formaty_kolumn_szablonu(ws, n_kol_cc)
        mapa_kolumn = _raport_pominietych_kolumn(kolumny, naglowki_cc, n_kol_cc)
        for r_i, row in enumerate(rows):
            for nowy, stary in enumerate(kolumny):
                if stary < len(row):
                    ws.cell(row=2 + r_i, column=1 + nowy).value = to_value(row[stary])
        for nowy, stary in enumerate(kolumny):
            ws.cell(row=1, column=1 + nowy).value = naglowki_cc[stary]
        _wyczysc_pozostalosci_szablonu(ws, len(kolumny) + 1, n_kol_cc)
        _przenies_formaty_kolumn(ws, kolumny, formaty_szablonu, len(rows))

        # J1 – aktualizacja czujnika wzorcowego
        j1 = ws.cell(row=1, column=10)
        if j1.value and isinstance(j1.value, str):
            new_val = re.sub(r'Pt100-\d+', sensor_name, j1.value)
            if new_val == j1.value and sensor_name:
                print(f"  UWAGA: Nie znaleziono wzorca 'Pt100-XX' w J1: {j1.value!r}")
            j1.value = new_val
        else:
            j1.value = sensor_name
        print(f"J1 -> {j1.value!r}")

        # N92 – data,  O92 – podpis
        ws.cell(row=92, column=14).value = today
        ws.cell(row=92, column=15).value = PODPIS

        print("\nAnaliza stabilnosci...")
        rep_groups = analyze_and_highlight(ws, rows, file_type='CC', punkty_pz=punkty_pz)

    _zapisz_bezpiecznie(wb, output_path, "obserwacje")
    _przywroc_wykresy_z_szablonu(template_path, output_path, mapa_kolumn)
    print(f"\nZapisano: {output_name}")

    if rep_groups:
        generuj_protokol(rep_groups, rows, measurement_id, file_type,
                         sensor_names=sensor_names, pz_mapa=pz_mapa, zest=zest,
                         pz_lista=_pz_lista)
        # Znaczniki punktow w zestawieniu przyrzadow (nawigacja po czasie)
        oznacz_zestawienie_punkty(rep_groups, rows)
        # Zdjecia punktow (gdy wlaczone w konfiguracji)
        if KOPIUJ_FOTO:
            kopiuj_foto_punktow(rep_groups, rows)

    print("Gotowe!")


if __name__ == '__main__':
    main()
