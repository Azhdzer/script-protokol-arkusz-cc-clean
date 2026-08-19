# -*- coding: utf-8 -*-
"""
pz_dane.py — wspolny modul do wczytywania danych przyrzadow z:

  1) plikow PDF "Potwierdzenie zamowienia" (folder PZ/),
  2) pliku "Zestawienie wzorcowanych przyrzadow.xlsx" (arkusz Termohigrometry).

Uzywany przez generuj_obserwacje.py (wypelnianie Strony 2 protokolu) oraz
generuj_arkusze.py (fallback, gdy Strona 2 pusta). Modul jest CZYSTY — tylko
czyta i parsuje, bez efektow ubocznych.

Kluczowe pojecia:
  • nr_fabryczny (serial) — laczy przyrzad z PZ z kolumna pomiarowa Strony 3
    (pliki wynikow nazywaja sie <SERIAL>_wynik.xlsx).
  • rozdzielczosc t/RH — z Zestawienia (po producencie+typie); gdy brak,
    liczona z wahania cyfr po przecinku w danych pomiarowych.
"""

import os
import re
import glob

try:
    import logging
    from pypdf import PdfReader
    logging.getLogger('pypdf').setLevel(logging.ERROR)
    _PDF_OK = True
except ImportError:
    _PDF_OK = False

try:
    import openpyxl
    _XLSX_OK = True
except ImportError:
    _XLSX_OK = False


# =============================================================================
# MODEL DANYCH
# =============================================================================

class PZPrzyrzad:
    """Jeden przyrzad z PZ (po rozwinieciu 'N szt.' — dokladnie jeden nr fabryczny)."""
    __slots__ = ("nr_zlecenia", "wytworca", "typ", "nr_fabr", "nr_ewid",
                 "czuj_wytworca", "czuj_typ", "czuj_nr_fabr", "uzytkownik",
                 "zleceniodawca", "pozycja", "komora")

    def __init__(self, nr_zlecenia="", wytworca="", typ="", nr_fabr="", nr_ewid="",
                 czuj_wytworca="", czuj_typ="", czuj_nr_fabr="", uzytkownik="",
                 zleceniodawca="", pozycja=None, komora=False):
        self.nr_zlecenia = nr_zlecenia
        self.wytworca = wytworca
        self.typ = typ
        self.nr_fabr = nr_fabr
        self.nr_ewid = nr_ewid
        self.czuj_wytworca = czuj_wytworca
        self.czuj_typ = czuj_typ
        self.czuj_nr_fabr = czuj_nr_fabr
        self.uzytkownik = uzytkownik   # blok adresowy z pola 'UZYTKOWNIK:' (moze byc pusty)
        self.zleceniodawca = zleceniodawca  # blok adresowy z pola 'ZLECENIODAWCA:' 
        self.pozycja = pozycja         # numer pozycji w 'Obiekty wzorcowania' (1,2,3...)
        self.komora = komora           # True = wzorcowany w KOMORZE KLIMATYCZNEJ (ILAJ 5.4/11)

    def __repr__(self):
        return (f"PZPrzyrzad(zlec={self.nr_zlecenia}, wytw={self.wytworca!r}, "
                f"typ={self.typ!r}, fabr={self.nr_fabr!r}, ewid={self.nr_ewid!r}, "
                f"czuj={self.czuj_typ!r}/{self.czuj_nr_fabr!r})")


# =============================================================================
# NORMALIZACJA
# =============================================================================

def normalizuj_serial(s):
    """Klucz do dopasowania: bez spacji, wielkie litery, bez kropek/przecinkow na koncu."""
    if s is None:
        return ""
    return re.sub(r'\s+', '', str(s)).strip().strip('.,;').upper()


def _norm_txt(s):
    """Normalizacja nazw (producent/typ) do dopasowania rozmytego."""
    if s is None:
        return ""
    return re.sub(r'[\s\-_.]+', '', str(s)).strip().lower()


def _oczysc(s):
    """Porzadkuje wartosc do wpisania: zbedne spacje, koncowe znaki interpunkcyjne."""
    if s is None:
        return ""
    return re.sub(r'\s+', ' ', str(s)).strip().strip('.,;').strip()


# =============================================================================
# PARSER PDF (Potwierdzenie zamowienia)
# =============================================================================

# Etykiety pol wewnatrz opisu przyrzadu — DWUJEZYCZNE (PL / EN; PZ bywa po angielsku).
# Granica konca wartosci serialu = poczatek nastepnej etykiety.
# UWAGA na warianty zapisu numeru ewidencyjnego w PZ: 'nr wew.:', 'nr wewn.:', 'nr ewid.:',
# 'nr ewid .:'. Musza byc GRANICA konca numeru fabrycznego — inaczej tekst 'nr wew.: CL-1318A'
# wpada do wartosci 'nr fabr.' i po rozdzieleniu przecinkiem tworzy fikcyjny drugi przyrzad.
_END = (r'wytw[oó]rca|manufacturer|nr\s*wewn?|nr\s*ewid|identification|'
        r'typ\s*:|type\s*:|nr\s*kat|year\s+of\s+production|oraz\s+czujnika|adres\s*:')
_RE_TYP    = re.compile(r'(?:typ|type|nr\s*kat\.?)\s*:\s*([^,\n]+)', re.I)
# nr ewidencyjny: 'nr wew.:' / 'nr wewn.:' (starsze PZ) albo 'nr ewid.:' / 'nr ewid .:' (nowsze)
# Srednik konczy wartosc tak samo jak przecinek: w PZ z lista wypunktowana
# kazdy podpunkt konczy sie ';', a po ostatnim stoi jeszcze 'wytworca: ...'.
# Bez tej granicy nr ewidencyjny wychodzil jako 'UR00044; wytworca: Testo'.
_RE_WEWN   = re.compile(
    r'(?:nr\s*wewn?\s*\.?|nr\s*ewid\s*\.?|identification\s+number)\s*:\s*([^,;\n]+)', re.I)
_RE_WYTW   = re.compile(r'(?:wytw[oó]rca|manufacturer)\s*:\s*([^,.\n]+)', re.I)
# serial: "nr fabr.:" / "serial number(s):" albo bare "nr:" (nie "nr kat.", nie "nr wewn.")
_RE_FABR   = re.compile(
    r'(?:nr\s*fabr\.?|serial\s*numbers?)\s*:\s*(.+?)(?=,?\s*(?:' + _END + r')|$)', re.I)
_RE_NR     = re.compile(
    r'(?<!kat)(?<!wewn)\bnr\s*:\s*(.+?)(?=,?\s*(?:' + _END + r')|$)', re.I)
# podzial obiekt / czujnik pomiarowy (PL/EN, rozne sformulowania)
_RE_CZUJNIK_SPLIT = re.compile(
    r'\b(?:z\s+czujnikiem|oraz\s+czujnika|i\s+(?:zewn[eę]trznego\s+)?czujnik\w*'
    r'|with\s+\d+\b[^,]*?sensors?)\b', re.I)


def _wytnij_serial_liste(part):
    """Zwraca liste serial-i z czesci 'obiekt' (obsluguje kilka po przecinku)."""
    m = _RE_FABR.search(part) or _RE_NR.search(part)
    if not m:
        return []
    return [s.strip() for s in m.group(1).split(',') if s.strip()]


def _parsuj_pole(part):
    """Zwraca (typ, wytworca) z fragmentu tekstu."""
    mt = _RE_TYP.search(part)
    mw = _RE_WYTW.search(part)
    return (_oczysc(mt.group(1)) if mt else "",
            _oczysc(mw.group(1)) if mw else "")


# Znaki wypunktowania w PZ (pypdf zwraca rozne warianty, m.in. z fontu Symbol).
_RE_BULLET = re.compile(r'[•·●▪‣]\s*')


# Goly numer fabryczny na poczatku podpunktu — gdy etykieta 'nr fabr.:' stoi
# raz, w naglowku pozycji. Dopuszczamy cyfry, litery, myslnik i ukosnik.
_RE_SERIAL_NA_POCZATKU = re.compile(r'\s*([A-Za-z0-9][\w\-/]{2,})\s*(?:,|;|$)')


def _parsuj_bullet(fragment, typ, wytworca, nr_zlecenia, serial_z_naglowka=False):
    """
    Jeden wypunktowany podpunkt: '• nr fabr.: 40118669, nr ewid.: Q/LOG/36'
    (opcjonalnie z dopiskiem 'oraz czujnika temperatury typ: 0572 1001, nr ewid.: 8669').
    Typ i wytworca dziedziczy z naglowka pozycji. Zwraca liste PZPrzyrzad.

    `serial_z_naglowka=True` obsluguje wariant PZ dla wielu przyrzadow na tych
    samych punktach, gdzie etykieta stoi RAZ, w naglowku pozycji:

        Termohigrometr (rejestrator, 9 szt.) typ: testo 174H, nr fabr.:
          • 83623973, nr wew.: UR00045;
          • 83617608, nr wew.: UR00052;

    Podpunkt zaczyna sie wtedy wprost od numeru fabrycznego, bez etykiety.
    """
    czesci = _RE_CZUJNIK_SPLIT.split(fragment, maxsplit=1)
    obiekt  = czesci[0]
    czujnik = czesci[1] if len(czesci) > 1 else ""

    serials = _wytnij_serial_liste(obiekt)
    if not serials and serial_z_naglowka:
        m = _RE_SERIAL_NA_POCZATKU.match(obiekt)
        if m:
            serials = [m.group(1)]
    mw = _RE_WEWN.search(obiekt)
    nr_ewid = _oczysc(mw.group(1)) if mw else ""

    czuj_typ, czuj_wytworca = _parsuj_pole(czujnik) if czujnik else ("", "")
    czuj_serials = _wytnij_serial_liste(czujnik) if czujnik else []
    czuj_serial = _oczysc(czuj_serials[0]) if czuj_serials else ""

    return [PZPrzyrzad(
        nr_zlecenia=nr_zlecenia, wytworca=wytworca, typ=typ,
        nr_fabr=_oczysc(s), nr_ewid=nr_ewid,
        czuj_wytworca=czuj_wytworca, czuj_typ=czuj_typ, czuj_nr_fabr=czuj_serial,
    ) for s in (serials or [""])]


def _parsuj_wpis(wpis, nr_zlecenia):
    """
    Parsuje jeden wpis 'Obiekty wzorcowania' -> lista PZPrzyrzad.

    Dwa uklady w PZ:
      A) WYPUNKTOWANY (wiele przyrzadow jednego typu):
           'Termometr ... (rejestrator, 7 szt.) typ: testo 175T2,
              • nr fabr.: 40118669, nr ewid.: Q/LOG/36,
              • nr fabr.: 40118614, nr ewid.: Q/LOG/37, ...
            wytworca: Testo.'
         -> po jednym rekordzie na wypunktowanie (typ/wytworca z naglowka).
      B) JEDNOLINIOWY: 'typ: M1, nr fabr.: TMM160500502, nr ewid.: Q/LOG/19, wytworca: Tempmate.'
         (obsluguje tez kilka seriali po przecinku oraz opcjonalny czujnik pomiarowy).
    """
    bullety = _RE_BULLET.split(wpis)
    naglowek = bullety[0]
    podpunkty = [b for b in bullety[1:] if b.strip()]

    typ, wytworca = _parsuj_pole(naglowek)
    if not wytworca:
        # W ukladzie wypunktowanym wytworca stoi PO ostatnim podpunkcie — szukamy w calosci.
        mw = _RE_WYTW.search(wpis)
        wytworca = _oczysc(mw.group(1)) if mw else ""

    if podpunkty:
        # Etykieta 'nr fabr.:' bez wartosci na koncu naglowka = numery stoja
        # dopiero w podpunktach (PZ dla wielu przyrzadow na tych samych punktach).
        serial_z_naglowka = bool(
            re.search(r'(?:nr\s*fabr\.?|serial\s*numbers?)\s*:\s*$',
                      naglowek.strip(), re.I))
        out = []
        for frag in podpunkty:
            out.extend(_parsuj_bullet(frag, typ, wytworca, nr_zlecenia,
                                      serial_z_naglowka))
        return out

    # --- uklad B: jedna linia ---
    czesci = _RE_CZUJNIK_SPLIT.split(wpis, maxsplit=1)
    obiekt = czesci[0]
    czujnik = czesci[1] if len(czesci) > 1 else ""

    mw = _RE_WEWN.search(obiekt)
    nr_ewid = _oczysc(mw.group(1)) if mw else ""
    serials = _wytnij_serial_liste(obiekt)

    czuj_typ, czuj_wytworca = _parsuj_pole(czujnik) if czujnik else ("", "")
    czuj_serials = _wytnij_serial_liste(czujnik) if czujnik else []
    czuj_serial = _oczysc(czuj_serials[0]) if czuj_serials else ""

    # Gdy wytworca podany raz na koncu (wspolny dla obiektu i czujnika) — trafia do
    # czesci 'czujnik'; obiekt zostaje bez wytworcy. Dziedziczymy go wtedy do obiektu.
    if not wytworca and czuj_wytworca:
        wytworca = czuj_wytworca

    return [PZPrzyrzad(
        nr_zlecenia=nr_zlecenia, wytworca=wytworca, typ=typ,
        nr_fabr=_oczysc(s), nr_ewid=nr_ewid,
        czuj_wytworca=czuj_wytworca, czuj_typ=czuj_typ, czuj_nr_fabr=czuj_serial,
    ) for s in (serials or [""])]


# Pole UZYTKOWNIK / USER (opcjonalne) — blok adresowy do '[użytkownik]' w Word.
# Naglowek zawsze wielkimi literami (jak APPLICANT/ZLECENIODAWCA) — bez re.I, by nie
# lapac slowa 'user' w zdaniach.
_RE_UZYT = re.compile(
    r'(?:UŻYTKOWNIK|UZYTKOWNIK|USER)\s*:(.*?)'
    r'(?:Uwaga\s*:|Note\s*:|Opis\s+us[łl]ugi|Service\s+description|$)', re.S)


# ZLECENIODAWCA / APPLICANT — blok adresowy do '[zleceniodawca]' w Word.
# Bierzemy tekst miedzy naglowkiem a 'Osoba odpowiedzialna' (dalej ida dane kontaktowe
# osoby, telefon, e-mail, NIP — te NIE naleza do adresu zleceniodawcy).
_RE_ZLEC = re.compile(
    r'(?:ZLECENIODAWCA|APPLICANT|ORDERING\s+PARTY)\s*:(.*?)'
    r'(?:Osoba\s+odpowiedzialna|Contact\s+person|tel\.?\s*:|e-?mail\s*:|NIP\s*:'
    r'|U[ŻZ]YTKOWNIK\s*:|USER\s*:|$)', re.S | re.I)


def _parsuj_zleceniodawce(text):
    """
    Zwraca blok adresowy ZLECENIODAWCY (nazwa + ulica + kod/miasto) jako tekst
    wieloliniowy, albo ''. Puste linie z PDF sa pomijane, kolejnosc zachowana:
        'DANLAB Danuta Katryńska\\nul. Handlowa 6D\\n15-399 Białystok'
    """
    m = _RE_ZLEC.search(text)
    if not m:
        return ""
    linie = [ln.strip() for ln in m.group(1).splitlines() if ln.strip()]
    return "\n".join(linie)


def _parsuj_uzytkownikow_wg_pozycji(text):
    """
    Mapa: numer pozycji przyrzadu -> blok adresowy JEGO uzytkownika.

    Pole 'UŻYTKOWNIK:' bywa lista numerowana, w ktorej numer odpowiada POZYCJI przyrzadu
    w 'Obiekty wzorcowania' (przyrzad z pozycji 5 -> uzytkownik '5) GBA Polska ...').
    Do swiadectwa musi trafic TYLKO jego uzytkownik — wczesniej wpisywana byla cala lista.
    Numer z nawiasem jest usuwany. Gdy listy nie ma (jeden uzytkownik) -> {None: blok}.
    """
    m = _RE_UZYT.search(text)
    if not m:
        return {}
    linie = [ln.strip() for ln in m.group(1).splitlines() if ln.strip()]
    if not linie:
        return {}
    if not any(re.match(r'^\d+\)', ln) for ln in linie):
        return {None: "\n".join(linie)}          # jeden uzytkownik, bez numeracji

    mapa, biezacy = {}, None
    for ln in linie:
        m2 = re.match(r'^(\d+)\)\s*(.*)$', ln)
        if m2:
            biezacy = int(m2.group(1))
            reszta = m2.group(2).strip()
            mapa[biezacy] = [reszta] if reszta else []
        elif biezacy is not None:
            mapa[biezacy].append(ln)
    return {nr: "\n".join(w) for nr, w in mapa.items() if w}


def _parsuj_uzytkownik(text):
    """Zwraca blok adresowy uzytkownika (bez pustych linii i notatki 'Uwaga'), albo ''."""
    m = _RE_UZYT.search(text)
    if not m:
        return ""
    linie = [ln.strip() for ln in m.group(1).splitlines() if ln.strip()]
    return "\n".join(linie)


def _numer_zlecenia_th(text):
    """Zwraca numer zlecenia (samo '119'); przy wielu zleceniach preferuje /LA/TH/."""
    # PZ bywa z jednym albo kilkoma zleceniami: 'Numer zlecenia laboratorium: 119/LA/TH/2026'
    # oraz 'Numery zleceń laboratorium: 182/LA/TH/2026; 425/LA/TH/2026'.
    m = re.search(r'(?:Numer(?:y)?\s+zlece(?:nia|ń|n)\s+laboratorium|Order\s+numbers?)'
                  r'\s*:\s*([^\n]+)', text, re.I)
    if not m:
        return ""
    linia = m.group(1)
    th = re.search(r'(\d+)\s*/\s*LA\s*/\s*TH', linia, re.I)
    if th:
        return th.group(1)
    any_nr = re.search(r'(\d+)\s*/', linia)
    return any_nr.group(1) if any_nr else _oczysc(linia)


_RE_METODA = re.compile(
    r'(?:Metoda\s+wzorcowania|Metody\s+wzorcowania|Calibration\s+methods?)\s*:(.*?)'
    r'(?:Zakres\s+wzorcowania|Calibration\s+range|Uzupe[łl]nia\s+zleceniodawca|$)',
    re.I | re.S)


def _numery_pozycji(glowa):
    """Numery pozycji z naglowka wiersza metody: '1), 8), 9)' oraz zakresy '2) ÷ 7)'."""
    nums = set()
    for a, b in re.findall(r'(\d+)\s*\)\s*[÷\-–—]\s*(\d+)\s*\)', glowa):
        nums.update(range(int(a), int(b) + 1))
    for n in re.findall(r'(\d+)\s*\)', glowa):
        nums.add(int(n))
    return nums


def _pozycje_komory(text):
    """
    Numery pozycji wzorcowanych w KOMORZE KLIMATYCZNEJ (nasz protokol CC/CC-04) —
    z sekcji 'Metoda wzorcowania', np.:
        1), 8), 9) Metoda porownawcza ... w termostacie cieczowym ... ILAJ 5.4/3 ...
        2) ÷ 7)    Metoda porownawcza w komorze klimatycznej ... ILAJ 5.4/11 ...
    -> {2,3,4,5,6,7}.
    Zwraca None, gdy sekcji nie ma albo nie rozpoznano zadnej pozycji (wtedy nie filtrujemy).
    """
    m = _RE_METODA.search(text)
    if not m:
        return None
    sekcja = m.group(1)
    pozycje = set()
    # Kazdy wiersz metody zaczyna sie od numerow pozycji, potem opis ('Metoda porownawcza ...').
    for frag in re.split(r'(?m)^(?=\s*\d+\s*\))', sekcja):
        if not frag.strip():
            continue
        if not re.search(r'komor|5\.4/11|termohigrometr', frag, re.I):
            continue           # inna metoda (np. termostat cieczowy) — pomijamy
        glowa = re.split(r'metod', frag, maxsplit=1, flags=re.I)[0] or frag[:80]
        pozycje |= _numery_pozycji(glowa)
    return pozycje or None


_RE_ZAKRES = re.compile(
    r'(?:Zakres\s+wzorcowania|Calibration\s+range)\s*:(.*?)'
    r'(?:Termin\s+wykonania|Dokument\s+us[łl]ugi|Koszt\s+us[łl]ugi|Uzupe[łl]nia|$)',
    re.I | re.S)
# '(25 °C, 30 %rh)'  — punkt temperatura + wilgotnosc
_RE_PUNKT_TRH = re.compile(
    r'\(\s*(-?[\d.,]+)\s*°?\s*C\s*[;,]\s*(-?[\d.,]+)\s*%\s*rh\s*\)', re.I)
# '(0; 10; 20; 30) °C'  — lista punktow samej temperatury
_RE_PUNKT_T   = re.compile(r'\(\s*(-?[\d.,]+(?:\s*;\s*-?[\d.,]+)+)\s*\)\s*°?\s*C', re.I)


def _punkty_z_fragmentu(frag):
    """
    Punkty (T, RH|None) z jednego fragmentu sekcji 'Zakres wzorcowania'.

    Fragment potrafi laczyc OBA warianty zapisu, np.:

        (-20; 0; 40) °C, (25 °C, 30 %rh); (25 °C, 60 %rh); (25 °C, 85 %rh)

    czyli trzy punkty samej temperatury PLUS punkty z wilgotnoscia. Wczesniej,
    gdy trafil sie choc jeden punkt z wilgotnoscia, funkcja konczyla prace i
    lista '(-20; 0; 40) °C' przepadala — a poniewaz punkty do protokolu wybiera
    sie wg PZ, te trzy po prostu znikaly z protokolu.

    Zbieramy oba warianty i ustawiamy je w KOLEJNOSCI WYSTAPIENIA w tekscie,
    zeby punkty w protokole szly tak jak w zamowieniu.
    """
    znalezione = []   # (pozycja_w_tekscie, indeks_w_grupie, (T, RH))

    for m in _RE_PUNKT_TRH.finditer(frag):
        t, rh = _do_float(m.group(1)), _do_float(m.group(2))
        if t is not None:
            znalezione.append((m.start(), 0, (t, rh)))

    # '(0; 10; 20) °C' — jedna lista to kilka punktow; moze wystapic wiele razy.
    for m in _RE_PUNKT_T.finditer(frag):
        for i, kawalek in enumerate(m.group(1).split(';')):
            t = _do_float(kawalek)
            if t is not None:
                znalezione.append((m.start(), i, (t, None)))

    znalezione.sort(key=lambda x: (x[0], x[1]))
    return [punkt for _poz, _i, punkt in znalezione]


def _scal_punkty(grupy):
    """
    Laczy punkty z wielu POZYCJI (przyrzadow) i wielu PZ w jedna liste punktow komory.

    Ten sam punkt zamowiony przez kilka przyrzadow to JEDEN wsad komory — nie powielamy go.
    Ale powtorzenie WEWNATRZ jednej pozycji jest znaczace (np. drugi raz 50 %rh na
    histereze), wiec dla kazdej wartosci zostawiamy tyle wystapien, ile MAKSYMALNIE
    zamowiono w pojedynczej pozycji. Kolejnosc = pierwsze wystapienie.
    """
    maks, kolejnosc = {}, []
    for grupa in grupy:
        licznik = {}
        for p in grupa:
            licznik[p] = licznik.get(p, 0) + 1
        for p, n in licznik.items():
            if p not in maks:
                kolejnosc.append(p)
            maks[p] = max(maks.get(p, 0), n)
    out = []
    for p in kolejnosc:
        out.extend([p] * maks[p])
    return out


def punkty_wzorcowania(text, pozycje=None):
    """
    Punkty pomiarowe ZAMOWIONE w PZ (sekcja 'Zakres wzorcowania') — lista krotek
    (temperatura, wilgotnosc|None) W KOLEJNOSCI z PZ, razem z powtorzeniami.

    `pozycje` — zbior numerow pozycji, ktore nas interesuja (np. wzorcowane w komorze
    klimatycznej). Sekcja bywa rozbita per pozycja i dotyczy roznych wielkosci, np.:
        1) (25 °C, 30 %rh); (25 °C, 50 %rh); ...        <- termohigrometr (komora)
        2) (950 ÷ 1050) hPa ...                          <- barometr (inne stanowisko)
    Bez filtrowania wzielibysmy punkty obu. Gdy `pozycje` = None albo w sekcji nie ma
    numeracji — bierzemy calosc.

    Obslugiwane zapisy punktow:
      '(25 °C, 30 %rh); (25 °C, 50 %rh); (25 °C, 70 %rh); (25 °C, 50 %rh) 1)'
         -> [(25,30), (25,50), (25,70), (25,50)]   (ostatni to powtorka na histereze)
      '2) ÷ 7) (0; 10; 20; 30) °C'
         -> [(0,None), (10,None), (20,None), (30,None)]
    """
    return [p for grupa in punkty_wzorcowania_grupy(text, pozycje) for p in grupa]


def punkty_wzorcowania_grupy(text, pozycje=None):
    """
    To samo co punkty_wzorcowania, ale z podzialem na POZYCJE (przyrzady) — jedna grupa
    na pozycje. Kazda pozycja PZ ma zwykle WLASNY zestaw punktow, np.:
        1) (29; 30; 31) °C     <- przyrzad 1
        2) (36; 38) °C         <- przyrzad 2
        3) (2; 8) °C           <- przyrzady 3
    Podzial jest potrzebny, by odroznic powtorzenie punktu w JEDNEJ pozycji (histereza)
    od tego samego punktu zamowionego przez ROZNE przyrzady (jeden wsad komory).
    """
    m = _RE_ZAKRES.search(text)
    if not m:
        return []
    sekcja = m.group(1)

    fragmenty = [f for f in re.split(r'(?m)^(?=\s*\d+\s*\))', sekcja) if f.strip()]
    if len(fragmenty) <= 1:
        punkty = _punkty_z_fragmentu(sekcja)   # brak numeracji pozycji — calosc jako jedna grupa
        return [punkty] if punkty else []

    grupy = []
    for frag in fragmenty:
        glowa = frag.split('(', 1)[0]         # numery stoja PRZED pierwszym nawiasem
        nums = _numery_pozycji(glowa)
        if pozycje and nums and not (nums & set(pozycje)):
            continue                           # fragment innej pozycji (np. cisnienie)
        punkty = _punkty_z_fragmentu(frag)
        if punkty:
            grupy.append(punkty)
    return grupy


def punkty_wg_pozycji(text, pozycje=None):
    """
    Mapa: numer pozycji PZ -> lista punktow zamowionych DLA TEJ pozycji (przyrzadu).

    W PZ kazdy przyrzad (pozycja) ma zwykle wlasny zestaw punktow:
        1) (29; 30; 31) °C     <- pozycja 1
        3) (2; 8) °C           <- pozycja 3
    Gdy sekcja nie ma numeracji (jeden przyrzad), zwracamy {None: [punkty]}.
    """
    m = _RE_ZAKRES.search(text)
    if not m:
        return {}
    sekcja = m.group(1)
    fragmenty = [f for f in re.split(r'(?m)^(?=\s*\d+\s*\))', sekcja) if f.strip()]
    if len(fragmenty) <= 1:
        punkty = _punkty_z_fragmentu(sekcja)
        return {None: punkty} if punkty else {}

    mapa = {}
    for frag in fragmenty:
        glowa = frag.split('(', 1)[0]
        nums = _numery_pozycji(glowa)
        if pozycje and nums and not (nums & set(pozycje)):
            continue
        punkty = _punkty_z_fragmentu(frag)
        if not punkty:
            continue
        for nr in (nums or [None]):
            mapa.setdefault(nr, []).extend(punkty)
    return mapa


def wczytaj_punkty_przyrzadow(folder_pz):
    """
    Mapa: nr fabryczny/ewidencyjny przyrzadu -> punkty ZAMOWIONE wlasnie dla niego.

    Jeden wsad komory obsluguje kilka zlecen i kilka przyrzadow, a kazdy przyrzad bywa
    zamowiony na INNE punkty. Protokol zawiera sume punktow, wiec odczyty przyrzadu w
    punktach spoza JEGO zamowienia sa zbedne — dzieki tej mapie mozna je wyszarzyc.
    """
    wynik = {}
    if not _PDF_OK or not os.path.isdir(folder_pz):
        return wynik
    for path in sorted(glob.glob(os.path.join(folder_pz, "*.pdf"))):
        try:
            reader = PdfReader(path)
            text = "\n".join((p.extract_text() or "") for p in reader.pages)
            komora = _pozycje_komory(text)
            mapa_poz = punkty_wg_pozycji(text, komora)
            przyrzady = parsuj_pdf(path)
        except Exception as e:
            print(f"  [PZ] Blad odczytu punktow przyrzadow z '{os.path.basename(path)}': {e}")
            continue
        if not mapa_poz:
            continue
        # Gdy sekcja punktow nie ma numeracji — ten sam zestaw dotyczy wszystkich pozycji.
        wspolne = mapa_poz.get(None)
        for p in przyrzady:
            if not getattr(p, 'komora', True):
                continue
            punkty = mapa_poz.get(p.pozycja) or wspolne
            if not punkty:
                continue
            for klucz in (normalizuj_serial(p.nr_fabr), normalizuj_serial(p.nr_ewid)):
                if klucz:
                    wynik.setdefault(klucz, list(punkty))
    return wynik


def _opis_punktow(punkty):
    return ", ".join(f"{t:g}C/{rh:g}%" if rh is not None else f"{t:g}C" for t, rh in punkty)


def wczytaj_punkty(folder_pz):
    """
    Punkty zamowione we WSZYSTKICH PZ z folderu i we wszystkich ich pozycjach.

    Jeden wsad komory obsluguje zwykle kilka zlecen, a w kazdym zleceniu kazdy przyrzad
    moze miec inny zestaw punktow. Do obserwacji potrzebna jest ICH SUMA — bez powielania
    punktow zamowionych przez kilka przyrzadow (to jeden i ten sam wsad komory).
    Zwraca liste (temperatura, wilgotnosc|None). Cicha gdy brak folderu/pypdf.
    """
    if not _PDF_OK or not os.path.isdir(folder_pz):
        return []
    grupy = []
    for path in sorted(glob.glob(os.path.join(folder_pz, "*.pdf"))):
        try:
            reader = PdfReader(path)
            text = "\n".join((p.extract_text() or "") for p in reader.pages)
            # Punkty TYLKO z pozycji wzorcowanych w komorze klimatycznej — ten sam
            # przyrzad bywa w PZ takze na innym stanowisku (np. barometr, hPa).
            g = punkty_wzorcowania_grupy(text, _pozycje_komory(text))
        except Exception as e:
            print(f"  [PZ] Blad odczytu punktow z '{os.path.basename(path)}': {e}")
            continue
        if g:
            plaskie = [p for grupa in g for p in grupa]
            print(f"  [PZ] Punkty zamowione ({os.path.basename(path)}, "
                  f"{len(g)} poz.): {_opis_punktow(plaskie)}")
            grupy.extend(g)

    punkty = _scal_punkty(grupy)
    if punkty:
        print(f"  [PZ] Razem punktow do obserwacji ({len(punkty)}): {_opis_punktow(punkty)}")
    return punkty


def parsuj_pdf(path):
    """Parsuje jeden PDF PZ -> lista PZPrzyrzad."""
    if not _PDF_OK:
        return []
    reader = PdfReader(path)
    text = "\n".join((p.extract_text() or "") for p in reader.pages)
    nr_zlec = _numer_zlecenia_th(text)
    uzyt = _parsuj_uzytkownik(text)
    uzyt_wg_poz = _parsuj_uzytkownikow_wg_pozycji(text)
    zlec_adres = _parsuj_zleceniodawce(text)

    m = re.search(
        r'(?:Obiekty\s+wzorcowania|Calibration\s+objects)\s*:(.*?)'
        r'(?:(?:Metoda\s+wzorcowania|Calibration\s+methods?)\s*:|$)', text, re.I | re.S)
    if not m:
        return []
    sekcja = re.sub(r'\s+', ' ', m.group(1)).strip()

    # Pozycje wzorcowane w komorze klimatycznej (nasz protokol) — z sekcji 'Metoda wzorcowania'.
    komora_poz = _pozycje_komory(text)

    # podzial na wpisy numerowane '1) ... 2) ...'; brak numeracji => jeden wpis
    wpisy = re.split(r'(?<!\d)([1-9])\)\s', sekcja)
    przyrzady = []
    if len(wpisy) > 1:
        # re.split z grupa: [prefix, '1', tekst1, '2', tekst2, ...]
        for i in range(1, len(wpisy), 2):
            nr_poz = int(wpisy[i])
            tekst  = wpisy[i + 1] if i + 1 < len(wpisy) else ""
            for p in _parsuj_wpis(tekst, nr_zlec):
                p.pozycja = nr_poz
                # brak sekcji metody => nie filtrujemy (kazda pozycja traktowana jak nasza)
                p.komora  = True if komora_poz is None else (nr_poz in komora_poz)
                przyrzady.append(p)
    else:
        for p in _parsuj_wpis(sekcja, nr_zlec):
            p.komora = True
            przyrzady.append(p)
    for p in przyrzady:
        # Uzytkownik JEGO pozycji; gdy lista nie jest numerowana — wspolny dla wszystkich.
        p.uzytkownik = (uzyt_wg_poz.get(p.pozycja)
                        or uzyt_wg_poz.get(None)
                        or (uzyt if not uzyt_wg_poz else ""))
        p.zleceniodawca = zlec_adres
    return przyrzady


def wczytaj_pz(folder_pz):
    """
    Wczytuje wszystkie PDFy z folder_pz -> dict[normalizuj_serial(nr_fabr) -> PZPrzyrzad].
    Zwraca (mapa, lista_wszystkich). Cichy gdy brak folderu/plikow.
    """
    mapa, lista = {}, []
    if not _PDF_OK:
        print("  [PZ] Brak biblioteki 'pypdf' — pomijam dane z PZ.")
        return mapa, lista
    if not os.path.isdir(folder_pz):
        print(f"  [PZ] Brak folderu '{folder_pz}' — pomijam dane z PZ.")
        return mapa, lista

    pliki = sorted(glob.glob(os.path.join(folder_pz, "*.pdf")))
    for path in pliki:
        try:
            przyrzady = parsuj_pdf(path)
        except Exception as e:
            print(f"  [PZ] Blad odczytu '{os.path.basename(path)}': {e}")
            continue
        lista.extend(przyrzady)

    # Klucz po nr fabrycznym I nr ewidencyjnym — pliki logerow bywaja nazwane jednym albo
    # drugim (np. Vaisala: plik 'EV363499' = nr ewidencyjny, a nr fabryczny 'D4940027').
    # KOLEJNOSC MA ZNACZENIE: najpierw pozycje z KOMORY KLIMATYCZNEJ (nasz protokol),
    # potem pozostale. Ten sam przyrzad bywa w PZ dwa razy — raz jako wzorcowany w komorze
    # (czujnik wewnetrzny, kanal 1), raz w termostacie z czujnikiem zewnetrznym (0572 1001).
    # Bez tego do protokolu trafialy dane czujnika zewnetrznego.
    for tylko_komora in (True, False):
        for p in lista:
            if bool(p.komora) is not tylko_komora:
                continue
            for klucz in (normalizuj_serial(p.nr_fabr), normalizuj_serial(p.nr_ewid)):
                if klucz:
                    mapa.setdefault(klucz, p)

    n_kom = sum(1 for p in lista if p.komora)
    print(f"  [PZ] Wczytano {len(pliki)} PDF, {len(lista)} przyrzadow "
          f"({n_kom} z komory klimatycznej, {len(mapa)} kluczy).")
    return mapa, lista


# =============================================================================
# ZESTAWIENIE (rozdzielczosc t / RH)
# =============================================================================

_RE_RES_T  = re.compile(r't\s*:?\s*([\d.,]+)', re.I)
_RE_RES_RH = re.compile(r'(?:rh|wilg)\s*:?\s*([\d.,]+)', re.I)


def _do_float(s):
    try:
        return float(str(s).replace(',', '.'))
    except (ValueError, TypeError):
        return None


def wczytaj_zestawienie(sciezka, arkusz="Termohigrometry"):
    """
    Wczytuje liste wpisow: [(producent_norm, typ_norm, t_res, rh_res), ...].
    Producent w kol. B, Typ w kol. D (czesto wieloliniowy, np. 'wskaznik:\\ntesto 625\\n...'),
    Rozdzielczosc w kol. J (np. 't: 0,1 °C\\nRH: 0,1 %'). Cichy gdy brak pliku.
    Zwraca liste (nie dict) — dopasowanie jest rozmyte (podciag), patrz rozdzielczosc_zestawienie.
    """
    wpisy = []
    if not _XLSX_OK or not os.path.exists(sciezka):
        if not os.path.exists(sciezka):
            print(f"  [PZ] Brak pliku Zestawienia '{os.path.basename(sciezka)}' — "
                  f"rozdzielczosc tylko z danych.")
        return wpisy
    try:
        wb = openpyxl.load_workbook(sciezka, data_only=True, read_only=True)
    except Exception as e:
        print(f"  [PZ] Blad otwarcia Zestawienia: {e}")
        return wpisy
    if arkusz not in wb.sheetnames:
        print(f"  [PZ] Brak arkusza '{arkusz}' w Zestawieniu.")
        return wpisy
    ws = wb[arkusz]
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=11, values_only=True):
        producent, typ, rozdz = row[1], row[3], row[9]   # B, D, J
        if not (producent and typ and rozdz):
            continue
        rozdz = str(rozdz)
        mt = _RE_RES_T.search(rozdz)
        mrh = _RE_RES_RH.search(rozdz)
        t_res = _do_float(mt.group(1)) if mt else None
        rh_res = _do_float(mrh.group(1)) if mrh else None
        wpisy.append((_norm_txt(producent), _norm_txt(typ), t_res, rh_res))
    wb.close()
    print(f"  [PZ] Zestawienie: {len(wpisy)} przyrzadow z rozdzielczoscia.")
    return wpisy


def rozdzielczosc_zestawienie(zest, producent, typ):
    """
    Rozmyte dopasowanie (producent, typ) do Zestawienia. Kolumna Typ bywa wieloliniowa
    ('wskaznik:\\ntesto 625\\nczujnik:\\n...'), dlatego typ przyrzadu szukamy jako PODCIAG.
    Zwraca (t_res, rh_res) albo (None, None).
    """
    pn, tn = _norm_txt(producent), _norm_txt(typ)
    if not tn:
        return (None, None)
    for zp, zt, t_res, rh_res in zest:
        prod_ok = bool(pn) and (pn == zp or pn in zp or zp in pn)
        typ_ok = tn in zt or zt in tn
        if prod_ok and typ_ok:
            return (t_res, rh_res)
    return (None, None)


# =============================================================================
# ROZDZIELCZOSC Z DANYCH (fallback)
# =============================================================================

def rozdzielczosc_z_kolumny(values):
    """
    Rozdzielczosc = najmniejsze miejsce dziesietne, ktore sie zmienia w kolumnie:
      same calkowite (40.00...) -> 1;  zmiana na 1. miejscu (40.1) -> 0.1;
      na 2. (40.04) -> 0.01.  None gdy brak danych liczbowych.
    """
    vals = [v for v in values if isinstance(v, (int, float))]
    if not vals:
        return None
    max_dec = 0
    for v in vals:
        s = f"{float(v):.4f}".rstrip('0')
        dec = len(s.split('.')[1]) if '.' in s else 0
        if dec > max_dec:
            max_dec = dec
    return {0: 1.0, 1: 0.1, 2: 0.01, 3: 0.001}.get(max_dec, 10 ** (-max_dec))
